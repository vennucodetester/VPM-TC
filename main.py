import sys
import datetime
import calendar
import json
import xml.etree.ElementTree as ET
from enum import Enum
from dateutil.relativedelta import relativedelta
from functools import partial
import os
import importlib.util
import csv
import pandas as pd

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QToolBar, QLabel, QFrame, QPushButton, QProgressBar, QMessageBox,
    QFileDialog, QDialog, QFormLayout, QLineEdit, QDateEdit, QDialogButtonBox, QSizePolicy,
    QMenu, QTextEdit, QCheckBox, QComboBox, QTabWidget, QTreeWidget, QTreeWidgetItem, QListWidget, QListWidgetItem, QScrollArea,
    QTableWidget, QTableWidgetItem, QHeaderView, QRadioButton, QButtonGroup, QGroupBox,
    QSplitter, QTabBar, QStackedWidget, QInputDialog,
    QGraphicsView, QGraphicsScene, QGraphicsItem, QGraphicsTextItem, QGraphicsRectItem,
    QGraphicsProxyWidget, QPlainTextEdit, QAbstractItemView,
    QToolButton, QFontComboBox, QColorDialog
)
from PyQt6.QtGui import (
    QAction, QFont, QFontMetrics, QPainter, QColor, QPen, QPolygonF, QGuiApplication, QBrush, QTextCursor, QKeyEvent, QTextOption, QFocusEvent, QCursor, QPalette,
    QTextCharFormat, QTextListFormat
)
from PyQt6.QtCore import Qt, QRectF, QPointF, pyqtSignal, QDate, QTimer, QSizeF, QObject, QEvent

import advanced_excel_importer as ecn_engine

# ============================================================================
# DEBUG LOGGING - CAN BE EASILY REMOVED
# ============================================================================
# TO DISABLE: Set DEBUG_NOTES = False
# TO REMOVE: Delete this entire section and search for "debug_log(" to remove all calls
# ============================================================================
DEBUG_NOTES = True  # Set to False to disable all debug logging

def debug_log(message, level="INFO"):
    """
    Debug logging function for tracking crashes and issues.

    Usage: debug_log("message", "INFO|ERROR")

    Logs include:
    - Timestamp (millisecond precision)
    - Level (INFO, ERROR)
    - Message
    - Full stack trace for ERRORs

    To remove all debug logging:
    1. Delete this entire DEBUG LOGGING section (lines 31-58)
    2. Search for 'debug_log(' and delete all debug_log calls
    3. You can use regex to find all calls
    """
    if DEBUG_NOTES:
        import traceback
        timestamp = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        print(f"[{timestamp}] [{level}] {message}")
        if level == "ERROR":
            print(f"    Stack: {traceback.format_exc()}")

# ============================================================================

class TimescaleView(Enum):
    WEEK = 1
    MONTH = 2
    YEAR = 3

class Priority(Enum):
    LOW = 1
    MEDIUM = 2
    HIGH = 3

# ============================================================================
# NOTES SYSTEM DATA MODEL
# ============================================================================

class Note:
    """Individual note with text, timestamp, and optional reminder."""
    def __init__(self, text, timestamp=None, reminder_date=None, x_pos=0, y_pos=0, width=200, height=100, heading="", note_type="canvas"):
        self.text = text
        self.heading = heading  # Optional title/heading for the note
        self.timestamp = timestamp or datetime.datetime.now()
        self.reminder_date = reminder_date
        self.is_completed = False
        self.uid = f"note_{id(self)}_{self.timestamp.timestamp()}"

        # Canvas position and size
        self.x_pos = x_pos
        self.y_pos = y_pos
        self.width = width
        self.height = height

        # Note type: "canvas" (plain text) or "grid" (Excel-like)
        self.note_type = note_type

        # Grid data (only for grid type notes)
        self.grid_data = {}  # {(row, col): value}

        # Grid position (for grid mode)
        self.row_index = None
        self.col_index = None

        # Status: "no_timeline", "active", "overdue", "completed"
        self.status = "no_timeline"

    def update_status(self):
        """Update status based on reminder_date and is_completed."""
        if not self.reminder_date:
            self.status = "no_timeline"
        elif self.is_completed:
            self.status = "completed"
        elif datetime.datetime.now() > self.reminder_date:
            self.status = "overdue"
        else:
            self.status = "active"
        return self.status

    def to_dict(self):
        # Convert grid_data tuple keys to strings for JSON serialization
        # Format: {(row, col): value} -> {"row,col": value}
        grid_data_serialized = {}
        if self.grid_data:
            for (row, col), value in self.grid_data.items():
                key = f"{row},{col}"
                grid_data_serialized[key] = value

        print(f"[DEBUG] Note.to_dict() uid={self.uid}, type={self.note_type}, grid_data entries={len(self.grid_data)}")
        if self.note_type == 'grid' and self.grid_data:
            print(f"[DEBUG]   Grid data sample: {list(self.grid_data.items())[:3]}")

        return {
            'uid': self.uid,
            'text': self.text,
            'heading': self.heading,
            'timestamp': self.timestamp.isoformat(),
            'reminder_date': self.reminder_date.isoformat() if self.reminder_date else None,
            'is_completed': self.is_completed,
            'x_pos': self.x_pos,
            'y_pos': self.y_pos,
            'width': self.width,
            'height': self.height,
            'note_type': self.note_type,
            'grid_data': grid_data_serialized,
            'row_index': self.row_index,
            'col_index': self.col_index,
            'status': self.status
        }

    @staticmethod
    def from_dict(data):
        note = Note(
            data['text'],
            datetime.datetime.fromisoformat(data['timestamp']),
            datetime.datetime.fromisoformat(data['reminder_date']) if data.get('reminder_date') else None,
            data.get('x_pos', 0),
            data.get('y_pos', 0),
            data.get('width', 200),
            data.get('height', 100),
            data.get('heading', ''),  # Default to empty string for backwards compatibility
            data.get('note_type', 'canvas')  # Default to canvas for backwards compatibility
        )
        note.uid = data.get('uid', note.uid)
        note.is_completed = data.get('is_completed', False)
        # Convert grid_data string keys back to tuples for internal use
        # Format: {"row,col": value} -> {(row, col): value}
        grid_data_raw = data.get('grid_data', {})
        note.grid_data = {}
        if grid_data_raw:
            for key, value in grid_data_raw.items():
                try:
                    # Try to parse as "row,col" format (new format)
                    if ',' in key:
                        row_str, col_str = key.split(',', 1)
                        row, col = int(row_str), int(col_str)
                        note.grid_data[(row, col)] = value
                    else:
                        # Legacy format: might be tuple already (shouldn't happen in JSON, but handle gracefully)
                        # If it's already a tuple from old data, keep it
                        if isinstance(key, tuple):
                            note.grid_data[key] = value
                except (ValueError, AttributeError):
                    # Skip invalid keys
                    continue
        note.row_index = data.get('row_index', None)
        note.col_index = data.get('col_index', None)
        note.status = data.get('status', 'no_timeline')

        print(f"[DEBUG] Note.from_dict() uid={note.uid}, type={note.note_type}, grid_data entries={len(note.grid_data)}")
        if note.note_type == 'grid' and note.grid_data:
            print(f"[DEBUG]   Grid data sample: {list(note.grid_data.items())[:3]}")

        return note


class NotesItem:
    """Item in the notes navigation structure (can be linked to Task or custom page)."""
    def __init__(self, name, uid=None, linked_task=None, is_custom=False, level=2):
        self.name = name
        self.uid = uid or f"notesitem_{id(self)}"
        self.linked_task = linked_task  # Reference to Task object if linked
        self.is_custom = is_custom  # True if manually created, False if auto-populated
        self.level = level  # 2=middle column, 3=top tabs
        # SIMPLE: Automatically copy notes from linked task
        self.notes = list(linked_task.notes) if linked_task and hasattr(linked_task, 'notes') else []
        self.children = []  # Child NotesItems (for hierarchy)
        self.parent = None
        self.view_mode = "canvas"  # "canvas" or "grid"
        self.grid_data = {}  # For grid mode: stores cell data as {row,col: value}

    def add_note(self, note):
        """Add a note to this item."""
        self.notes.append(note)

    def get_all_notes(self):
        """Get all notes including from children."""
        all_notes = list(self.notes)
        for child in self.children:
            all_notes.extend(child.get_all_notes())
        return all_notes

    def has_overdue_notes(self):
        """Check if any note in this item is overdue."""
        for note in self.notes:
            note.update_status()
            if note.status == "overdue":
                return True
        return False

    def has_overdue_children(self):
        """Check if any child has overdue notes (for cascading indicators)."""
        for child in self.children:
            if child.has_overdue_notes() or child.has_overdue_children():
                return True
        return False

    def to_dict(self):
        print(f"[DEBUG] NotesItem.to_dict() for '{self.name}': {len(self.notes)} notes, {len(self.children)} children, view_mode={self.view_mode}")
        if self.notes:
            print(f"[DEBUG]   First note: uid={self.notes[0].uid}, type={self.notes[0].note_type}")
        return {
            'uid': self.uid,
            'name': self.name,
            'linked_task_uid': self.linked_task.uid if self.linked_task else None,
            'is_custom': self.is_custom,
            'level': self.level,
            'view_mode': self.view_mode,
            'grid_data': {f"{row},{col}": value for (row, col), value in self.grid_data.items()} if self.grid_data else {},
            'notes': [note.to_dict() for note in self.notes],
            'children': [child.to_dict() for child in self.children]  # Recursively serialize children
        }

    @staticmethod
    def from_dict(data, task_map=None):
        """Reconstruct NotesItem from dict. task_map is {uid: Task} for linking."""
        linked_task = None
        if data.get('linked_task_uid') and task_map:
            linked_task = task_map.get(data['linked_task_uid'])

        item = NotesItem(
            data['name'],
            data['uid'],
            linked_task,
            data.get('is_custom', False),
            data.get('level', 2)
        )
        item.notes = [Note.from_dict(n) for n in data.get('notes', [])]
        item.view_mode = data.get('view_mode', 'canvas')
        # Convert grid_data string keys back to tuples for internal use (same as Note)
        grid_data_raw = data.get('grid_data', {})
        item.grid_data = {}
        if grid_data_raw:
            for key, value in grid_data_raw.items():
                try:
                    if ',' in key:
                        row_str, col_str = key.split(',', 1)
                        row, col = int(row_str), int(col_str)
                        item.grid_data[(row, col)] = value
                    elif isinstance(key, tuple):
                        item.grid_data[key] = value
                except (ValueError, AttributeError):
                    continue

        # Recursively deserialize children
        item.children = [NotesItem.from_dict(c, task_map) for c in data.get('children', [])]
        for child in item.children:
            child.parent = item

        return item


class NotesStructure:
    """Manages the overall notes organization for all projects."""
    def __init__(self):
        self.projects_notes = {}  # {project_uid: {'items': [NotesItem], 'inbox_notes': [Note]}}
        self.global_inbox = []  # Notes not tied to any project

    def get_project_structure(self, project_uid):
        """Get or create notes structure for a project."""
        if project_uid not in self.projects_notes:
            self.projects_notes[project_uid] = {'items': [], 'inbox_notes': []}
        return self.projects_notes[project_uid]

    def auto_populate_from_project(self, project):
        """Auto-populate notes structure from a Gantt project."""
        project_structure = self.get_project_structure(project.uid)

        # Helper function to recursively collect all existing task UIDs
        def collect_existing_uids(notes_item):
            """Recursively collect task UIDs from a NotesItem and its children."""
            uids = set()
            if notes_item.linked_task:
                uids.add(notes_item.linked_task.uid)
            for child in notes_item.children:
                uids.update(collect_existing_uids(child))
            return uids

        # Track UIDs of tasks we've already created items for (check both flat list and children)
        existing_uids = set()
        for item in project_structure['items']:
            existing_uids.update(collect_existing_uids(item))

        # Helper to find existing NotesItem by task UID
        def find_notes_item(task_uid):
            """Recursively find NotesItem with matching linked_task.uid"""
            def search_in_item(item):
                if item.linked_task and item.linked_task.uid == task_uid:
                    return item
                for child in item.children:
                    result = search_in_item(child)
                    if result:
                        return result
                return None

            for root_item in project_structure['items']:
                result = search_in_item(root_item)
                if result:
                    return result
            return None

        def add_task_as_item(task, level=2):
            # Check if item already exists
            if task.uid in existing_uids:
                # Item exists - find it and update its children references to current NotesItems
                existing_item = find_notes_item(task.uid)
                if existing_item and task.children:
                    # Build correct children list from current task.children
                    correct_children = []
                    for child_task in task.children:
                        child_item = find_notes_item(child_task.uid)
                        if child_item:
                            correct_children.append(child_item)
                            child_item.parent = existing_item
                    # Replace children array with correct references
                    existing_item.children = correct_children
                return

            # Create NotesItem for this task
            item = NotesItem(task.name, f"notes_{task.uid}", task, False, level)
            project_structure['items'].append(item)
            existing_uids.add(task.uid)

            # Add children as sub-items
            for child in task.children:
                if child.uid not in existing_uids:
                    # Create new child
                    child_item = NotesItem(child.name, f"notes_{child.uid}", child, False, level + 1)
                    item.children.append(child_item)
                    child_item.parent = item
                    existing_uids.add(child.uid)
                else:
                    # Reuse existing child
                    child_item = find_notes_item(child.uid)
                    if child_item and child_item not in item.children:
                        item.children.append(child_item)
                        child_item.parent = item

        # Add all root tasks
        for task in project.root_tasks:
            add_task_as_item(task, level=2)

    def add_task_manually(self, task, project_uid):
        """
        Manually add any task/item to notes structure.
        Works for: Task, ProblemReport, ECN, Item, Project, etc.
        """
        project_structure = self.get_project_structure(project_uid)

        # Check if already exists
        existing_uids = {item.linked_task.uid for item in project_structure['items'] if item.linked_task}
        if task.uid in existing_uids:
            return False  # Already exists

        # Create NotesItem for this task
        item = NotesItem(task.name, f"notes_{task.uid}", task, False, level=2)

        # Get children - works for Task, ProblemReport, ECN, etc.
        children = []
        if hasattr(task, 'children'):
            children = task.children
        elif hasattr(task, 'items'):  # For ECN objects
            children = task.items
        elif hasattr(task, 'ecns'):  # For ProblemReport
            children = list(task.ecns.values()) if task.ecns else []
        elif hasattr(task, 'root_tasks'):  # For Project
            children = task.root_tasks

        # Add children as sub-items (level 3)
        for child in children:
            child_item = NotesItem(child.name, f"notes_{child.uid}", child, False, level=3)
            item.children.append(child_item)
            child_item.parent = item

        project_structure['items'].append(item)
        return True  # Successfully added

    def to_dict(self):
        """Serialize to dictionary for JSON storage."""
        return {
            'projects_notes': {
                proj_uid: {
                    'items': [item.to_dict() for item in struct['items']],
                    'inbox_notes': [note.to_dict() for note in struct['inbox_notes']]
                }
                for proj_uid, struct in self.projects_notes.items()
            },
            'global_inbox': [note.to_dict() for note in self.global_inbox]
        }

    @staticmethod
    def from_dict(data, task_map=None):
        """Reconstruct from dictionary."""
        structure = NotesStructure()
        structure.global_inbox = [Note.from_dict(n) for n in data.get('global_inbox', [])]

        for proj_uid, proj_data in data.get('projects_notes', {}).items():
            structure.projects_notes[proj_uid] = {
                'items': [NotesItem.from_dict(i, task_map) for i in proj_data.get('items', [])],
                'inbox_notes': [Note.from_dict(n) for n in proj_data.get('inbox_notes', [])]
            }

        # Rebuild parent-child relationships (for backward compatibility with old config files)
        # New config files use 'children' field which is handled in NotesItem.from_dict()
        for proj_uid, struct in structure.projects_notes.items():
            item_map = {item.uid: item for item in struct['items']}
            for item_data in data.get('projects_notes', {}).get(proj_uid, {}).get('items', []):
                parent_item = item_map.get(item_data['uid'])
                if parent_item and 'children_uids' in item_data:
                    # Old format compatibility: children_uids exists but children data doesn't
                    for child_uid in item_data.get('children_uids', []):
                        child = item_map.get(child_uid)
                        if child and child not in parent_item.children:
                            parent_item.children.append(child)
                            child.parent = parent_item

        return structure


# ============================================================================
# NATURAL LANGUAGE DATE PARSER
# ============================================================================

import re

def parse_natural_date(text):
    """
    Parse natural language dates from text with comprehensive format support.
    Recognizes absolute times, relative times, and date phrases in many formats.
    Returns: (datetime object, cleaned_text) or (None, text) if no date found
    """
    try:
        text_lower = text.lower().strip()
        now = datetime.datetime.now()

        # Month names mapping
        month_names = {
            'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'august': 8, 'sep': 9, 'september': 9,
            'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
        }

        # ==== ISO DATE FORMATS ====
        # ISO 8601: YYYY-MM-DD or YYYY-MM-DDTHH:MM or YYYY-MM-DD HH:MM
        match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:[T\s](\d{1,2}):(\d{2})(?::(\d{2}))?)?', text_lower)
        if match:
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
            hour = int(match.group(4)) if match.group(4) else 0
            minute = int(match.group(5)) if match.group(5) else 0
            second = int(match.group(6)) if match.group(6) else 0
            try:
                reminder_date = datetime.datetime(year, month, day, hour, minute, second)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # ==== TEXT MONTH FORMATS ====
        # "December 25", "Dec 25", "December 25 2024", "Dec 25, 2024"
        match = re.search(r'\b(' + '|'.join(month_names.keys()) + r')\s+(\d{1,2})(?:,?\s+(\d{2,4}))?', text_lower)
        if match:
            month = month_names[match.group(1)]
            day = int(match.group(2))
            year = int(match.group(3)) if match.group(3) else now.year
            if year < 100:
                year += 2000
            try:
                reminder_date = datetime.datetime(year, month, day)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # "25 December", "25 Dec 2024", "25 December 2024"
        match = re.search(r'\b(\d{1,2})\s+(' + '|'.join(month_names.keys()) + r')(?:,?\s+(\d{2,4}))?', text_lower)
        if match:
            day = int(match.group(1))
            month = month_names[match.group(2)]
            year = int(match.group(3)) if match.group(3) else now.year
            if year < 100:
                year += 2000
            try:
                reminder_date = datetime.datetime(year, month, day)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # ==== TIME WITH AM/PM ====
        # "3pm", "3:30pm", "3:30 PM", "15:30"
        match = re.search(r'\b(\d{1,2})(?::(\d{2}))?\s*(am|pm)\b', text_lower)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2)) if match.group(2) else 0
            ampm = match.group(3)
            if ampm == 'pm' and hour < 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            # If time is in the past today, schedule for tomorrow
            reminder_date = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if reminder_date < now:
                reminder_date += datetime.timedelta(days=1)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # 24-hour format: "15:30", "09:45"
        match = re.search(r'\b(\d{1,2}):(\d{2})\b(?!\s*(?:am|pm))', text_lower)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            if 0 <= hour <= 23 and 0 <= minute <= 59:
                reminder_date = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if reminder_date < now:
                    reminder_date += datetime.timedelta(days=1)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())

        # ==== DATE WITH TIME COMBINATIONS ====
        # "HH:MM am/pm on MM/DD" or "HH:MM am/pm on MM/DD/YYYY"
        match = re.search(r'(\d{1,2}):(\d{2})\s*(am|pm)?\s+on\s+(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?', text_lower)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            ampm = match.group(3)
            month = int(match.group(4))
            day = int(match.group(5))
            year = int(match.group(6)) if match.group(6) else now.year
            if year < 100:
                year += 2000
            if ampm == 'pm' and hour < 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            try:
                reminder_date = datetime.datetime(year, month, day, hour, minute)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # "MM/DD HH:MM am/pm", "MM/DD/YYYY HH:MM am/pm", "MM-DD-YYYY HH:MM"
        match = re.search(r'(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?\s+(\d{1,2}):(\d{2})\s*(am|pm)?', text_lower)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            year = int(match.group(3)) if match.group(3) else now.year
            hour = int(match.group(4))
            minute = int(match.group(5))
            ampm = match.group(6)
            if year < 100:
                year += 2000
            if ampm == 'pm' and hour < 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            try:
                reminder_date = datetime.datetime(year, month, day, hour, minute)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # ==== DATE ONLY FORMATS ====
        # "MM/DD/YYYY", "MM-DD-YYYY", "MM.DD.YYYY"
        match = re.search(r'\b(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})\b', text_lower)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            year = int(match.group(3))
            if year < 100:
                year += 2000
            try:
                reminder_date = datetime.datetime(year, month, day)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # "MM/DD", "MM-DD" (current year)
        match = re.search(r'\b(\d{1,2})[/-](\d{1,2})\b', text_lower)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            try:
                reminder_date = datetime.datetime(now.year, month, day)
                # If date has passed this year, use next year
                if reminder_date < now:
                    reminder_date = datetime.datetime(now.year + 1, month, day)
                cleaned_text = text[:match.start()] + text[match.end():]
                return (reminder_date, cleaned_text.strip())
            except ValueError:
                pass

        # ==== RELATIVE TIME WITH DECIMALS ====
        # "1.5 hours", "2.5 days", "0.5 weeks"
        match = re.search(r'(?:in\s+)?(\d+(?:\.\d+)?)\s*(?:hours?|hrs?)\b', text_lower)
        if match:
            hours = float(match.group(1))
            reminder_date = now + datetime.timedelta(hours=hours)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        match = re.search(r'(?:in\s+)?(\d+(?:\.\d+)?)\s*(?:days?)\b', text_lower)
        if match:
            days = float(match.group(1))
            reminder_date = now + datetime.timedelta(days=days)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # ==== COMPOUND RELATIVE TIME ====
        # "2h30m", "1d12h", "2w3d"
        match = re.search(r'(\d+)h(?:\s*(\d+)m)?', text_lower)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2)) if match.group(2) else 0
            reminder_date = now + datetime.timedelta(hours=hours, minutes=minutes)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # ==== STANDARD RELATIVE TIME ====
        # "in X minutes" or "X minutes" or "X min"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:minutes?|mins?)\b', text_lower)
        if match:
            minutes = int(match.group(1))
            reminder_date = now + datetime.timedelta(minutes=minutes)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "in X hours" or "X hours" or "X hrs"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:hours?|hrs?)\b', text_lower)
        if match:
            hours = int(match.group(1))
            reminder_date = now + datetime.timedelta(hours=hours)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "in X seconds" or "X seconds" or "X secs"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:seconds?|secs?)\b', text_lower)
        if match:
            seconds = int(match.group(1))
            reminder_date = now + datetime.timedelta(seconds=seconds)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "X days" or "Xd"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:days?|d)\b', text_lower)
        if match:
            days = int(match.group(1))
            reminder_date = now + datetime.timedelta(days=days)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "X weeks" or "Xw" or "X wk"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:weeks?|wks?|w)\b', text_lower)
        if match:
            weeks = int(match.group(1))
            reminder_date = now + datetime.timedelta(weeks=weeks)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "X months" or "Xm" (but not time format like "3:30m")
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:months?)\b', text_lower)
        if match:
            months = int(match.group(1))
            reminder_date = now + relativedelta(months=months)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # "X years"
        match = re.search(r'(?:in\s+)?(\d+)\s*(?:years?|yrs?)\b', text_lower)
        if match:
            years = int(match.group(1))
            reminder_date = now + relativedelta(years=years)
            cleaned_text = text[:match.start()] + text[match.end():]
            return (reminder_date, cleaned_text.strip())

        # ==== NATURAL LANGUAGE ====
        # "today", "tonight"
        if 'tonight' in text_lower:
            reminder_date = now.replace(hour=20, minute=0, second=0, microsecond=0)
            if reminder_date < now:
                reminder_date += datetime.timedelta(days=1)
            cleaned_text = text_lower.replace('tonight', '').strip()
            return (reminder_date, cleaned_text)

        if 'today' in text_lower:
            reminder_date = now.replace(hour=17, minute=0, second=0, microsecond=0)
            if reminder_date < now:
                # If past 5pm, set to current time + 1 hour
                reminder_date = now + datetime.timedelta(hours=1)
            cleaned_text = text_lower.replace('today', '').strip()
            return (reminder_date, cleaned_text)

        # "tomorrow"
        if 'tomorrow' in text_lower:
            reminder_date = now + datetime.timedelta(days=1)
            reminder_date = reminder_date.replace(hour=9, minute=0, second=0, microsecond=0)
            cleaned_text = text_lower.replace('tomorrow', '').strip()
            return (reminder_date, cleaned_text)

        # "next week", "next month", "next year"
        if 'next week' in text_lower:
            reminder_date = now + datetime.timedelta(weeks=1)
            cleaned_text = text_lower.replace('next week', '').strip()
            return (reminder_date, cleaned_text)

        if 'next month' in text_lower:
            reminder_date = now + relativedelta(months=1)
            cleaned_text = text_lower.replace('next month', '').strip()
            return (reminder_date, cleaned_text)

        if 'next year' in text_lower:
            reminder_date = now + relativedelta(years=1)
            cleaned_text = text_lower.replace('next year', '').strip()
            return (reminder_date, cleaned_text)

        # Day of week: "Monday", "Tuesday", etc.
        weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        for i, day in enumerate(weekdays):
            if day in text_lower:
                current_weekday = now.weekday()
                target_weekday = i
                days_ahead = (target_weekday - current_weekday) % 7
                if days_ahead == 0:
                    days_ahead = 7  # Next week if same day
                reminder_date = now + datetime.timedelta(days=days_ahead)
                cleaned_text = text_lower.replace(day, '').strip()
                return (reminder_date, cleaned_text)

        # "next Monday", "next Friday", etc.
        for i, day in enumerate(weekdays):
            pattern = f'next {day}'
            if pattern in text_lower:
                current_weekday = now.weekday()
                target_weekday = i
                days_ahead = (target_weekday - current_weekday) % 7
                if days_ahead == 0:
                    days_ahead = 7
                days_ahead += 7  # "next" means the following week
                reminder_date = now + datetime.timedelta(days=days_ahead)
                cleaned_text = text_lower.replace(pattern, '').strip()
                return (reminder_date, cleaned_text)

        # No date found
        return (None, text)
    except Exception as e:
        import traceback
        error_msg = f"Error in parse_natural_date: {e}\n\n{traceback.format_exc()}"
        print(f"\n{'='*60}")
        print("PARSE NATURAL DATE ERROR")
        print(f"{'='*60}")
        print(error_msg)
        print(f"{'='*60}\n")
        # Return gracefully on any error
        return (None, text)


# ============================================================================

class Project:
    """A container for a single project's tasks and metadata."""
    def __init__(self, name, uid):
        self.name = name
        self.uid = uid
        self.root_tasks = []
        self.start, self.finish, self.is_milestone = None, None, False
        self.bar_rect, self.progress, self.is_active, self.comments = QRectF(), 0, True, []
        self.children = self.root_tasks
        self.source_plugin = None

    @property
    def effective_progress(self):
        if not self.root_tasks: return 0
        total_duration, weighted_progress = 0, 0
        for child in self.root_tasks:
            if hasattr(child, 'start') and child.start and hasattr(child, 'finish') and child.finish and hasattr(child, 'effective_progress'):
                duration = (child.finish - child.start).days
                if duration > 0:
                    total_duration += duration
                    weighted_progress += child.effective_progress * duration
        return 0 if total_duration == 0 else weighted_progress / total_duration

    def update_dates(self):
        all_tasks = self.get_all_tasks()
        if not all_tasks: 
            self.start, self.finish = None, None
            return
        valid_tasks = [t for t in all_tasks if hasattr(t, 'start') and t.start and hasattr(t, 'finish') and t.finish]
        if valid_tasks:
            self.start = min(t.start for t in valid_tasks)
            self.finish = max(t.finish for t in valid_tasks)
        else:
            self.start, self.finish = None, None

    def get_all_tasks(self, tasks=None):
        if tasks is None: tasks = self.root_tasks
        flat_list = []
        for task in tasks:
            flat_list.append(task)
            if hasattr(task, 'get_all_subtasks'):
                flat_list.extend(task.get_all_subtasks())
        return flat_list

    def get_savable_tasks(self):
        savable_tasks = []
        items_to_check = list(self.root_tasks)
        while items_to_check:
            item = items_to_check.pop(0)
            if isinstance(item, Task):
                savable_tasks.append(item)
            if hasattr(item, 'children'):
                items_to_check.extend(item.children)
        return savable_tasks

class Task:
    """A simple class to hold our task data in a structured way."""
    def __init__(self, uid, name, start, finish, outline_level, priority=Priority.MEDIUM):
        self.uid, self.name, self.start, self.finish, self.outline_level, self.priority = uid, name, start, finish, outline_level, priority
        self.is_milestone = (finish.date() - start.date()).days <= 0
        self.children, self.parent, self.bar_rect, self.progress = [], None, QRectF(), 0
        self.comments = []
        self.notes = []  # For grid notes from Excel
        self.project = None
        self.is_active = True

    @property
    def effective_progress(self):
        if not self.children: return self.progress
        total_duration, weighted_progress = 0, 0
        for child in self.children:
            if child.start and child.finish:
                duration = (child.finish - child.start).days
                if duration > 0:
                    total_duration += duration
                    weighted_progress += child.effective_progress * duration
        return 0 if total_duration == 0 else weighted_progress / total_duration

    def to_dict(self):
        return {'uid': self.uid, 'name': self.name, 'start': self.start.isoformat(), 'finish': self.finish.isoformat(), 'outline_level': self.outline_level, 'parent_uid': self.parent.uid if self.parent else None, 'progress': self.progress, 'priority': self.priority.value, 'is_active': self.is_active, 'comments': self.comments}

    @staticmethod
    def from_dict(data):
        start, finish = datetime.datetime.fromisoformat(data['start']), datetime.datetime.fromisoformat(data['finish'])
        task = Task(data['uid'], data['name'], start, finish, data['outline_level'], Priority(data.get('priority', Priority.MEDIUM.value)))
        task.progress, task.parent_uid = data.get('progress', 0), data.get('parent_uid')
        task.is_active = data.get('is_active', True)
        task.comments = data.get('comments', [])
        return task
    
    def get_all_subtasks(self):
        all_subtasks = []
        for child in self.children:
            all_subtasks.append(child)
            all_subtasks.extend(child.get_all_subtasks())
        return all_subtasks

class AddItemDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Items")
        self.setMinimumSize(800, 500)
        self.setStyleSheet("""
            QTableWidget::item:selected { background-color: #cce8ff; }
            QTableWidget::item { padding: 5px; }
        """)

        layout = QVBoxLayout(self)

        self.table = QTableWidget(1, 6)
        self.table.setHorizontalHeaderLabels(["Name", "Description", "Milestone?", "Start Date", "End Date", "Priority"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 200)
        self.table.setColumnWidth(1, 250)
        layout.addWidget(self.table)
        self.add_row(0)

        button_layout = QHBoxLayout()
        paste_button = QPushButton("Paste Names & Descriptions")
        paste_button.clicked.connect(self.paste_from_clipboard)
        button_layout.addWidget(paste_button)
        add_row_button = QPushButton("Add Row")
        add_row_button.clicked.connect(self.add_new_row)
        button_layout.addWidget(add_row_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        dialog_buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Add Items")
        dialog_buttons.accepted.connect(self.validate_and_accept)
        dialog_buttons.rejected.connect(self.reject)
        layout.addWidget(dialog_buttons)

    def add_row(self, row_num):
        self.table.setItem(row_num, 0, QTableWidgetItem())
        self.table.setItem(row_num, 1, QTableWidgetItem())

        milestone_check = QCheckBox()
        milestone_check.stateChanged.connect(lambda state, r=row_num: self.toggle_end_date(r, state))
        cell_widget = QWidget()
        layout = QHBoxLayout(cell_widget)
        layout.addWidget(milestone_check)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setContentsMargins(0,0,0,0)
        self.table.setCellWidget(row_num, 2, cell_widget)
        
        start_date_edit = QDateEdit(QDate.currentDate(), calendarPopup=True)
        start_date_edit.dateChanged.connect(lambda date, r=row_num: self.sync_milestone_date(r, date))
        end_date_edit = QDateEdit(QDate.currentDate(), calendarPopup=True)
        self.table.setCellWidget(row_num, 3, start_date_edit)
        self.table.setCellWidget(row_num, 4, end_date_edit)

        priority_combo = QComboBox()
        for p in Priority: priority_combo.addItem(p.name.capitalize(), p)
        priority_combo.setCurrentIndex(1)
        self.table.setCellWidget(row_num, 5, priority_combo)

    def add_new_row(self):
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        self.add_row(row_count)

    def paste_from_clipboard(self):
        clipboard = QApplication.clipboard().text()
        lines = clipboard.strip().split('\n')
        current_row = self.table.rowCount() -1
        
        for i, line in enumerate(lines):
            if not line.strip(): continue
            parts = line.split('\t')
            if i > 0: self.add_new_row()
            
            if current_row + i >= self.table.rowCount(): self.add_new_row()

            name_item = self.table.item(current_row + i, 0)
            if name_item: name_item.setText(parts[0].strip() if len(parts) > 0 else "")
            
            desc_item = self.table.item(current_row + i, 1)
            if desc_item: desc_item.setText(parts[1].strip() if len(parts) > 1 else "")
    
    def toggle_end_date(self, row, state):
        end_date_widget = self.table.cellWidget(row, 4)
        is_milestone = state == Qt.CheckState.Checked.value
        end_date_widget.setEnabled(not is_milestone)
        if is_milestone:
            self.sync_milestone_date(row, self.table.cellWidget(row, 3).date())

    def sync_milestone_date(self, row, date):
        milestone_checkbox = self.table.cellWidget(row, 2).findChild(QCheckBox)
        if milestone_checkbox and milestone_checkbox.isChecked():
            end_date_widget = self.table.cellWidget(row, 4)
            end_date_widget.setDate(date)

    def validate_and_accept(self):
        self.tasks_to_create = []
        errors = []
        
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                if self.table.item(r, c): self.table.item(r, c).setBackground(QColor("white"))

        for r in range(self.table.rowCount()):
            name_item = self.table.item(r, 0)
            name = name_item.text().strip() if name_item else ""
            
            if not name:
                if self.table.rowCount() > 1: continue
                else: errors.append(f"Row {r+1}: Name cannot be empty."); self.table.item(r, 0).setBackground(QColor("#ffcccc")); continue

            desc_item = self.table.item(r, 1)
            description = desc_item.text().strip() if desc_item else ""

            milestone_widget = self.table.cellWidget(r, 2)
            is_milestone = milestone_widget.findChild(QCheckBox).isChecked()

            start_date = self.table.cellWidget(r, 3).date().toPyDate()
            end_date = self.table.cellWidget(r, 4).date().toPyDate()
            
            if start_date > end_date and not is_milestone:
                errors.append(f"Row {r+1}: Start date cannot be after end date.");
                self.table.item(r,0).setBackground(QColor("#ffcccc")); continue

            priority_widget = self.table.cellWidget(r, 5)
            priority = priority_widget.currentData()
            
            task_data = {
                "name": name, "description": description, "is_milestone": is_milestone,
                "start": start_date, "end": end_date, "priority": priority
            }
            self.tasks_to_create.append(task_data)
        
        if errors:
            QMessageBox.warning(self, "Validation Error", "Please fix the highlighted errors before proceeding.\n\n" + "\n".join(errors))
            return
        
        super().accept()

    def get_tasks(self):
        return self.tasks_to_create

class CommentDialog(QDialog):
    def __init__(self, task, parent=None):
        super().__init__(parent)
        self.task = task
        self.setWindowTitle(f"Comments for '{task.name}'"); self.setMinimumSize(500, 400); layout = QVBoxLayout(self)
        self.comment_list = QListWidget(); layout.addWidget(self.comment_list)
        self.comment_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.comment_list.customContextMenuRequested.connect(self.show_comment_context_menu)
        for comment in self.task.comments: self.add_comment_to_list(comment)
        self.new_comment_edit = QTextEdit(); self.new_comment_edit.setPlaceholderText("Enter new comment or reminder..."); layout.addWidget(self.new_comment_edit)
        reminder_layout = QHBoxLayout()
        self.reminder_check = QCheckBox("Set Reminder")
        self.reminder_date_edit = QDateEdit(QDate.currentDate()); self.reminder_date_edit.setCalendarPopup(True); self.reminder_date_edit.setVisible(False)
        self.priority_combo = QComboBox(); self.priority_combo.setVisible(False)
        for p in Priority: self.priority_combo.addItem(p.name.capitalize(), p)
        self.reminder_check.stateChanged.connect(self.reminder_date_edit.setVisible); self.reminder_check.stateChanged.connect(self.priority_combo.setVisible)
        reminder_layout.addWidget(self.reminder_check); reminder_layout.addWidget(self.reminder_date_edit); reminder_layout.addWidget(QLabel("Priority:")); reminder_layout.addWidget(self.priority_combo); reminder_layout.addStretch()
        layout.addLayout(reminder_layout)
        add_button = QPushButton("Add Comment"); add_button.clicked.connect(self.add_comment); layout.addWidget(add_button)

    def add_comment_to_list(self, comment_data):
        display_text = f"[{comment_data['timestamp']}] {comment_data['text']}"
        if reminder := comment_data.get('reminder'): display_text += f" (Reminder: {reminder})"
        item = QListWidgetItem(display_text)
        item.setData(Qt.ItemDataRole.UserRole, comment_data)
        self.comment_list.addItem(item)

    def show_comment_context_menu(self, pos):
        item = self.comment_list.itemAt(pos)
        if not item: return
        menu = QMenu()
        delete_action = QAction("Delete Comment", self)
        delete_action.triggered.connect(lambda: self.delete_comment(item))
        menu.addAction(delete_action)
        menu.exec(self.comment_list.mapToGlobal(pos))

    def delete_comment(self, item):
        comment_data = item.data(Qt.ItemDataRole.UserRole)
        if comment_data in self.task.comments:
            self.task.comments.remove(comment_data)
        self.comment_list.takeItem(self.comment_list.row(item))

    def add_comment(self):
        comment_text = self.new_comment_edit.toPlainText().strip()
        if not comment_text: return
        new_comment = {'text': comment_text, 'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        if self.reminder_check.isChecked():
            new_comment['reminder'] = self.reminder_date_edit.date().toString("yyyy-MM-dd")
            new_comment['priority'] = self.priority_combo.currentData().value
            new_comment['reminder_complete'] = False
        self.task.comments.append(new_comment)
        self.add_comment_to_list(new_comment)
        self.new_comment_edit.clear(); self.reminder_check.setChecked(False)

class SearchResultsDialog(QDialog):
    task_selected = pyqtSignal(object)
    def __init__(self, results, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Search Results")
        self.setMinimumSize(500, 400)
        layout = QVBoxLayout(self)
        self.results_list = QListWidget()
        layout.addWidget(self.results_list)

        if not results:
            self.results_list.addItem("No results found.")
        else:
            for task, context in results:
                item = QListWidgetItem(f"{context}")
                item.setData(Qt.ItemDataRole.UserRole, task)
                self.results_list.addItem(item)
        
        self.results_list.itemDoubleClicked.connect(self.emit_task_selected)
        self.results_list.itemClicked.connect(self.emit_task_selected)  # Also handle single click

    def emit_task_selected(self, item):
        if item.data(Qt.ItemDataRole.UserRole) is not None:
            self.task_selected.emit(item.data(Qt.ItemDataRole.UserRole))
            self.accept()

class NotificationDialog(QDialog):
    tasks_to_complete, reminders_to_complete = pyqtSignal(list), pyqtSignal(list)
    task_selected = pyqtSignal(object)
    reminder_selected = pyqtSignal(object)
    
    def __init__(self, late_tasks, due_soon_tasks, reminder_tasks, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Notification Center"); self.setMinimumSize(600, 400); layout = QVBoxLayout(self)
        self.tab_widget = QTabWidget(); layout.addWidget(self.tab_widget)
        self.late_list, self.due_soon_list, self.reminder_list = QListWidget(), QListWidget(), QListWidget()
        self.late_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection); self.due_soon_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        self.tab_widget.addTab(self.late_list, f"Past Due ({len(late_tasks)})"); self.tab_widget.addTab(self.due_soon_list, f"Due This Week ({len(due_soon_tasks)})"); self.tab_widget.addTab(self.reminder_list, f"Reminders ({len(reminder_tasks)})")
        
        priority_colors = {Priority.LOW.value: "green", Priority.MEDIUM.value: "orange", Priority.HIGH.value: "red"}
        
        for task in sorted(late_tasks, key=lambda t: t.finish):
            item = QListWidgetItem(f"{task.name} (Due: {task.finish.strftime('%Y-%m-%d')})"); item.setData(Qt.ItemDataRole.UserRole, task.uid); self.late_list.addItem(item)
        for task in sorted(due_soon_tasks, key=lambda t: t.finish):
            item = QListWidgetItem(f"{task.name} (Due: {task.finish.strftime('%Y-%m-%d')})"); item.setData(Qt.ItemDataRole.UserRole, task.uid); self.due_soon_list.addItem(item)
        for task, comment in sorted(reminder_tasks, key=lambda x: x[1]['reminder']):
            priority = comment.get('priority', Priority.MEDIUM.value)
            color = priority_colors.get(priority, "black")
            item_text = f"<span style='font-size: 18px; color: {color};'>●</span> <b>'{task.name}'</b>: {comment['text']} (Reminder: {comment['reminder']})"
            label = QLabel(item_text)
            label.setWordWrap(True)
            item = QListWidgetItem()
            item.setSizeHint(label.sizeHint())
            self.reminder_list.addItem(item)
            self.reminder_list.setItemWidget(item, label)
            print(f"DEBUG: Storing reminder for task.uid: {task.uid} (type: {type(task.uid)})")
            item.setData(Qt.ItemDataRole.UserRole, (task.uid, comment['timestamp']))

        self.late_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu); self.due_soon_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu); self.reminder_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.late_list.customContextMenuRequested.connect(self.show_task_context_menu); self.due_soon_list.customContextMenuRequested.connect(self.show_task_context_menu); self.reminder_list.customContextMenuRequested.connect(self.show_reminder_context_menu)
        self.late_list.itemDoubleClicked.connect(self.emit_task_selected); self.due_soon_list.itemDoubleClicked.connect(self.emit_task_selected)
        self.reminder_list.itemDoubleClicked.connect(self.emit_reminder_selected)

    def emit_task_selected(self, item):
        self.task_selected.emit(item.data(Qt.ItemDataRole.UserRole))
        self.accept()
        
    def emit_reminder_selected(self, item):
        uid, _ = item.data(Qt.ItemDataRole.UserRole)
        self.reminder_selected.emit(uid)
        self.accept()

    def show_task_context_menu(self, pos):
        list_widget = self.sender()
        if not list_widget.selectedItems(): return
        menu = QMenu(); mark_complete_action = QAction("Mark Selected as Complete", self)
        mark_complete_action.triggered.connect(lambda: self.mark_selected_tasks_complete(list_widget)); menu.addAction(mark_complete_action)
        menu.exec(list_widget.mapToGlobal(pos))

    def show_reminder_context_menu(self, pos):
        if not self.reminder_list.selectedItems(): return
        menu = QMenu(); mark_complete_action = QAction("Mark Reminder as Complete", self)
        mark_complete_action.triggered.connect(self.mark_selected_reminders_complete); menu.addAction(mark_complete_action)
        menu.exec(self.reminder_list.mapToGlobal(pos))

    def mark_selected_tasks_complete(self, list_widget):
        selected_items = list_widget.selectedItems()
        if not selected_items: return
        self.tasks_to_complete.emit([item.data(Qt.ItemDataRole.UserRole) for item in selected_items])
        for item in selected_items: list_widget.takeItem(list_widget.row(item))
        self.tab_widget.setTabText(0, f"Past Due ({self.late_list.count()})"); self.tab_widget.setTabText(1, f"Due This Week ({self.due_soon_list.count()})")

    def mark_selected_reminders_complete(self):
        selected_items = self.reminder_list.selectedItems()
        if not selected_items: return
        self.reminders_to_complete.emit([item.data(Qt.ItemDataRole.UserRole) for item in selected_items])
        for item in selected_items: self.reminder_list.takeItem(self.reminder_list.row(item))
        self.tab_widget.setTabText(2, f"Reminders ({self.reminder_list.count()})")

class FilterDialog(QDialog):
    def __init__(self, data, current_filters, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filter Data")
        self.setMinimumSize(400, 500)
        layout = QVBoxLayout(self)

        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        self.data = data
        self.current_filters = current_filters if current_filters is not None else {}
        self.list_widgets = {}

        if not data:
            layout.addWidget(QLabel("No data to filter."))
        else:
            headers = data[0].keys()
            for header in headers:
                tab = QWidget()
                tab_layout = QVBoxLayout(tab)
                
                btn_layout = QHBoxLayout()
                select_all_btn = QPushButton("Select All")
                deselect_all_btn = QPushButton("Deselect All")
                btn_layout.addWidget(select_all_btn)
                btn_layout.addWidget(deselect_all_btn)
                btn_layout.addStretch()
                tab_layout.addLayout(btn_layout)

                list_widget = QListWidget()
                self.list_widgets[header] = list_widget
                unique_values = sorted(list(set(str(d.get(header, '')) for d in self.data)))
                
                active_values = self.current_filters.get(header)

                for value in unique_values:
                    item = QListWidgetItem(value)
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    if active_values is None or value in active_values:
                        item.setCheckState(Qt.CheckState.Checked)
                    else:
                        item.setCheckState(Qt.CheckState.Unchecked)
                    list_widget.addItem(item)
                
                tab_layout.addWidget(list_widget)
                self.tab_widget.addTab(tab, header)

                select_all_btn.clicked.connect(lambda _, lw=list_widget: self.toggle_all(lw, True))
                deselect_all_btn.clicked.connect(lambda _, lw=list_widget: self.toggle_all(lw, False))
        
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        dialog_buttons.accepted.connect(self.accept)
        dialog_buttons.rejected.connect(self.reject)
        layout.addWidget(dialog_buttons)

    def toggle_all(self, list_widget, check):
        state = Qt.CheckState.Checked if check else Qt.CheckState.Unchecked
        for i in range(list_widget.count()):
            list_widget.item(i).setCheckState(state)

    def get_filters(self):
        filters = {}
        if not self.data:
            return filters
            
        for header, list_widget in self.list_widgets.items():
            selected_values = []
            all_values_present = list_widget.count() > 0
            all_checked = all_values_present
            
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                if item.checkState() == Qt.CheckState.Checked:
                    selected_values.append(item.text())
                else:
                    all_checked = False

            if not all_checked and all_values_present:
                 filters[header] = selected_values
        return filters

class AdvancedSummaryDialog(QDialog):
    def __init__(self, summary_data, parent=None):
        super().__init__(parent)
        self.all_summary_data = summary_data
        self.active_filters = {}
        self.setWindowTitle("ECN/PR Advanced Summary")
        self.setMinimumSize(900, 500)
        
        layout = QVBoxLayout(self)
        
        filter_layout = QHBoxLayout()
        filter_button = QPushButton("Apply Filters...")
        filter_button.clicked.connect(self.show_filter_dialog)
        filter_layout.addWidget(filter_button)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.table = QTableWidget()
        layout.addWidget(self.table)
        
        button_layout = QHBoxLayout()
        export_button = QPushButton("Export to CSV")
        export_button.clicked.connect(self.export_to_csv)
        button_layout.addStretch()
        button_layout.addWidget(export_button)
        layout.addLayout(button_layout)
        
        self.populate_table(self.all_summary_data)

    def show_filter_dialog(self):
        dialog = FilterDialog(self.all_summary_data, self.active_filters, self)
        if dialog.exec():
            self.active_filters = dialog.get_filters()
            self.apply_filters()

    def apply_filters(self):
        if not self.active_filters:
            self.populate_table(self.all_summary_data)
            return

        filtered_data = self.all_summary_data
        for header, values in self.active_filters.items():
            filtered_data = [d for d in filtered_data if str(d.get(header, '')) in values]
            
        self.populate_table(filtered_data)

    def populate_table(self, data):
        self.table.clear()
        if not data:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        headers = list(data[0].keys())
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(data))

        for row_idx, row_data in enumerate(data):
            current_task = str(row_data.get("Current Task / Name", ""))
            overall_status = str(row_data.get("Overall Status", ""))
            progress = str(row_data.get("Progress", ""))

            # Determine row color
            if "On Hold" in current_task:
                bg_color = QColor("#FFF3CD")  # yellow for on-hold
            elif overall_status == "Rejected":
                bg_color = QColor("#F8D7DA")  # light red for rejected
            elif overall_status == "Completed" or progress == "100%":
                bg_color = QColor("#D4EDDA")  # light green for completed
            else:
                bg_color = None

            for col_idx, header in enumerate(headers):
                item = QTableWidgetItem(str(row_data.get(header, '')))
                if bg_color:
                    item.setBackground(bg_color)
                self.table.setItem(row_idx, col_idx, item)

        self.table.resizeColumnsToContents()

    def export_to_csv(self):
        row_count = self.table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "No Data", "There is no data to export.")
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "Save Summary", "", "CSV Files (*.csv)")
        if not file_name:
            return
            
        try:
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            data_to_export = []
            for row in range(row_count):
                row_data = {header: self.table.item(row, col).text() for col, header in enumerate(headers)}
                data_to_export.append(row_data)

            with open(file_name, 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data_to_export)
            QMessageBox.information(self, "Export Successful", f"Summary successfully exported to {file_name}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not export file:\n{e}")

class SummaryDialog(QDialog):
    def __init__(self, report_text, parent=None):
        super().__init__(parent); self.setWindowTitle("Project Status Summary"); self.setMinimumSize(500, 400); layout = QVBoxLayout(self)
        self.report_display = QTextEdit(); self.report_display.setReadOnly(True); self.report_display.setText(report_text); layout.addWidget(self.report_display)
        self.copy_button = QPushButton("Copy to Clipboard"); self.copy_button.clicked.connect(self.copy_to_clipboard); layout.addWidget(self.copy_button)
    def copy_to_clipboard(self): QGuiApplication.clipboard().setText(self.report_display.toPlainText()); self.copy_button.setText("Copied!"); self.copy_button.setEnabled(False)

class DeleteOptionsDialog(QDialog):
    def __init__(self, task_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Delete Level")
        self.setMinimumSize(400, 200)
        
        layout = QVBoxLayout(self)
        
        # Question label
        question_label = QLabel(f"How do you want to delete '{task_name}'?")
        question_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(question_label)
        
        # Radio button group
        self.button_group = QButtonGroup(self)
        
        # Option 1: Delete everything below
        self.delete_all_radio = QRadioButton("Delete everything below")
        self.delete_all_radio.setChecked(True)  # Default selection
        self.button_group.addButton(self.delete_all_radio, 0)
        layout.addWidget(self.delete_all_radio)
        
        delete_all_desc = QLabel("(Removes this level and all sub-levels permanently)")
        delete_all_desc.setStyleSheet("color: #666; margin-left: 20px;")
        layout.addWidget(delete_all_desc)
        
        # Option 2: Delete level only, move children up
        self.delete_level_radio = QRadioButton("Delete level only, move children up")
        self.button_group.addButton(self.delete_level_radio, 1)
        layout.addWidget(self.delete_level_radio)
        
        delete_level_desc = QLabel("(Removes this level, children move up one level)")
        delete_level_desc.setStyleSheet("color: #666; margin-left: 20px;")
        layout.addWidget(delete_level_desc)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QHBoxLayout()
        delete_button = QPushButton("Delete")
        delete_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(delete_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
    
    def get_delete_option(self):
        """Returns 'delete_all' or 'delete_level_only'"""
        if self.delete_all_radio.isChecked():
            return 'delete_all'
        else:
            return 'delete_level_only'

class PRPlacementDialog(QDialog):
    """Dialog for configuring PR placement into projects and hierarchy levels."""
    def __init__(self, pr_data, available_projects, parent=None):
        super().__init__(parent)
        self.pr_data = pr_data  # List of dicts with PR info
        self.available_projects = available_projects  # List of Project objects
        self.placement_config = {}  # Will store user selections
        
        self.setWindowTitle("PR Placement Configuration")
        self.setMinimumSize(1000, 600)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        # Header
        header_label = QLabel("Map each PR to its target project and hierarchy level:")
        header_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
        layout.addWidget(header_label)
        
        # Create table for PR placement
        self.table = QTableWidget(len(pr_data), 5)
        self.table.setHorizontalHeaderLabels(["PR ID", "Project Description", "Target Project", "Target Level", "Status"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 120)  # PR ID
        self.table.setColumnWidth(1, 300)  # Project Description (wider for longer descriptions)
        self.table.setColumnWidth(2, 200)  # Target Project
        self.table.setColumnWidth(3, 200)  # Target Level
        self.table.setColumnWidth(4, 100)  # Status
        
        # Populate table
        for row, pr_info in enumerate(pr_data):
            # PR ID
            pr_id_item = QTableWidgetItem(pr_info.get('pr_id', ''))
            pr_id_item.setFlags(pr_id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 0, pr_id_item)
            
            # Project Description (from Excel column C)
            project_description_item = QTableWidgetItem(pr_info.get('project_description', ''))
            project_description_item.setFlags(project_description_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 1, project_description_item)
            
            # Target Project Dropdown
            project_combo = QComboBox()
            project_combo.addItem("Select Project...", None)
            for project in available_projects:
                project_combo.addItem(project.name, project.uid)
            project_combo.currentIndexChanged.connect(lambda idx, r=row: self.on_project_changed(r, idx))
            self.table.setCellWidget(row, 2, project_combo)
            
            # Target Level Dropdown
            level_combo = QComboBox()
            level_combo.addItem("Select Level...", None)
            level_combo.setEnabled(False)  # Disabled until project is selected
            self.table.setCellWidget(row, 3, level_combo)
            
            # Status
            status_item = QTableWidgetItem("Pending")
            status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 4, status_item)
        
        layout.addWidget(self.table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        apply_all_button = QPushButton("Apply All")
        apply_all_button.clicked.connect(self.apply_all_placements)
        button_layout.addWidget(apply_all_button)
        
        save_default_button = QPushButton("Save as Default")
        save_default_button.clicked.connect(self.save_as_default)
        button_layout.addWidget(save_default_button)
        
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        ok_button = QPushButton("Apply")
        ok_button.clicked.connect(self.accept)
        button_layout.addWidget(ok_button)
        
        layout.addLayout(button_layout)
    
    def on_project_changed(self, row, project_index):
        """Handle project selection change - populate level dropdown"""
        project_combo = self.table.cellWidget(row, 2)
        level_combo = self.table.cellWidget(row, 3)
        
        if project_index > 0:  # A project was selected
            project_uid = project_combo.currentData()
            project = next((p for p in self.available_projects if p.uid == project_uid), None)
            
            if project:
                level_combo.clear()
                level_combo.addItem("Select Level...", None)
                
                # Add project root as an option
                level_combo.addItem(f"{project.name} (Root)", f"{project_uid}_root")
                
                # Add all tasks as potential levels
                self._populate_levels(level_combo, project.root_tasks, project_uid, 1)
                
                level_combo.setEnabled(True)
        else:
            level_combo.clear()
            level_combo.addItem("Select Level...", None)
            level_combo.setEnabled(False)
    
    def _populate_levels(self, combo, tasks, project_uid, level):
        """Recursively populate level dropdown with task hierarchy"""
        for task in tasks:
            indent = "  " * level
            combo.addItem(f"{indent}{task.name}", f"{project_uid}_{task.uid}")
            if task.children:
                self._populate_levels(combo, task.children, project_uid, level + 1)
    
    def apply_all_placements(self):
        """Apply all configured placements"""
        for row in range(self.table.rowCount()):
            self._apply_placement(row)
    
    def _apply_placement(self, row):
        """Apply placement for a single row"""
        project_combo = self.table.cellWidget(row, 2)
        level_combo = self.table.cellWidget(row, 3)
        status_item = self.table.item(row, 4)
        
        if project_combo.currentIndex() > 0 and level_combo.currentIndex() > 0:
            pr_id = self.table.item(row, 0).text()
            project_uid = project_combo.currentData()
            level_uid = level_combo.currentData()
            
            self.placement_config[pr_id] = {
                'project_uid': project_uid,
                'level_uid': level_uid,
                'last_updated': datetime.datetime.now().isoformat()
            }
            
            status_item.setText("✓")
            status_item.setBackground(QColor("#d4edda"))
        else:
            status_item.setText("Pending")
            status_item.setBackground(QColor("#fff3cd"))
    
    def save_as_default(self):
        """Save current configuration as default for future imports"""
        self.apply_all_placements()
        # This will be handled by the main window
        QMessageBox.information(self, "Saved", "Configuration saved as default for future imports.")
    
    def get_placement_config(self):
        """Return the placement configuration"""
        self.apply_all_placements()
        return self.placement_config

class LegendDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Timeline Legend")
        layout = QFormLayout(self)
        layout.setRowWrapPolicy(QFormLayout.RowWrapPolicy.WrapAllRows)
        
        items = [
            ("#ffc107", "<b>In Progress (ECN)</b>: Task has started but not yet complete."),
            ("#17a2b8", "<b>ECN Released</b>: All items in the ECN have been released (90% complete)."),
            ("#28a745", "<b>MCN Released</b>: At least one item is fully MCN released (100% complete)."),
            ("#dc3545", "<b>Needs Attention (Outline)</b>: An item is Rejected or has a stalled MCN task."),
            ("#dc3545", "<b>Overdue (Standard Task)</b>: A non-ECN task that is past its finish date.")
        ]

        for color, text in items:
            color_label = QLabel()
            if "Outline" in text:
                color_label.setStyleSheet(f"border: 2px solid {color}; background-color: #f0f0f0; min-width: 40px;")
            else:
                 color_label.setStyleSheet(f"background-color: {color}; min-width: 40px;")
            
            text_label = QLabel(text)
            text_label.setWordWrap(True)
            layout.addRow(color_label, text_label)

# ============================================================================
# NOTES VIEW - ONENOTE-STYLE INTERFACE
# ============================================================================

class NotesView(QWidget):
    """Main OneNote-style notes interface with 3-column layout."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self.notes_structure = NotesStructure()
        self.current_project = None
        self.current_item = None

        self.init_ui()

    def init_ui(self):
        """Initialize the 3-column layout."""
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Create splitter for 3 columns
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # LEFT COLUMN: Projects List
        self.project_list = ProjectListWidget(self)
        self.project_list.project_selected.connect(self.on_project_selected)
        splitter.addWidget(self.project_list)

        # MIDDLE COLUMN: Items List
        self.items_list = ItemsListWidget(self)
        self.items_list.item_selected.connect(self.on_item_selected)
        self.items_list.item_moved.connect(self.on_item_moved)
        splitter.addWidget(self.items_list)

        # RIGHT COLUMN: Tabs + Notes Editor
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)

        # TOP: Tabs for children
        self.child_tabs = ChildTabWidget(self)
        self.child_tabs.tab_selected.connect(self.on_child_tab_selected)
        right_layout.addWidget(self.child_tabs)

        # Formatting toolbar (OneNote-style) — starts disabled until a note is focused
        self.formatting_toolbar = FormattingToolbar(self)
        self.formatting_toolbar.set_editor(None)
        right_layout.addWidget(self.formatting_toolbar)

        # Small toolbar with grid toggle
        mode_toolbar = QWidget()
        mode_toolbar.setStyleSheet("background: #F9F7FB;")
        mode_toolbar_layout = QHBoxLayout(mode_toolbar)
        mode_toolbar_layout.setContentsMargins(5, 2, 5, 2)
        mode_toolbar_layout.setSpacing(5)
        mode_toolbar_layout.addStretch()

        # Small checkbox for "Toggle Grid Mode" - controls note type for new notes
        self.grid_toggle_checkbox = QCheckBox("Grid Mode")
        self.grid_toggle_checkbox.setToolTip("When checked, new notes will be created as grid/table notes instead of text notes")
        self.grid_toggle_checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
                font-size: 11px;
                background: transparent;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
        """)
        # No need to connect to toggle_view_mode - NotesCanvas checks checkbox state directly
        mode_toolbar_layout.addWidget(self.grid_toggle_checkbox)
        right_layout.addWidget(mode_toolbar)

        # BOTTOM: Notes canvas (always shown)
        self.notes_canvas = NotesCanvas(self)
        self.notes_canvas.note_added.connect(self.on_note_added)
        right_layout.addWidget(self.notes_canvas, 1)

        splitter.addWidget(right_container)

        # Set splitter proportions (LEFT:MIDDLE:RIGHT = 1:2:4)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        splitter.setStretchFactor(2, 4)

        main_layout.addWidget(splitter)

        # Test grid note will be created in set_projects() after projects are loaded

    def set_projects(self, projects):
        """Set the list of projects from MainWindow - SIMPLIFIED"""
        print(f"[DEBUG] set_projects() called with {len(projects)} projects")

        self.project_list.set_projects(projects)

        # SIMPLE: Always auto-populate (grid notes are in task.notes, will be picked up)
        print("[DEBUG] Auto-populating NotesItems from project tasks...")
        for project in projects:
            self.notes_structure.auto_populate_from_project(project)

    def on_project_selected(self, project):
        """Handle project selection from LEFT column."""
        print(f"[DEBUG] on_project_selected() called for project: {project.name}, uid={project.uid}")
        self.current_project = project
        self.current_item = None

        # Update MIDDLE column with items for this project
        project_structure = self.notes_structure.get_project_structure(project.uid)
        print(f"[DEBUG] Retrieved project_structure with {len(project_structure['items'])} items")
        for i, item in enumerate(project_structure['items']):
            safe_name = item.name.encode('ascii', errors='replace').decode('ascii')
            print(f"[DEBUG]   Item {i}: {safe_name}, notes={len(item.notes)}, id={id(item)}")
        self.items_list.set_items(project_structure['items'])

        # Clear RIGHT side
        self.child_tabs.clear()
        self.notes_canvas.clear()

    def on_item_selected(self, item):
        """Handle item selection from MIDDLE column."""
        safe_name = item.name.encode('ascii', errors='replace').decode('ascii') if item else None
        print(f"[DEBUG] NotesView.on_item_selected() called with item: {safe_name}")
        self.current_item = item

        # Update TOP tabs with children
        self.child_tabs.set_children(item.children if item else [])

        # Display the item in canvas (always use notes_canvas)
        if item:
            safe_name = item.name.encode('ascii', errors='replace').decode('ascii')
            print(f"[DEBUG]   Calling notes_canvas.set_item() for: {safe_name}")
            self.notes_canvas.set_item(item)
        else:
            print(f"[DEBUG]   Item is None, not calling set_item")

        # Refresh indicators to show current status
        self.items_list.update_indicators()
        self.child_tabs.update_indicators()

    def on_child_tab_selected(self, child_item):
        """Handle tab selection from TOP."""
        safe_name = child_item.name.encode('ascii', errors='replace').decode('ascii') if child_item else None
        print(f"[DEBUG] NotesView.on_child_tab_selected() called with child: {safe_name}")
        self.current_item = child_item

        # Display the item in canvas
        if child_item:
            safe_name = child_item.name.encode('ascii', errors='replace').decode('ascii')
            print(f"[DEBUG]   Calling notes_canvas.set_item() for child: {safe_name}")
            self.notes_canvas.set_item(child_item)
        else:
            print(f"[DEBUG]   Child item is None, not calling set_item")

        # Refresh indicators to show current status
        self.items_list.update_indicators()
        self.child_tabs.update_indicators()

    def on_note_added(self, note):
        """Handle new note added."""
        if self.current_item:
            self.current_item.add_note(note)
            # Update indicator in Gantt view if needed
            if hasattr(self.main_window, 'update_view'):
                self.main_window.update_view()
            # Save notes structure
            if hasattr(self.main_window, 'save_notes_structure'):
                self.main_window.save_notes_structure()

    def on_item_moved(self, item, new_level):
        """Handle item moved between levels."""
        item.level = new_level
        # Refresh the display
        if self.current_project:
            self.on_project_selected(self.current_project)

    # Note: toggle_view_mode removed - checkbox now only controls note_type for new notes
    # The NotesCanvas checks grid_toggle_checkbox.isChecked() directly when creating new notes


class ProjectListWidget(QListWidget):
    """LEFT column - List of projects."""

    project_selected = pyqtSignal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.projects = []
        self.setStyleSheet("""
            QListWidget {
                background-color: #F7F3FA;
                border: none;
                border-right: 2px solid #7719AA;
                font-size: 13px;
                font-family: 'Calibri', sans-serif;
            }
            QListWidget::item {
                padding: 10px 12px;
                border-bottom: 1px solid #EDE7F3;
            }
            QListWidget::item:selected {
                background-color: #7719AA;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #EDE7F3;
            }
        """)
        self.itemClicked.connect(self.on_item_clicked)

    def set_projects(self, projects):
        """Set the list of projects."""
        self.clear()
        self.projects = projects
        for project in projects:
            item = QListWidgetItem(project.name)
            item.setData(Qt.ItemDataRole.UserRole, project)
            self.addItem(item)

    def on_item_clicked(self, list_item):
        """Emit signal when project is clicked."""
        project = list_item.data(Qt.ItemDataRole.UserRole)
        self.project_selected.emit(project)


class ItemsListWidget(QListWidget):
    """MIDDLE column - List of items (PRs, tasks, custom pages)."""

    item_selected = pyqtSignal(object)
    item_moved = pyqtSignal(object, int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.notes_view = parent  # Store reference to NotesView
        self.items = []
        self.setStyleSheet("""
            QListWidget {
                background-color: #FAFAFA;
                border: none;
                border-right: 1px solid #E0D8E8;
                font-size: 12px;
                font-family: 'Calibri', sans-serif;
            }
            QListWidget::item {
                padding: 8px 10px;
                border-bottom: 1px solid #F0F0F0;
            }
            QListWidget::item:selected {
                background-color: #F3EDF8;
                color: #1A1A2E;
                border-left: 3px solid #7719AA;
            }
            QListWidget::item:hover {
                background-color: #F7F3FA;
            }
        """)
        self.setDragDropMode(QListWidget.DragDropMode.InternalMove)
        self.itemClicked.connect(self.on_item_clicked)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    def set_items(self, items):
        """Set the list of items."""
        print(f"[DEBUG] ItemsListWidget.set_items() called with {len(items)} items")
        self.clear()
        self.items = items
        for i, item in enumerate(items):
            safe_name = item.name.encode('ascii', errors='replace').decode('ascii')
            # Add icon if item has notes
            if item.notes:
                display_name = f"📝 {item.name}"
                print(f"[DEBUG]   Adding item {i}: {safe_name} WITH notes ({len(item.notes)} notes)")
            else:
                display_name = item.name
                print(f"[DEBUG]   Adding item {i}: {safe_name} WITHOUT notes")

            list_item = QListWidgetItem(display_name)
            list_item.setData(Qt.ItemDataRole.UserRole, item)
            self.addItem(list_item)
        print(f"[DEBUG] ItemsListWidget now has {self.count()} items displayed")

    def on_item_clicked(self, list_item):
        """Emit signal when item is clicked."""
        item = list_item.data(Qt.ItemDataRole.UserRole)
        self.item_selected.emit(item)

    def show_context_menu(self, position):
        """Show context menu for item operations."""
        menu = QMenu(self)

        # Add custom page
        add_action = menu.addAction("Add Custom Page")
        add_action.triggered.connect(self.add_custom_page)

        # If an item is selected, show more options
        current = self.currentItem()
        if current:
            item = current.data(Qt.ItemDataRole.UserRole)
            menu.addSeparator()

            # Move to level options
            if item.level == 3:
                move_to_2 = menu.addAction("Move to Items List (Level 2)")
                move_to_2.triggered.connect(lambda: self.move_item_level(item, 2))

            # Delete
            delete_action = menu.addAction("Delete")
            delete_action.triggered.connect(lambda: self.delete_item(item))

        menu.exec(self.mapToGlobal(position))

    def add_custom_page(self):
        """Add a custom page."""
        name, ok = QInputDialog.getText(self, "Add Custom Page", "Page name:")
        if ok and name:
            # Ask if it should be a Gantt task too
            reply = QMessageBox.question(
                self,
                "Create as Task?",
                "Do you want to create this as a task in the Gantt chart too?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            custom_item = NotesItem(name, is_custom=True, level=2)
            self.items.append(custom_item)

            if reply == QMessageBox.StandardButton.Yes:
                # Signal to create task in Gantt (implementation needed in MainWindow)
                pass

            # Refresh display
            self.set_items(self.items)

            # Auto-select the newly created item
            for i in range(self.count()):
                if self.item(i).data(Qt.ItemDataRole.UserRole) == custom_item:
                    self.setCurrentRow(i)
                    self.item_selected.emit(custom_item)
                    break

    def move_item_level(self, item, new_level):
        """Move item to a different level."""
        self.item_moved.emit(item, new_level)

    def delete_item(self, item):
        """Delete an item."""
        if item.linked_task:
            # Ask if delete from Gantt too
            reply = QMessageBox.question(
                self,
                "Delete Item",
                f"Delete '{item.name}' from notes only, or from Gantt chart too?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )

            if reply == QMessageBox.StandardButton.Cancel:
                return
            elif reply == QMessageBox.StandardButton.Yes:
                # Delete from Gantt too (implementation needed)
                pass

        # Delete from notes
        if item in self.items:
            self.items.remove(item)
            self.set_items(self.items)

    def update_indicators(self):
        """Update red indicators for items with overdue notes."""
        for i in range(self.count()):
            list_item = self.item(i)
            item = list_item.data(Qt.ItemDataRole.UserRole)

            # Check if item has overdue notes or overdue children
            has_overdue = item.has_overdue_notes() or item.has_overdue_children()

            # Update display with indicators
            display_text = item.name
            if has_overdue:
                display_text = f"🔴 {display_text}"
            elif item.notes:
                display_text = f"📝 {display_text}"

            list_item.setText(display_text)


class ChildTabWidget(QTabWidget):
    """TOP tabs - Children of selected item."""

    tab_selected = pyqtSignal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.children = []
        self.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #F9F7FB;
                padding: 6px 16px;
                margin-right: 2px;
                border: none;
                border-bottom: 2px solid transparent;
                font-family: 'Calibri', sans-serif;
                font-size: 12px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #7719AA;
                font-weight: bold;
            }
            QTabBar::tab:hover {
                background-color: #EDE7F3;
            }
        """)
        self.currentChanged.connect(self.on_tab_changed)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    def set_children(self, children):
        """Set the child items as tabs."""
        print(f"[DEBUG] ChildTabsBar.set_children() called with {len(children)} children")
        for i, child in enumerate(children):
            # Safely encode child name to avoid Unicode errors
            safe_name = child.name.encode('ascii', errors='replace').decode('ascii')
            print(f"[DEBUG]   Child {i}: {safe_name}, id={id(child)}, notes={len(child.notes)}")
        self.clear()
        self.children = children

        if not children:
            # No children, show single "Notes" tab
            placeholder = QWidget()
            self.addTab(placeholder, "Notes")
        else:
            for child in children:
                placeholder = QWidget()
                tab_name = child.name
                if child.notes:
                    tab_name = f"📝 {tab_name}"
                self.addTab(placeholder, tab_name)

    def on_tab_changed(self, index):
        """Emit signal when tab is selected."""
        if 0 <= index < len(self.children):
            selected_child = self.children[index]
            # Safely encode child name to avoid Unicode errors
            safe_name = selected_child.name.encode('ascii', errors='replace').decode('ascii')
            print(f"[DEBUG] ChildTabsBar.on_tab_changed() index={index}, emitting: {safe_name}, id={id(selected_child)}, notes={len(selected_child.notes)}")
            self.tab_selected.emit(selected_child)

    def show_context_menu(self, position):
        """Show context menu for tab operations."""
        index = self.tabBar().tabAt(position)
        if index >= 0 and index < len(self.children):
            child = self.children[index]

            menu = QMenu(self)
            move_action = menu.addAction("Move to Items List")
            move_action.triggered.connect(lambda: self.move_to_items(child))
            menu.exec(self.mapToGlobal(position))

    def move_to_items(self, child):
        """Move a child from tabs to items list."""
        # Signal parent to handle move
        parent_widget = self.parent()
        if hasattr(parent_widget, 'on_item_moved'):
            parent_widget.on_item_moved(child, 2)

    def update_indicators(self):
        """Update red indicators for tabs with overdue notes."""
        for i in range(self.count()):
            if i < len(self.children):
                child = self.children[i]

                # Check if child has overdue notes
                has_overdue = child.has_overdue_notes()

                # Update tab text with indicators
                tab_name = child.name
                if has_overdue:
                    tab_name = f"🔴 {tab_name}"
                elif child.notes:
                    tab_name = f"📝 {tab_name}"

                self.setTabText(i, tab_name)


# ============================================================================
# ONENOTE-STYLE CANVAS SYSTEM
# ============================================================================

class TimelineEventFilter(QObject):
    """Event filter for timeline field to parse natural language dates."""
    def __init__(self, note_container):
        super().__init__()
        self.note_container = note_container

    def eventFilter(self, obj, event):
        """Parse timeline on Enter key or focus loss."""
        if obj == self.note_container.timeline_edit:
            if event.type() == QEvent.Type.KeyPress:
                if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                    # Parse timeline when Enter is pressed
                    self.parse_and_update_timeline()
                    event.accept()
                    return True
            elif event.type() == QEvent.Type.FocusOut:
                # Parse timeline when field loses focus
                self.parse_and_update_timeline()
        return super().eventFilter(obj, event)

    def parse_and_update_timeline(self):
        """Parse the timeline text and update the note's reminder_date."""
        try:
            timeline_text = self.note_container.timeline_edit.text().strip()

            if not timeline_text:
                # Clear reminder if field is empty
                if self.note_container.note.reminder_date:
                    self.note_container.note.reminder_date = None
                    self.note_container.note.update_status()
                    self.note_container.update()
                    if self.note_container.parent_canvas:
                        self.note_container.parent_canvas.update_status_indicators()
                        self.note_container.parent_canvas.save_notes()
            else:
                # Try to parse natural language date
                reminder_date, _ = parse_natural_date(timeline_text)
                if reminder_date:
                    # Always update reminder and status when parsing succeeds
                    # (removed condition check - datetime comparison can fail due to time differences)
                    self.note_container.note.reminder_date = reminder_date
                    self.note_container.note.update_status()
                    self.note_container.update()
                    # Update display to show formatted date
                    # Temporarily remove event filter to prevent re-triggering when setText() is called
                    self.note_container.timeline_edit.removeEventFilter(self.note_container.timeline_filter)
                    self.note_container.timeline_edit.setText(reminder_date.strftime("%b %d"))
                    self.note_container.timeline_edit.installEventFilter(self.note_container.timeline_filter)
                    if self.note_container.parent_canvas:
                        self.note_container.parent_canvas.update_status_indicators()
                        self.note_container.parent_canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error parsing timeline: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("TIMELINE PARSE ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")


class NoteEventFilter(QObject):
    """Event filter for NoteContainer to handle Enter key and Delete key events."""
    def __init__(self, note_container):
        super().__init__()
        self.note_container = note_container

    def eventFilter(self, obj, event):
        """Filter events for date detection on Enter key and deletion on Delete key."""
        if obj == self.note_container.content_widget:
            if event.type() == QEvent.Type.KeyPress:
                if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                    # Trigger date detection when Enter is pressed
                    text = self.note_container.content_widget.toPlainText()
                    reminder_date, _ = parse_natural_date(text)
                    if reminder_date:
                        self.note_container.note.reminder_date = reminder_date
                        # Don't clean text automatically - just set the date
                        self.note_container.note.update_status()
                        self.note_container.update()
                        # Trigger cascade update and save
                        if self.note_container.parent_canvas:
                            self.note_container.parent_canvas.update_status_indicators()
                            self.note_container.parent_canvas.save_notes()
                elif event.key() == Qt.Key.Key_Delete or event.key() == Qt.Key.Key_Backspace:
                    # Check if the text editor is empty and cursor is at the start
                    text = self.note_container.content_widget.toPlainText().strip()
                    cursor = self.note_container.content_widget.textCursor()

                    # Delete the note container if:
                    # 1. The note is completely empty, OR
                    # 2. For Delete key: cursor is at the very end and text is empty
                    # 3. For Backspace key: cursor is at the very start and text is empty
                    should_delete = False

                    if not text:  # Completely empty
                        if event.key() == Qt.Key.Key_Delete and cursor.atEnd():
                            should_delete = True
                        elif event.key() == Qt.Key.Key_Backspace and cursor.atStart():
                            should_delete = True

                    if should_delete:
                        # Delete the note container
                        try:
                            canvas = self.note_container.parent_canvas

                            # Remove from parent canvas containers list
                            if self.note_container in canvas.containers:
                                canvas.containers.remove(self.note_container)

                            # Remove from pending containers if present
                            if self.note_container in canvas.pending_containers:
                                canvas.pending_containers.discard(self.note_container)

                            # Remove from current item's notes
                            if canvas.current_item and self.note_container.note in canvas.current_item.notes:
                                canvas.current_item.notes.remove(self.note_container.note)

                            # Remove from scene
                            if self.note_container.scene():
                                self.note_container.scene().removeItem(self.note_container)

                            # Save changes
                            canvas.save_notes()

                            # Accept the event to prevent default behavior
                            event.accept()
                            return True
                        except Exception as e:
                            import traceback
                            error_msg = f"Error deleting note container: {e}\n\n{traceback.format_exc()}"
                            print(f"\n{'='*60}")
                            print("DELETE NOTE CONTAINER ERROR")
                            print(f"{'='*60}")
                            print(error_msg)
                            print(f"{'='*60}\n")
            elif event.type() == QEvent.Type.FocusIn:
                # Connect formatting toolbar to this editor
                try:
                    canvas = self.note_container.parent_canvas
                    if canvas and hasattr(canvas, 'notes_view_parent') and canvas.notes_view_parent:
                        nv = canvas.notes_view_parent
                        if hasattr(nv, 'formatting_toolbar') and isinstance(obj, QTextEdit):
                            nv.formatting_toolbar.set_editor(obj)
                    # Show header when note gets focus (OneNote behavior)
                    if hasattr(self.note_container, 'header_proxy'):
                        self.note_container.header_proxy.setVisible(True)
                        if hasattr(self.note_container, '_header_hide_timer'):
                            self.note_container._header_hide_timer.stop()
                except Exception:
                    pass
            elif event.type() == QEvent.Type.FocusOut:
                # Commit or discard pending note on focus loss
                try:
                    if self.note_container.is_pending:
                        # Use is_note_empty() helper to check if note is empty
                        if self.note_container.is_note_empty():
                            self.note_container.discard_if_empty()
                        else:
                            self.note_container.commit_pending_note()
                    else:
                        # Not pending, just commit (commit_container will check if empty)
                        self.note_container.commit_pending_note()
                except Exception as e:
                    import traceback
                    error_msg = f"Error in FocusOut handler: {e}\n\n{traceback.format_exc()}"
                    print(f"\n{'='*60}")
                    print("FOCUS OUT ERROR")
                    print(f"{'='*60}")
                    print(error_msg)
                    print(f"{'='*60}\n")
        return super().eventFilter(obj, event)


class GridEventFilter(QObject):
    """Event filter for grid widgets to handle Excel-like keyboard navigation."""
    def __init__(self, note_container):
        super().__init__()
        self.note_container = note_container

    def _is_descendant(self, widget, ancestor):
        """Check if widget is a descendant of ancestor."""
        if not widget or not ancestor:
            return False
        parent = widget.parent()
        while parent:
            if parent == ancestor:
                return True
            parent = parent.parent()
        return False

    def _check_and_discard_if_needed(self, table):
        """Helper method to check if note should be discarded and do it."""
        try:
            # Final check - is table being edited or focus on child?
            current_focus = QApplication.focusWidget()
            if current_focus and self._is_descendant(current_focus, table):
                return  # Focus is on child, don't discard
            if table.state() == QAbstractItemView.State.EditingState:
                return  # Table is being edited, don't discard
            
            # Now check if note should be discarded
            if self.note_container.is_pending:
                if self.note_container.is_note_empty():
                    self.note_container.discard_if_empty()
                else:
                    self.note_container.commit_pending_note()
            else:
                self.note_container.commit_pending_note()
        except Exception as e:
            # Silently ignore errors in delayed check
            pass

    def eventFilter(self, obj, event):
        """Filter events for Excel-like keyboard navigation."""
        table = self.note_container.content_widget
        
        # Check if event is from the table widget or any of its descendants
        # This includes the table itself, cell editors, and other child widgets
        is_table_event = (obj == table)
        if not is_table_event and hasattr(obj, 'parent'):
            # Check if obj is a descendant of the table
            parent = obj.parent()
            while parent:
                if parent == table:
                    is_table_event = True
                    break
                parent = parent.parent()
        
        if is_table_event:
            if event.type() == QEvent.Type.KeyPress:
                key = event.key()
                modifiers = event.modifiers()
                
                # Delete/Backspace: Clear selected cells
                if (key == Qt.Key.Key_Delete or key == Qt.Key.Key_Backspace) and modifiers == Qt.KeyboardModifier.NoModifier:
                    selected_items = table.selectedItems()
                    if selected_items:
                        # Clear all selected cells
                        for item in selected_items:
                            if item:
                                item.setText('')
                                # Update grid_data
                                row = item.row()
                                col = item.column()
                                if (row, col) in self.note_container.note.grid_data:
                                    del self.note_container.note.grid_data[(row, col)]
                        # Save changes
                        if self.note_container.parent_canvas:
                            self.note_container.parent_canvas.save_notes()
                        return True
                    # If no selection, check if current cell should be cleared
                    current_item = table.currentItem()
                    if current_item:
                        current_item.setText('')
                        row = current_item.row()
                        col = current_item.column()
                        if (row, col) in self.note_container.note.grid_data:
                            del self.note_container.note.grid_data[(row, col)]
                        if self.note_container.parent_canvas:
                            self.note_container.parent_canvas.save_notes()
                        return True
                
                # Enter: Move down to next row (same column)
                if (key == Qt.Key.Key_Return or key == Qt.Key.Key_Enter) and modifiers == Qt.KeyboardModifier.NoModifier:
                    # Get current position (may be -1 if no selection)
                    current_row = table.currentRow()
                    current_col = table.currentColumn()
                    
                    # If no current cell, start at (0, 0)
                    if current_row < 0:
                        current_row = 0
                    if current_col < 0:
                        current_col = 0
                    
                    # Close editor if open (this commits the data)
                    # setCurrentCell will automatically close the editor and commit data
                    # We just need to make sure we have the correct current position
                    if table.state() == QAbstractItemView.State.EditingState:
                        # Get the current cell position before closing editor
                        # The editor is already handling the commit
                        pass
                    
                    # Move to next row
                    next_row = current_row + 1
                    if next_row >= table.rowCount():
                        # Add more rows if needed
                        old_row_count = table.rowCount()
                        new_row_count = old_row_count + 10
                        table.setRowCount(new_row_count)
                        # Initialize new cells
                        for row in range(old_row_count, new_row_count):
                            for col in range(table.columnCount()):
                                if not table.item(row, col):
                                    table.setItem(row, col, QTableWidgetItem(''))
                    
                    # Ensure the target cell exists
                    target_item = table.item(next_row, current_col)
                    if not target_item:
                        target_item = QTableWidgetItem('')
                        table.setItem(next_row, current_col, target_item)
                    
                    # Move to next row, same column
                    table.setCurrentCell(next_row, current_col)
                    # Start editing the new cell after a short delay
                    QTimer.singleShot(10, lambda r=next_row, c=current_col: table.editItem(table.item(r, c)))
                    return True
                
                # Shift+Enter: Move up to previous row (same column)
                elif (key == Qt.Key.Key_Return or key == Qt.Key.Key_Enter) and modifiers == Qt.KeyboardModifier.ShiftModifier:
                    # Get current position (may be -1 if no selection)
                    current_row = table.currentRow()
                    current_col = table.currentColumn()
                    
                    # If no current cell, start at (0, 0)
                    if current_row < 0:
                        current_row = 0
                    if current_col < 0:
                        current_col = 0
                    
                    # Close editor if open (this commits the data)
                    # setCurrentCell will automatically close the editor and commit data
                    # We just need to make sure we have the correct current position
                    if table.state() == QAbstractItemView.State.EditingState:
                        # Get the current cell position before closing editor
                        # The editor is already handling the commit
                        pass
                    
                    # Move to previous row (don't go below 0)
                    next_row = max(0, current_row - 1)
                    
                    # Ensure the target cell exists
                    target_item = table.item(next_row, current_col)
                    if not target_item:
                        target_item = QTableWidgetItem('')
                        table.setItem(next_row, current_col, target_item)
                    
                    # Move to previous row, same column
                    table.setCurrentCell(next_row, current_col)
                    # Start editing the new cell after a short delay
                    QTimer.singleShot(10, lambda r=next_row, c=current_col: table.editItem(table.item(r, c)))
                    return True
                
                # Delete/Backspace: Clear selected cells or delete note if grid is empty
                elif key == Qt.Key.Key_Delete or key == Qt.Key.Key_Backspace:
                    # First, check if there are selected cells to clear
                    selected_items = table.selectedItems()
                    if selected_items:
                        # Clear all selected cells
                        for item in selected_items:
                            if item:
                                item.setText('')
                                # Update grid_data
                                row = item.row()
                                col = item.column()
                                if (row, col) in self.note_container.note.grid_data:
                                    del self.note_container.note.grid_data[(row, col)]
                        # Save changes
                        if self.note_container.parent_canvas:
                            self.note_container.parent_canvas.save_notes()
                        return True
                    
                    # If no selection, check if current cell should be cleared
                    current_item = table.currentItem()
                    if current_item and current_item.text():
                        current_item.setText('')
                        row = current_item.row()
                        col = current_item.column()
                        if (row, col) in self.note_container.note.grid_data:
                            del self.note_container.note.grid_data[(row, col)]
                        if self.note_container.parent_canvas:
                            self.note_container.parent_canvas.save_notes()
                        return True
                    
                    # Don't auto-delete note - user must use delete button
                    # This prevents issues when clicking on blank canvas
                    return True
            
            elif event.type() == QEvent.Type.FocusOut:
                # Commit or discard pending grid note on focus loss
                # But only if focus is truly leaving the table (not moving to cell editor)
                try:
                    focus_event = event
                    if not isinstance(focus_event, QFocusEvent):
                        return False
                    
                    # Don't discard if table is currently being edited
                    if table.state() == QAbstractItemView.State.EditingState:
                        # Table is being edited, focus might be moving to cell editor
                        # Don't discard yet - wait until editing is complete
                        return False  # Let event pass through, don't handle it here
                    
                    # Check if focus is moving to a child widget (cell editor or other child)
                    # Use a delayed check because focus widget might not be set yet when FocusOut fires
                    focus_widget = QApplication.focusWidget()
                    if focus_widget and self._is_descendant(focus_widget, table):
                        # Focus is moving to a child widget (e.g., cell editor)
                        # Don't discard - user is still interacting with the table
                        return False  # Let event pass through
                    
                    # Check if focus is moving to a child of the table by looking at next focus widget
                    # When editing starts, the editor widget might not be in focus yet, but we can check
                    # if the event reason suggests editing is about to start
                    reason = focus_event.reason()
                    # For mouse clicks, always delay check to allow:
                    # 1. Cell to be selected (if clicking on table)
                    # 2. Editor to be created (if editing starts)
                    # 3. User to start typing
                    if reason == Qt.FocusReason.MouseFocusReason:
                        # Mouse click - delay check to allow cell selection and editor creation
                        # This is critical because clicking on a cell causes FocusOut before editing starts
                        # Use a single delayed check with sufficient time for editing to start
                        def delayed_check():
                            # Check if editing has started or if focus is on a child
                            current_focus = QApplication.focusWidget()
                            if current_focus and self._is_descendant(current_focus, table):
                                # Focus is on a child (editor), editing has started - don't discard
                                return
                            if table.state() == QAbstractItemView.State.EditingState:
                                # Table is being edited - don't discard
                                return
                            # Check if a cell is selected - if so, user clicked on table
                            # If cell is selected, give more time for editing to start
                            if table.currentRow() >= 0 and table.currentColumn() >= 0:
                                # Cell is selected - user clicked on table to edit
                                # Wait longer to see if editing starts (user might be about to type)
                                def second_check():
                                    # Check again if editing has started
                                    current_focus = QApplication.focusWidget()
                                    if current_focus and self._is_descendant(current_focus, table):
                                        return  # Editing started, don't discard
                                    if table.state() == QAbstractItemView.State.EditingState:
                                        return  # Editing started, don't discard
                                    # Cell is still selected but no editing - might be single click
                                    # If cell is selected, don't discard immediately
                                    # User might double-click to edit or press a key to edit
                                    # Only discard if user explicitly leaves the note
                                    # For now, if cell is selected, don't auto-discard
                                    # (User can still delete with Delete/Backspace)
                                    return  # Don't auto-discard if cell is selected
                                
                                # Wait 300ms for editing to start if cell is selected
                                QTimer.singleShot(300, second_check)
                                return
                            
                            # No cell selected and not editing - user likely clicked outside
                            # Only commit if note has content, don't auto-discard
                            # User must use delete button to delete note
                            if self.note_container.is_pending:
                                if not self.note_container.is_note_empty():
                                    self.note_container.commit_pending_note()
                            else:
                                # Not pending, just commit changes
                                self.note_container.commit_pending_note()
                        
                        # Delay by 200ms to allow cell selection and editor creation
                        QTimer.singleShot(200, delayed_check)
                        return False  # Don't handle immediately - let editing start
                    
                    # For other focus reasons (Tab, keyboard, etc.), check immediately
                    # But still verify focus isn't moving to a child
                    # Focus is truly leaving the table and all its descendants
                    # Only commit if note has content, don't auto-discard
                    # User must use delete button to delete note
                    if self.note_container.is_pending:
                        if not self.note_container.is_note_empty():
                            self.note_container.commit_pending_note()
                    else:
                        # Not pending, just commit changes
                        self.note_container.commit_pending_note()
                except Exception as e:
                    import traceback
                    error_msg = f"Error in GridEventFilter FocusOut handler: {e}\n\n{traceback.format_exc()}"
                    print(f"\n{'='*60}")
                    print("GRID FOCUS OUT ERROR")
                    print(f"{'='*60}")
                    print(error_msg)
                    print(f"{'='*60}\n")
                
        return super().eventFilter(obj, event)


class NoteContainer(QGraphicsItem):
    """
    A draggable, resizable note container that holds text.
    Represents an individual note box on the canvas.
    """

    def __init__(self, note, parent_canvas):
        debug_log(f"NoteContainer.__init__ ENTRY: note.uid={getattr(note, 'uid', 'NO_UID')}, note.text='{getattr(note, 'text', 'NO_TEXT')[:30]}...', parent_canvas={type(parent_canvas).__name__}")
        try:
            super().__init__()
            debug_log("  - super().__init__() OK")

            self.note = note
            self.parent_canvas = parent_canvas
            self.is_pending = False
            debug_log(f"  - Basic attributes set")

            # Make item movable and selectable
            self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
            self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
            self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges)
            debug_log(f"  - Flags set")

            # Set position from note data
            self.setPos(note.x_pos, note.y_pos)
            debug_log(f"  - Position set: ({note.x_pos}, {note.y_pos})")

            # Border margin — tight padding for OneNote-like floating text
            self.border_margin = 3

            # Header height constant — compact
            self.HEADER_HEIGHT = 28

            # Create header widget (date written | heading | timeline)
            debug_log(f"  - Creating header widget...")
            self.header_container = QWidget()
            self.header_container.setFixedHeight(self.HEADER_HEIGHT)
            self.header_container.setStyleSheet("background: transparent; border: none;")
            header_layout = QHBoxLayout(self.header_container)
            header_layout.setContentsMargins(4, 2, 4, 2)
            header_layout.setSpacing(8)

            # Left: Date written (read-only but copiable)
            self.date_written_edit = QLineEdit()
            self.date_written_edit.setText(note.timestamp.strftime("%b %d, %Y"))
            self.date_written_edit.setReadOnly(True)
            self.date_written_edit.setFixedWidth(95)
            self.date_written_edit.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    color: #6B6B80;
                    font-size: 9pt;
                    font-family: 'Calibri', sans-serif;
                    padding: 3px;
                }
                QLineEdit:focus {
                    background-color: #F0E6F6;
                    border: 1px solid #E0D8E8;
                }
            """)
            header_layout.addWidget(self.date_written_edit)

            # Middle: Heading (editable, optional)
            self.heading_edit = QLineEdit()
            self.heading_edit.setPlaceholderText("Add a title...")
            self.heading_edit.setText(note.heading)
            self.heading_edit.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    font-size: 11pt;
                    font-weight: bold;
                    font-family: 'Calibri', sans-serif;
                    color: #1A1A2E;
                    padding: 3px;
                }
                QLineEdit:focus {
                    border: 1px solid #7719AA;
                }
            """)
            header_layout.addWidget(self.heading_edit, stretch=1)

            # Right: Timeline (editable, free-form natural language)
            self.timeline_edit = QLineEdit()
            self.timeline_edit.setPlaceholderText("e.g., tomorrow")
            self.timeline_edit.setFixedWidth(120)
            # Pre-fill with existing reminder if present
            if note.reminder_date:
                self.timeline_edit.setText(note.reminder_date.strftime("%b %d"))
            self.timeline_edit.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    font-size: 9pt;
                    font-family: 'Calibri', sans-serif;
                    color: #666;
                    padding: 3px;
                }
                QLineEdit:focus {
                    border: 1px solid #0078d4;
                }
            """)
            header_layout.addWidget(self.timeline_edit)

            # Create proxy widget for header
            self.header_proxy = QGraphicsProxyWidget(self)
            self.header_proxy.setWidget(self.header_container)
            self.header_proxy.setPos(self.border_margin, self.border_margin)

            # CRITICAL: Set header width to fit within borders
            header_width = max(50, note.width - self.border_margin * 2)
            self.header_container.setFixedWidth(header_width)

            debug_log(f"  - Header widget created with width: {header_width}")

            # Wire up heading field
            self.heading_edit.textChanged.connect(self.on_heading_changed)

            # Wire up timeline field (parse on Enter or focus loss)
            self.timeline_filter = TimelineEventFilter(self)
            self.timeline_edit.installEventFilter(self.timeline_filter)

            # Create content widget based on note type
            print(f"[DEBUG] NoteContainer: note.note_type = '{note.note_type}', grid_data entries = {len(note.grid_data)}")
            if note.grid_data:
                print(f"[DEBUG] NoteContainer: grid_data sample: {list(note.grid_data.items())[:3]}")

            if note.note_type == "grid":
                print(f"[DEBUG] NoteContainer: CREATING GRID WIDGET")
                debug_log(f"  - Creating grid content widget...")
                self.content_widget = self.create_grid_widget()
            else:
                print(f"[DEBUG] NoteContainer: Creating text widget (not grid)")
                debug_log(f"  - Creating canvas (text) content widget...")
                self.content_widget = self.create_text_widget()

            debug_log(f"  - Content widget configured")

            # Proxy widget to embed content widget (positioned below header)
            debug_log(f"  - Creating proxy widget...")
            self.proxy = QGraphicsProxyWidget(self)
            self.proxy.setWidget(self.content_widget)
            self.proxy.setPos(self.border_margin, self.border_margin + self.HEADER_HEIGHT)
            debug_log(f"  - Proxy widget created and positioned below header")

            # Connect signals based on type
            debug_log(f"  - Connecting signals...")
            if note.note_type == "grid":
                self.content_widget.cellChanged.connect(self.on_grid_changed)
            else:
                self.content_widget.textChanged.connect(self.on_text_changed)
                self.content_widget.textChanged.connect(self.adjust_size_to_content)
                # Connect to formatting toolbar when this editor gets focus
                self.content_widget.installEventFilter(self)
            debug_log(f"  - Signals connected")

            # Set initial size - use saved dimensions first, then adjust after widget is shown
            # This prevents incorrect sizing when switching tabs (widget not fully rendered yet)
            debug_log(f"  - Setting initial size (saved: {self.note.width}x{self.note.height})...")
            if self.note.width > 0 and self.note.height > 0:
                # For grid notes, check if size is too small (old notes) and upgrade to larger size
                if note.note_type == "grid":
                    # Minimum reasonable size for grid notes: 600x400 (content)
                    min_grid_content_width = 600
                    min_grid_content_height = 400
                    min_grid_total_width = min_grid_content_width + self.border_margin * 2
                    min_grid_total_height = min_grid_content_height + self.HEADER_HEIGHT + self.border_margin * 2
                    
                    if self.note.width < min_grid_total_width or self.note.height < min_grid_total_height:
                        # Upgrade old small grid notes to larger default size
                        debug_log(f"  - Upgrading small grid note size from {self.note.width}x{self.note.height}")
                        grid_content_width = 700
                        grid_content_height = 500
                        self.note.width = grid_content_width + self.border_margin * 2
                        self.note.height = grid_content_height + self.HEADER_HEIGHT + self.border_margin * 2
                        debug_log(f"  - Upgraded to: {self.note.width}x{self.note.height}")
                
                # Use saved dimensions (or upgraded dimensions for grid notes)
                content_width = self.note.width - self.border_margin * 2
                content_height = self.note.height - self.border_margin * 2 - self.HEADER_HEIGHT
                self.content_widget.setFixedSize(max(50, content_width), max(30, content_height))
                self.proxy.resize(max(50, content_width), max(30, content_height))
                debug_log(f"  - Using size: {self.note.width}x{self.note.height} (content: {content_width}x{content_height})")
            else:
                # No saved size - calculate now (only for text widgets)
                debug_log(f"  - No saved size...")
                if note.note_type == "canvas":
                    self.adjust_size_to_content()
                else:
                    # Grid notes have larger default size for better visibility
                    # Content size: 700x500
                    grid_content_width = 700
                    grid_content_height = 500
                    self.content_widget.setFixedSize(grid_content_width, grid_content_height)
                    self.proxy.resize(grid_content_width, grid_content_height)
                    # Update note dimensions to match (content + header + borders)
                    note.width = grid_content_width + self.border_margin * 2
                    note.height = grid_content_height + self.HEADER_HEIGHT + self.border_margin * 2
                    debug_log(f"  - Grid note size set: {note.width}x{note.height}")

            # Schedule deferred size adjustment after widget is shown/rendered (text only)
            if note.note_type == "canvas":
                debug_log(f"  - Scheduling deferred size adjustment...")
                QTimer.singleShot(50, lambda: self._safe_adjust_size())

            # Install event filter for Enter key detection (text only) or keyboard navigation (grid)
            if note.note_type == "canvas":
                debug_log(f"  - Installing event filter for text notes...")
                self.event_filter = NoteEventFilter(self)
                self.content_widget.installEventFilter(self.event_filter)
                debug_log(f"  - Event filter installed")
            elif note.note_type == "grid":
                debug_log(f"  - Installing event filter for grid notes...")
                self.grid_event_filter = GridEventFilter(self)
                self.content_widget.installEventFilter(self.grid_event_filter)
                debug_log(f"  - Grid event filter installed")

            # Checkbox for complete toggle (positioned below header in text area)
            debug_log(f"  - Creating checkbox...")
            self.checkbox = QCheckBox()
            self.checkbox.setChecked(note.is_completed)
            self.checkbox.stateChanged.connect(self.on_checkbox_toggled)
            self.checkbox_proxy = QGraphicsProxyWidget(self)
            self.checkbox_proxy.setWidget(self.checkbox)
            self.checkbox_proxy.setPos(self.border_margin + 3, self.border_margin + self.HEADER_HEIGHT + 3)
            debug_log(f"  - Checkbox created")

            # HIDE checkbox by default - only show on hover
            self.checkbox_proxy.setVisible(False)
            
            # Delete button (positioned in top-right corner, shown on hover)
            debug_log(f"  - Creating delete button...")
            from PyQt6.QtWidgets import QPushButton
            self.delete_button = QPushButton("×")
            self.delete_button.setFixedSize(24, 24)
            self.delete_button.setStyleSheet("""
                QPushButton {
                    background-color: #ff4444;
                    color: white;
                    border: none;
                    border-radius: 12px;
                    font-size: 18px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #ff6666;
                }
                QPushButton:pressed {
                    background-color: #cc0000;
                }
            """)
            self.delete_button.clicked.connect(self.delete_note)
            self.delete_button_proxy = QGraphicsProxyWidget(self)
            self.delete_button_proxy.setWidget(self.delete_button)
            # Position in top-right corner
            delete_button_x = self.note.width - self.border_margin - 24
            delete_button_y = self.border_margin + 3
            self.delete_button_proxy.setPos(delete_button_x, delete_button_y)
            debug_log(f"  - Delete button created at ({delete_button_x}, {delete_button_y})")
            
            # HIDE delete button by default - only show on hover
            self.delete_button_proxy.setVisible(False)

            # State
            self.is_hovered = False
            self.setAcceptHoverEvents(True)
            debug_log(f"  - Hover events enabled")

            # OneNote-style: hide header by default, show on hover/focus
            # Exception: if note has a title, keep header visible
            has_title = bool(note.heading and note.heading.strip())
            self.header_proxy.setVisible(has_title)
            self._header_hide_timer = QTimer()
            self._header_hide_timer.setSingleShot(True)
            self._header_hide_timer.setInterval(300)
            self._header_hide_timer.timeout.connect(self._maybe_hide_header)

            # Update status
            debug_log(f"  - Updating note status...")
            self.note.update_status()
            debug_log(f"NoteContainer.__init__ SUCCESS: note.uid={note.uid}, status={self.note.status}")

        except Exception as e:
            debug_log(f"NoteContainer.__init__ FAILED: {e}", "ERROR")
            raise

    def create_text_widget(self):
        """Create a rich text editor widget for canvas mode."""
        text_edit = QTextEdit()
        text_edit.setAcceptRichText(True)

        # Backward compat: load plain text or HTML
        text = self.note.text
        if text and text.strip().startswith('<'):
            text_edit.setHtml(text)
        else:
            text_edit.setPlainText(text if text else '')

        # Disable scrollbars
        text_edit.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        text_edit.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        text_edit.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border: none;
                padding: 2px 4px;
                font-size: 11pt;
                font-family: 'Calibri', sans-serif;
                color: #1A1A2E;
            }
        """)
        # Set document default font so new text uses Calibri
        doc = text_edit.document()
        default_font = QFont("Calibri", 11)
        doc.setDefaultFont(default_font)
        text_edit.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        return text_edit

    def create_grid_widget(self):
        """Create a table widget for grid mode."""
        print(f"[DEBUG] create_grid_widget() called for note uid={self.note.uid}")
        print(f"[DEBUG]   note.grid_data type: {type(self.note.grid_data)}")
        print(f"[DEBUG]   note.grid_data entries: {len(self.note.grid_data)}")
        if self.note.grid_data:
            print(f"[DEBUG]   grid_data sample: {list(self.note.grid_data.items())[:5]}")

        table = QTableWidget()
        table.setRowCount(15)  # More rows for better visibility
        table.setColumnCount(6)  # More columns (A-F) for better visibility
        table.setHorizontalHeaderLabels(['A', 'B', 'C', 'D', 'E', 'F'])

        # Set larger column widths for better visibility
        for col in range(6):
            table.setColumnWidth(col, 110)  # Increased from 80 to 110
        table.verticalHeader().setDefaultSectionSize(30)  # Increased from 25 to 30

        # Load grid data if exists
        cells_populated = 0
        for (row, col), value in self.note.grid_data.items():
            if isinstance(row, str):
                row, col = eval(row) if ',' in row else (int(row), int(col))
            # Ensure table is large enough
            if row >= table.rowCount():
                table.setRowCount(row + 1)
            if col >= table.columnCount():
                table.setColumnCount(col + 1)
            item = QTableWidgetItem(str(value))
            table.setItem(row, col, item)
            cells_populated += 1

        print(f"[DEBUG]   Populated {cells_populated} cells in table")

        # Initialize empty cells
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if not table.item(row, col):
                    table.setItem(row, col, QTableWidgetItem(''))

        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d0d0d0;
                font-size: 12px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 3px;
                border: 1px solid #ccc;
                font-size: 11px;
            }
        """)
        
        # Enable Excel-like editing behavior
        table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | 
                             QAbstractItemView.EditTrigger.EditKeyPressed | 
                             QAbstractItemView.EditTrigger.AnyKeyPressed)
        
        # Enable extended selection for drag-to-select multiple cells
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        
        # Enable row and column selection by clicking headers
        table.verticalHeader().setSectionsClickable(True)
        table.horizontalHeader().setSectionsClickable(True)
        table.verticalHeader().sectionClicked.connect(lambda row: self._select_row(table, row))
        table.horizontalHeader().sectionClicked.connect(lambda col: self._select_column(table, col))
        
        return table
    
    def _select_row(self, table, row):
        """Select entire row when header is clicked."""
        # Clear current selection
        table.clearSelection()
        # Select all cells in the row
        for col in range(table.columnCount()):
            item = table.item(row, col)
            if not item:
                item = QTableWidgetItem('')
                table.setItem(row, col, item)
            item.setSelected(True)
        # Set current cell to first column of selected row
        table.setCurrentCell(row, 0)
    
    def _select_column(self, table, col):
        """Select entire column when header is clicked."""
        # Clear current selection
        table.clearSelection()
        # Select all cells in the column
        for row in range(table.rowCount()):
            item = table.item(row, col)
            if not item:
                item = QTableWidgetItem('')
                table.setItem(row, col, item)
            item.setSelected(True)
        # Set current cell to first row of selected column
        table.setCurrentCell(0, col)

    def boundingRect(self):
        return QRectF(0, 0, self.note.width, self.note.height)

    def paint(self, painter, option, widget):
        """OneNote-style: invisible by default, subtle outline on hover, purple on select."""
        rect = self.boundingRect()
        is_interacting = self.is_hovered or self.isSelected()

        if is_interacting:
            # Subtle container outline on hover — faint gray with light tint
            painter.setPen(QPen(QColor("#D0D0D0"), 1))
            painter.setBrush(QColor(250, 250, 250, 40))  # Very faint tint
            painter.drawRoundedRect(rect.adjusted(1, 1, -1, -1), 4, 4)

        if self.isSelected():
            # Purple selection border (OneNote uses blue, we use purple to match theme)
            painter.setPen(QPen(QColor("#7719AA"), 1.5))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRoundedRect(rect.adjusted(1, 1, -1, -1), 4, 4)

        # Status dot — small colored circle in top-left, only on hover
        if is_interacting and self.note.status != "no_timeline":
            status_colors = {
                "active": "#9B59B6",
                "overdue": "#dc3545",
                "completed": "#28a745",
            }
            dot_color = status_colors.get(self.note.status)
            if dot_color:
                painter.setPen(Qt.PenStyle.NoPen)
                painter.setBrush(QColor(dot_color))
                painter.drawEllipse(QPointF(8, 8), 3, 3)

        # Move handle dots — lighter, only on hover
        if is_interacting:
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QColor("#C0C0C0"))
            for i in range(4):
                x = rect.width() / 2 - 15 + i * 10
                painter.drawEllipse(QPointF(x, 6), 1.5, 1.5)

    def hoverEnterEvent(self, event):
        self.is_hovered = True
        # Cancel any pending hide
        if hasattr(self, '_header_hide_timer'):
            self._header_hide_timer.stop()
        # Show header, checkbox, delete button on hover
        self.header_proxy.setVisible(True)
        self.checkbox_proxy.setVisible(True)
        if hasattr(self, 'delete_button_proxy'):
            self.delete_button_proxy.setVisible(True)
        self.update()

    def hoverLeaveEvent(self, event):
        self.is_hovered = False
        self.checkbox_proxy.setVisible(False)
        if hasattr(self, 'delete_button_proxy'):
            self.delete_button_proxy.setVisible(False)
        # Start delayed header hide (prevents flickering)
        if hasattr(self, '_header_hide_timer'):
            self._header_hide_timer.start()
        self.update()

    def _maybe_hide_header(self):
        """Hide header after delay, unless note has a title or is focused."""
        # Keep header visible if note has a title
        has_title = bool(self.note.heading and self.note.heading.strip())
        if has_title:
            return
        # Keep header visible if any child widget has focus
        if hasattr(self, 'content_widget') and self.content_widget and self.content_widget.hasFocus():
            return
        if self.heading_edit.hasFocus() or self.timeline_edit.hasFocus():
            return
        self.header_proxy.setVisible(False)

    # Manual resize removed - auto-expand handles sizing

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged:
            # Update note position
            self.note.x_pos = int(self.pos().x())
            self.note.y_pos = int(self.pos().y())
            # Save after position change (triggered after drag completes)
            if self.parent_canvas and hasattr(self.parent_canvas, 'save_notes'):
                self.parent_canvas.save_notes()
        elif change == QGraphicsItem.GraphicsItemChange.ItemPositionChange:
            # Update delete button position when note moves
            if hasattr(self, 'delete_button_proxy'):
                delete_button_x = self.note.width - self.border_margin - 24
                delete_button_y = self.border_margin + 3
                self.delete_button_proxy.setPos(delete_button_x, delete_button_y)
        return super().itemChange(change, value)

    def on_heading_changed(self):
        """Handle heading field changes."""
        try:
            heading = self.heading_edit.text()
            self.note.heading = heading
            # Keep header visible if title is entered (OneNote behavior)
            if heading.strip() and hasattr(self, 'header_proxy'):
                self.header_proxy.setVisible(True)
            # Save changes for committed notes
            if self.parent_canvas and not self.is_pending:
                self.parent_canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in on_heading_changed: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("HEADING CHANGED ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def on_text_changed(self):
        """Handle text changes (canvas mode)."""
        try:
            if self.note.note_type != "canvas":
                return
            # Save as HTML for rich text, get plain text for parsing
            self.note.text = self.content_widget.toHtml()
            text = self.content_widget.toPlainText()

            if self.is_pending and text.strip():
                self.commit_pending_note()

            # Attempt to parse timeline information
            reminder_date, cleaned_text = parse_natural_date(text)
            if reminder_date:
                # Only update if changed to avoid repeated saves
                if self.note.reminder_date != reminder_date:
                    self.note.reminder_date = reminder_date
                    self.note.update_status()
                    self.update()
                    if self.parent_canvas:
                        self.parent_canvas.update_status_indicators()
                        self.parent_canvas.save_notes()

            # Persist text changes for committed notes
            if self.parent_canvas and not self.is_pending:
                self.parent_canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in on_text_changed: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("TEXT CHANGED ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def on_grid_changed(self, row, col):
        """Handle grid cell changes (grid mode)."""
        try:
            if self.note.note_type != "grid":
                return

            # Save grid data
            item = self.content_widget.item(row, col)
            if item:
                cell_text = item.text().strip()
                if cell_text:
                    # Cell has content, save it
                    self.note.grid_data[(row, col)] = cell_text
                else:
                    # Cell is empty, remove it from grid_data if it exists
                    if (row, col) in self.note.grid_data:
                        del self.note.grid_data[(row, col)]

            # Check if grid is empty and note is pending
            # If grid is empty, don't commit (will be discarded on focus loss)
            if self.is_pending:
                if not self.is_note_empty():
                    # Grid has content, commit the note
                    self.commit_pending_note()
                # If grid is empty, leave it pending (will be discarded on focus loss)

            if self.parent_canvas:
                self.parent_canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in on_grid_changed: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("GRID CHANGED ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def _safe_adjust_size(self):
        """Wrapper for deferred adjust_size_to_content with error handling."""
        debug_log(f"_safe_adjust_size CALLED (deferred timer fired)")
        try:
            self.adjust_size_to_content()
        except Exception as e:
            debug_log(f"_safe_adjust_size FAILED: {e}", "ERROR")

    def adjust_size_to_content(self):
        """
        Auto-expand height to fit text content. Width stays fixed (OneNote-style).
        Text wraps within the note width; only height changes.
        """
        try:
            # Only for canvas notes
            if hasattr(self.note, 'note_type') and self.note.note_type != "canvas":
                return
            if not hasattr(self, 'content_widget') or self.content_widget is None:
                return
            if not hasattr(self, 'note') or self.note is None:
                return

            # Ensure word wrap is always on
            self.content_widget.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)

            # Use current width (or default 300 for new notes)
            current_width = max(250, getattr(self.note, 'width', 300))
            content_area_width = current_width - self.border_margin * 2

            # Calculate document height at current width
            doc = self.content_widget.document()
            doc.setTextWidth(content_area_width)
            doc_height = doc.size().height()

            min_content_height = 60
            content_height = max(min_content_height, int(doc_height) + 10)
            total_height = content_height + self.HEADER_HEIGHT + self.border_margin * 2

            if current_width != self.note.width or total_height != self.note.height:
                self.prepareGeometryChange()
                self.note.width = current_width
                self.note.height = total_height

                text_edit_width = max(50, content_area_width)
                text_edit_height = max(30, content_height)
                self.content_widget.setFixedSize(text_edit_width, text_edit_height)
                self.proxy.resize(text_edit_width, text_edit_height)

                header_width = max(50, content_area_width)
                self.header_container.setFixedWidth(header_width)

                # Update delete button position
                if hasattr(self, 'delete_button_proxy'):
                    delete_x = self.note.width - self.border_margin - 24
                    self.delete_button_proxy.setPos(delete_x, self.border_margin + 3)

                self.update()
        except Exception as e:
            # Don't crash — keep current size
            try:
                if hasattr(self, 'note') and hasattr(self, 'HEADER_HEIGHT'):
                    if not hasattr(self.note, 'width') or self.note.width < 150:
                        self.note.width = 150
                    min_content = 60
                    min_total_height = min_content + self.HEADER_HEIGHT + self.border_margin * 2
                    if not hasattr(self.note, 'height') or self.note.height < min_total_height:
                        self.note.height = min_total_height
            except:
                pass

    def on_checkbox_toggled(self, state):
        """Handle completion checkbox toggle."""
        try:
            self.note.is_completed = (state == Qt.CheckState.Checked.value)
            self.note.update_status()
            self.update()
            # Trigger cascade update and save
            if self.parent_canvas:
                self.parent_canvas.update_status_indicators()
                self.parent_canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in checkbox toggle: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("CHECKBOX TOGGLE ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")
    
    def delete_note(self):
        """Delete this note container. Always works regardless of focus."""
        try:
            canvas = self.parent_canvas
            if not canvas:
                return
            
            # Remove from parent canvas containers list
            if self in canvas.containers:
                canvas.containers.remove(self)
            
            # Remove from pending containers if present
            if self in canvas.pending_containers:
                canvas.pending_containers.discard(self)
            
            # Remove from current item's notes
            if canvas.current_item and self.note in canvas.current_item.notes:
                canvas.current_item.notes.remove(self.note)
            
            # Remove from scene
            if self.scene():
                self.scene().removeItem(self)
            
            # Save changes
            canvas.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error deleting note: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("DELETE NOTE ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")
            
            # Try to show error dialog
            try:
                app = QApplication.instance()
                if app:
                    QMessageBox.critical(
                        None,
                        "Error",
                        f"An error occurred while toggling note completion:\n\n{type(e).__name__}: {e}\n\nSee console for details."
                    )
            except:
                pass

    def commit_pending_note(self):
        """Commit this note if it was pending."""
        try:
            if self.is_pending and self.parent_canvas:
                self.parent_canvas.commit_container(self)
        except Exception as e:
            import traceback
            error_msg = f"Error in commit_pending_note: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("COMMIT PENDING NOTE ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def is_note_empty(self):
        """Check if note is empty (no content)."""
        try:
            if self.note.note_type == "grid":
                # For grid notes, check both the table widget and grid_data
                # First check if heading exists
                has_heading = self.note.heading and self.note.heading.strip()
                
                # Check the actual table widget if available (more accurate for pending notes)
                if hasattr(self, 'content_widget') and self.content_widget:
                    table = self.content_widget
                    # Check if any cell in the table has non-empty content
                    for row in range(table.rowCount()):
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            if item and item.text().strip():
                                return False  # Found non-empty cell
                
                # Also check grid_data (for saved notes)
                if self.note.grid_data:
                    for value in self.note.grid_data.values():
                        if value and str(value).strip():
                            return False  # Found non-empty cell in saved data
                
                # No non-empty cells found, check if heading exists
                return not has_heading
            else:
                # Canvas note: check plain text (note.text may be HTML)
                if hasattr(self, 'content_widget') and self.content_widget:
                    return not self.content_widget.toPlainText().strip()
                return not (self.note.text and self.note.text.strip())
        except Exception as e:
            # If error, assume note is not empty (safer default)
            return False

    def discard_if_empty(self):
        """Remove this container if it is pending and empty."""
        try:
            if self.is_pending and self.parent_canvas and self.is_note_empty():
                self.parent_canvas.discard_pending_container(self)
        except Exception as e:
            import traceback
            error_msg = f"Error in discard_if_empty: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("DISCARD IF EMPTY ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")


class FormattingToolbar(QWidget):
    """OneNote-style rich text formatting toolbar."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._editor = None
        self.setFixedHeight(34)
        self.setStyleSheet("""
            QWidget { background: #F9F7FB; border-bottom: 1px solid #E0D8E8; }
            QToolButton { border: none; padding: 4px 7px; border-radius: 3px; font-size: 12px; min-width: 24px; font-family: 'Calibri'; }
            QToolButton:hover { background: #EDE7F3; }
            QToolButton:checked { background: #D5C0E8; }
            QToolButton:disabled { color: #C0C0C0; }
            QFontComboBox, QComboBox { border: 1px solid #E0D8E8; border-radius: 3px; padding: 2px 4px; font-size: 11px; max-width: 130px; background: white; font-family: 'Calibri'; }
            QFontComboBox:disabled, QComboBox:disabled { background: #F0F0F0; color: #C0C0C0; }
        """)
        h = QHBoxLayout(self)
        h.setContentsMargins(6, 2, 6, 2)
        h.setSpacing(2)

        # Bold
        self.bold_btn = QToolButton(); self.bold_btn.setText("B"); self.bold_btn.setCheckable(True)
        self.bold_btn.setStyleSheet(self.bold_btn.styleSheet() + "QToolButton { font-weight: bold; }")
        self.bold_btn.setShortcut("Ctrl+B"); self.bold_btn.clicked.connect(self.toggle_bold)
        h.addWidget(self.bold_btn)

        # Italic
        self.italic_btn = QToolButton(); self.italic_btn.setText("I"); self.italic_btn.setCheckable(True)
        self.italic_btn.setStyleSheet(self.italic_btn.styleSheet() + "QToolButton { font-style: italic; }")
        self.italic_btn.setShortcut("Ctrl+I"); self.italic_btn.clicked.connect(self.toggle_italic)
        h.addWidget(self.italic_btn)

        # Underline
        self.underline_btn = QToolButton(); self.underline_btn.setText("U"); self.underline_btn.setCheckable(True)
        self.underline_btn.setStyleSheet(self.underline_btn.styleSheet() + "QToolButton { text-decoration: underline; }")
        self.underline_btn.setShortcut("Ctrl+U"); self.underline_btn.clicked.connect(self.toggle_underline)
        h.addWidget(self.underline_btn)

        # Separator
        sep1 = QFrame(); sep1.setFrameShape(QFrame.Shape.VLine); sep1.setFixedWidth(1)
        sep1.setStyleSheet("background: #E0D8E8;"); h.addWidget(sep1)

        # Font family
        self.font_combo = QFontComboBox()
        self.font_combo.setCurrentFont(QFont("Calibri"))
        self.font_combo.currentFontChanged.connect(self.set_font_family)
        h.addWidget(self.font_combo)

        # Font size
        self.size_combo = QComboBox()
        self.size_combo.setEditable(True)
        for s in [8, 9, 10, 11, 12, 14, 16, 18, 20, 24, 28, 36]:
            self.size_combo.addItem(str(s))
        self.size_combo.setCurrentText("11")
        self.size_combo.currentTextChanged.connect(self.set_font_size)
        self.size_combo.setFixedWidth(55)
        h.addWidget(self.size_combo)

        # Separator
        sep2 = QFrame(); sep2.setFrameShape(QFrame.Shape.VLine); sep2.setFixedWidth(1)
        sep2.setStyleSheet("background: #E0D8E8;"); h.addWidget(sep2)

        # Text color
        self.color_btn = QToolButton(); self.color_btn.setText("A"); self._text_color = QColor("#1A1A2E")
        self.color_btn.setStyleSheet(self.color_btn.styleSheet() + f"QToolButton {{ color: {self._text_color.name()}; font-weight: bold; }}")
        self.color_btn.clicked.connect(self.pick_text_color); h.addWidget(self.color_btn)

        # Highlight
        self.highlight_btn = QToolButton(); self.highlight_btn.setText("ab"); self._highlight_color = QColor("#FFFF00")
        self.highlight_btn.setStyleSheet(self.highlight_btn.styleSheet() + f"QToolButton {{ background: {self._highlight_color.name()}; }}")
        self.highlight_btn.clicked.connect(self.pick_highlight); h.addWidget(self.highlight_btn)

        # Separator
        sep3 = QFrame(); sep3.setFrameShape(QFrame.Shape.VLine); sep3.setFixedWidth(1)
        sep3.setStyleSheet("background: #E0D8E8;"); h.addWidget(sep3)

        # Bullet list
        self.bullet_btn = QToolButton(); self.bullet_btn.setText("•≡")
        self.bullet_btn.clicked.connect(self.toggle_bullet_list); h.addWidget(self.bullet_btn)

        # Clear formatting
        self.clear_btn = QToolButton(); self.clear_btn.setText("Tx")
        self.clear_btn.setToolTip("Clear formatting")
        self.clear_btn.clicked.connect(self.clear_formatting); h.addWidget(self.clear_btn)

        h.addStretch()

    # --- Editor connection ---
    def set_editor(self, editor):
        """Connect toolbar to a QTextEdit instance."""
        if self._editor:
            try: self._editor.cursorPositionChanged.disconnect(self.update_state)
            except: pass
        self._editor = editor
        if editor:
            editor.cursorPositionChanged.connect(self.update_state)
            self.update_state()
        self._update_enabled_state()

    def _update_enabled_state(self):
        """Gray out toolbar buttons when no editor is connected."""
        enabled = self._editor is not None
        for child in self.findChildren(QToolButton):
            child.setEnabled(enabled)
        self.font_combo.setEnabled(enabled)
        self.size_combo.setEnabled(enabled)

    def update_state(self):
        """Sync button checked states with current cursor formatting."""
        if not self._editor: return
        fmt = self._editor.currentCharFormat()
        self.bold_btn.setChecked(fmt.fontWeight() == QFont.Weight.Bold)
        self.italic_btn.setChecked(fmt.fontItalic())
        self.underline_btn.setChecked(fmt.fontUnderline())

    # --- Formatting actions ---
    def _apply_char_format(self, modify_fn):
        if not self._editor: return
        fmt = self._editor.currentCharFormat()
        modify_fn(fmt)
        self._editor.mergeCurrentCharFormat(fmt)

    def toggle_bold(self):
        self._apply_char_format(lambda f: f.setFontWeight(
            QFont.Weight.Normal if f.fontWeight() == QFont.Weight.Bold else QFont.Weight.Bold))

    def toggle_italic(self):
        self._apply_char_format(lambda f: f.setFontItalic(not f.fontItalic()))

    def toggle_underline(self):
        self._apply_char_format(lambda f: f.setFontUnderline(not f.fontUnderline()))

    def set_font_family(self, font):
        if not self._editor: return
        self._apply_char_format(lambda f: f.setFontFamilies([font.family()]))

    def set_font_size(self, size_str):
        if not self._editor: return
        try:
            size = int(size_str)
            if 1 <= size <= 200:
                self._apply_char_format(lambda f: f.setFontPointSize(size))
        except ValueError:
            pass

    def pick_text_color(self):
        if not self._editor: return
        color = QColorDialog.getColor(self._text_color, self, "Text Color")
        if color.isValid():
            self._text_color = color
            self.color_btn.setStyleSheet(f"QToolButton {{ color: {color.name()}; font-weight: bold; border: none; padding: 4px 7px; border-radius: 3px; }} QToolButton:hover {{ background: #E8D5F0; }}")
            self._apply_char_format(lambda f: f.setForeground(QBrush(color)))

    def pick_highlight(self):
        if not self._editor: return
        color = QColorDialog.getColor(self._highlight_color, self, "Highlight Color")
        if color.isValid():
            self._highlight_color = color
            self.highlight_btn.setStyleSheet(f"QToolButton {{ background: {color.name()}; border: none; padding: 4px 7px; border-radius: 3px; }} QToolButton:hover {{ background: #E8D5F0; }}")
            self._apply_char_format(lambda f: f.setBackground(QBrush(color)))

    def toggle_bullet_list(self):
        if not self._editor: return
        cursor = self._editor.textCursor()
        current_list = cursor.currentList()
        if current_list:
            # Remove list formatting
            block_fmt = cursor.blockFormat()
            block_fmt.setIndent(0)
            cursor.setBlockFormat(block_fmt)
            current_list.remove(cursor.block())
        else:
            list_fmt = QTextListFormat()
            list_fmt.setStyle(QTextListFormat.Style.ListDisc)
            cursor.createList(list_fmt)

    def clear_formatting(self):
        if not self._editor: return
        fmt = QTextCharFormat()
        self._editor.setCurrentCharFormat(fmt)


class NotesCanvas(QGraphicsView):
    """
    OneNote-style canvas for free-form note containers.
    2000x2000 pixel canvas where users can click anywhere to create notes.
    """

    note_added = pyqtSignal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_item = None
        self.notes_view_parent = parent
        self.show_grid = False  # Toggle for grid lines
        self.grid_size = 30  # Grid cell size in pixels

        # Create scene
        self.scene = QGraphicsScene(0, 0, 2000, 2000)
        self.scene.setBackgroundBrush(QBrush(QColor("white")))
        self.setScene(self.scene)

        # View settings
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setDragMode(QGraphicsView.DragMode.RubberBandDrag)

        # Container tracking
        self.containers = []
        self.pending_containers = set()

    def toggle_grid(self, enabled):
        """Toggle grid lines on/off."""
        self.show_grid = enabled
        self.viewport().update()  # Force redraw

    def drawBackground(self, painter, rect):
        """Draw the background with optional grid lines."""
        super().drawBackground(painter, rect)

        if not self.show_grid:
            return

        # Draw grid lines
        painter.setPen(QPen(QColor(220, 220, 220), 1))

        # Vertical lines
        left = int(rect.left()) - (int(rect.left()) % self.grid_size)
        top = int(rect.top()) - (int(rect.top()) % self.grid_size)

        x = left
        while x < rect.right():
            painter.drawLine(x, int(rect.top()), x, int(rect.bottom()))
            x += self.grid_size

        # Horizontal lines
        y = top
        while y < rect.bottom():
            painter.drawLine(int(rect.left()), y, int(rect.right()), y)
            y += self.grid_size

    def set_item(self, item):
        """Set the current item and load its notes as containers."""
        safe_name = item.name.encode('ascii', errors='replace').decode('ascii') if item else None
        print(f"[DEBUG] NotesCanvas.set_item() called with item: {safe_name}")
        self.current_item = item

        # Clear existing containers
        self.scene.clear()
        self.containers = []
        self.pending_containers.clear()

        if not item:
            print(f"[DEBUG]   No item provided, returning")
            return

        print(f"[DEBUG]   Item has {len(item.notes)} notes")

        # Create containers for each note
        for note in item.notes:
            print(f"[DEBUG]   Creating container for note: uid={note.uid}, type={note.note_type}")
            container = NoteContainer(note, self)
            self.scene.addItem(container)
            self.containers.append(container)
            print(f"[DEBUG]   Container created and added to scene")

        print(f"[DEBUG]   Total containers created: {len(self.containers)}")

    def mousePressEvent(self, event):
        """Handle click to create new note container."""
        debug_log(f"NotesCanvas.mousePressEvent ENTRY: pos={event.pos()}, button={event.button()}")
        try:
            # Check if clicked on empty space
            item = self.itemAt(event.pos())
            debug_log(f"  - itemAt result: {item}")

            if not item and event.button() == Qt.MouseButton.LeftButton:
                debug_log(f"  - Clicked on empty space, creating new note...")
                # Create new note at click position
                scene_pos = self.mapToScene(event.pos())
                debug_log(f"  - Scene position: ({scene_pos.x()}, {scene_pos.y()})")

                # Determine note type from checkbox
                note_type = "canvas"  # default
                if self.notes_view_parent and hasattr(self.notes_view_parent, 'grid_toggle_checkbox'):
                    note_type = "grid" if self.notes_view_parent.grid_toggle_checkbox.isChecked() else "canvas"
                debug_log(f"  - Note type: {note_type}")

                # Set default size based on note type
                # For grid notes: 700x500 content + header(35) + borders(6*2) = 712x547
                # For canvas notes: use default 200x100 (will auto-expand)
                if note_type == "grid":
                    # Grid notes: larger default size
                    border_margin = 3
                    header_height = 28
                    grid_content_width = 700
                    grid_content_height = 500
                    default_width = grid_content_width + border_margin * 2
                    default_height = grid_content_height + header_height + border_margin * 2
                else:
                    # Canvas notes: OneNote-style fixed width, auto-height
                    default_width = 300
                    default_height = 100

                # Create new Note
                debug_log(f"  - Creating Note object...")
                new_note = Note("", x_pos=int(scene_pos.x()), y_pos=int(scene_pos.y()), 
                              note_type=note_type, width=default_width, height=default_height)
                debug_log(f"  - Note created: uid={new_note.uid}, pos=({new_note.x_pos}, {new_note.y_pos}), type={new_note.note_type}, size={new_note.width}x{new_note.height}")

                # Create container (pending until user enters text)
                debug_log(f"  - Creating NoteContainer...")
                container = NoteContainer(new_note, self)
                debug_log(f"  - NoteContainer created successfully")

                container.is_pending = True
                debug_log(f"  - Adding to scene...")
                self.scene.addItem(container)
                self.containers.append(container)
                self.pending_containers.add(container)
                debug_log(f"  - Added to scene. Total containers: {len(self.containers)}, Pending: {len(self.pending_containers)}")

                # Focus the content widget
                debug_log(f"  - Setting focus to content widget...")
                container.content_widget.setFocus()
                debug_log(f"  - Focus set")

                # Accept the event so base class doesn't steal focus (which discards pending note)
                event.accept()
                debug_log(f"NotesCanvas.mousePressEvent SUCCESS: New note container created")
                return

            debug_log(f"  - Delegating to super().mousePressEvent()")
            super().mousePressEvent(event)
            debug_log(f"NotesCanvas.mousePressEvent COMPLETE")

        except Exception as e:
            debug_log(f"NotesCanvas.mousePressEvent FAILED: {e}", "ERROR")
            # Don't crash the app - just log and continue
            event.accept()  # Accept event to prevent further processing

    def commit_container(self, container):
        """Commit a pending container once it has content."""
        try:
            if not container.parent_canvas is self:
                return

            # Check if note has content before committing
            # If note is empty, discard it instead of committing
            if container.is_note_empty():
                # Note is empty, discard it
                self.discard_pending_container(container)
                return

            if container.is_pending:
                container.is_pending = False
                self.pending_containers.discard(container)

                if self.current_item and container.note not in self.current_item.notes:
                    self.current_item.add_note(container.note)

                # Emit signal for committed note
                self.note_added.emit(container.note)

            # Persist updates
            self.update_status_indicators()
            self.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in commit_container: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("COMMIT CONTAINER ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def discard_pending_container(self, container):
        """Remove a pending container that was left empty."""
        try:
            if container in self.pending_containers:
                self.pending_containers.discard(container)
            if container in self.containers:
                self.containers.remove(container)

            container.is_pending = False

            # Ensure note is not stored on the current item
            if self.current_item and container.note in getattr(self.current_item, "notes", []):
                self.current_item.notes.remove(container.note)

            # Remove from scene
            if container.scene():
                container.scene().removeItem(container)

            # Persist cleanup
            self.save_notes()
        except Exception as e:
            import traceback
            error_msg = f"Error in discard_pending_container: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("DISCARD PENDING CONTAINER ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")

    def update_status_indicators(self):
        """Trigger update of status indicators in parent view."""
        if self.notes_view_parent and hasattr(self.notes_view_parent, 'update_indicators'):
            self.notes_view_parent.update_indicators()

    def clear(self):
        """Clear the canvas."""
        self.scene.clear()
        self.containers = []
        self.current_item = None
        self.pending_containers.clear()

    def save_notes(self):
        """Trigger save of notes structure to config."""
        try:
            if self.notes_view_parent and hasattr(self.notes_view_parent, 'main_window'):
                main_window = self.notes_view_parent.main_window
                if hasattr(main_window, 'save_notes_structure'):
                    main_window.save_notes_structure()
        except Exception as e:
            import traceback
            error_msg = f"Error in save_notes: {e}\n\n{traceback.format_exc()}"
            print(f"\n{'='*60}")
            print("SAVE NOTES ERROR")
            print(f"{'='*60}")
            print(error_msg)
            print(f"{'='*60}\n")


class TableNotesCanvas(QTableWidget):
    """
    Simple Excel-style grid view - just cells with toggleable grid lines.
    """

    note_added = pyqtSignal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_item = None
        self.notes_view_parent = parent
        self.grid_visible = True  # Grid lines visible by default

        # Setup table with many rows and columns (like Excel)
        num_rows = 100
        num_cols = 26  # A-Z columns
        self.setRowCount(num_rows)
        self.setColumnCount(num_cols)

        # Set column headers (A, B, C, ...)
        headers = [chr(65 + i) for i in range(num_cols)]
        self.setHorizontalHeaderLabels(headers)

        # Set default column width
        for col in range(num_cols):
            self.setColumnWidth(col, 100)

        # Set default row height
        self.verticalHeader().setDefaultSectionSize(25)

        # Initialize all cells
        for row in range(num_rows):
            for col in range(num_cols):
                item = QTableWidgetItem('')
                self.setItem(row, col, item)

        # Enable features
        self.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed | QTableWidget.EditTrigger.AnyKeyPressed)

        # Styling
        self.update_grid_style()

        # Install event filter for Excel-like navigation
        self.installEventFilter(self)

        # Connect signals
        self.cellChanged.connect(self.on_cell_changed)

    def update_grid_style(self):
        """Update the grid styling based on grid_visible flag."""
        if self.grid_visible:
            # Show grid lines
            self.setStyleSheet("""
                QTableWidget {
                    background-color: white;
                    gridline-color: #d0d0d0;
                    font-size: 12px;
                }
                QTableWidget::item {
                    padding: 3px;
                }
                QTableWidget::item:selected {
                    background-color: #cce8ff;
                    color: black;
                }
                QHeaderView::section {
                    background-color: #f0f0f0;
                    padding: 3px;
                    border: 1px solid #ccc;
                    font-size: 11px;
                }
            """)
            self.setShowGrid(True)
        else:
            # Hide grid lines
            self.setStyleSheet("""
                QTableWidget {
                    background-color: white;
                    gridline-color: transparent;
                    font-size: 12px;
                }
                QTableWidget::item {
                    padding: 3px;
                }
                QTableWidget::item:selected {
                    background-color: #cce8ff;
                    color: black;
                }
                QHeaderView::section {
                    background-color: #f0f0f0;
                    padding: 3px;
                    border: 1px solid #ccc;
                    font-size: 11px;
                }
            """)
            self.setShowGrid(False)

    def toggle_grid_lines(self):
        """Toggle grid lines visibility."""
        self.grid_visible = not self.grid_visible
        self.update_grid_style()

    def eventFilter(self, obj, event):
        """Handle Excel-like keyboard navigation."""
        if obj == self and event.type() == event.Type.KeyPress:
            key = event.key()
            modifiers = event.modifiers()

            # Tab: Move to next cell (right)
            if key == Qt.Key.Key_Tab and modifiers == Qt.KeyboardModifier.NoModifier:
                current = self.currentRow(), self.currentColumn()
                next_col = current[1] + 1
                if next_col >= self.columnCount():
                    next_col = 0
                    next_row = current[0] + 1
                    if next_row >= self.rowCount():
                        # Add more rows if needed
                        self.setRowCount(self.rowCount() + 20)
                        for row in range(self.rowCount() - 20, self.rowCount()):
                            for col in range(self.columnCount()):
                                if not self.item(row, col):
                                    self.setItem(row, col, QTableWidgetItem(''))
                    self.setCurrentCell(next_row if next_col == 0 else current[0], next_col)
                else:
                    self.setCurrentCell(current[0], next_col)
                return True

            # Enter: Move down
            elif (key == Qt.Key.Key_Return or key == Qt.Key.Key_Enter) and modifiers == Qt.KeyboardModifier.NoModifier:
                current = self.currentRow(), self.currentColumn()
                next_row = current[0] + 1
                if next_row >= self.rowCount():
                    # Add more rows if needed
                    self.setRowCount(self.rowCount() + 20)
                    for row in range(self.rowCount() - 20, self.rowCount()):
                        for col in range(self.columnCount()):
                            if not self.item(row, col):
                                self.setItem(row, col, QTableWidgetItem(''))
                self.setCurrentCell(next_row, current[1])
                return True

            # Right arrow: Move right
            elif key == Qt.Key.Key_Right and modifiers == Qt.KeyboardModifier.NoModifier:
                current = self.currentRow(), self.currentColumn()
                next_col = min(current[1] + 1, self.columnCount() - 1)
                self.setCurrentCell(current[0], next_col)
                return True

        return super().eventFilter(obj, event)

    def on_cell_changed(self, row, col):
        """Handle cell content change."""
        # Auto-save when cell changes
        self.save_grid_data()

    def set_item(self, item):
        """Set the current NotesItem and load its grid data."""
        self.current_item = item
        self.blockSignals(True)

        if not item:
            self.blockSignals(False)
            return

        # Load grid data from item if it exists
        if hasattr(item, 'grid_data') and item.grid_data:
            for cell_key, cell_value in item.grid_data.items():
                try:
                    row, col = map(int, cell_key.split(','))
                    if row < self.rowCount() and col < self.columnCount():
                        item_widget = self.item(row, col)
                        if item_widget:
                            item_widget.setText(cell_value)
                except:
                    pass

        self.blockSignals(False)

    def save_grid_data(self):
        """Save grid data to the current item."""
        if not self.current_item:
            return

        # Store non-empty cells
        grid_data = {}
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item and item.text().strip():
                    grid_data[f"{row},{col}"] = item.text()

        # Save to item
        self.current_item.grid_data = grid_data

        # Trigger main window save
        try:
            if self.notes_view_parent and hasattr(self.notes_view_parent, 'main_window'):
                main_window = self.notes_view_parent.main_window
                if hasattr(main_window, 'save_notes_structure'):
                    main_window.save_notes_structure()
        except Exception as e:
            pass

    def clear(self):
        """Clear the grid."""
        self.blockSignals(True)
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    item.setText('')
        self.current_item = None
        self.blockSignals(False)


# ============================================================================

# ============================================================================

class TimelineCanvas(QWidget):
    comment_requested, task_double_clicked, progress_changed, edit_task_requested, phase_complete_requested, active_zone_changed, delete_task_requested, summary_requested, edit_project_requested, add_to_notes_requested = pyqtSignal(object), pyqtSignal(object), pyqtSignal(), pyqtSignal(object), pyqtSignal(object), pyqtSignal(object, bool), pyqtSignal(object), pyqtSignal(object), pyqtSignal(object), pyqtSignal(object)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.tasks_to_display = []
        self.project_start_date, self.project_end_date = None, None
        self.pixels_per_day = 10
        self.timescale_view = TimescaleView.MONTH
        self.setStyleSheet("background-color: #ffffff;")
        self.header_height, self.row_height, self.bar_height_ratio = 50, 30, 0.6
        self.setMouseTracking(True)
        self.last_hovered_task = None
        self.highlighted_task = None  # Task to highlight from search
        self.highlight_timer = None  # Timer to clear highlight
        self._bom_blocking = {}  # item_id -> blocking info dict
        self._badge_rects = []  # [(QRectF, task, blocking_info), ...] for click detection

    def set_bom_blocking_data(self, blocking_dict):
        """Set BOM blocking data dict (item_id -> blocking info)."""
        self._bom_blocking = blocking_dict or {}

    def _get_blocking_for_task(self, task):
        """Check if a task/item has BOM blocking info. Returns info dict or None."""
        if not self._bom_blocking:
            return None
        # Match by item id - for ECN Items, the id field is the part number
        task_id = getattr(task, 'id', None) or ""
        if task_id and str(task_id) in self._bom_blocking:
            return self._bom_blocking[str(task_id)]
        # Also try matching by name (some tasks use the item_id as name)
        task_name = getattr(task, 'name', "") or ""
        if task_name in self._bom_blocking:
            return self._bom_blocking[task_name]
        return None

    def set_tasks(self, tasks):
        self.tasks_to_display = [t for t in tasks if t.start and t.finish and t.name]
        required_height = self.header_height + (len(self.tasks_to_display) * self.row_height) + 20
        self.setMinimumHeight(required_height)
        self.update()

    def highlight_task(self, task_uid):
        """Highlight a specific task by its UID.
        Highlight stays until user clicks the bar, does a new search, or navigates away."""
        for task in self.tasks_to_display:
            if hasattr(task, 'uid') and task.uid == task_uid:
                self.highlighted_task = task
                self.update()
                break

    def clear_highlight(self):
        """Clear the current highlight"""
        self.highlighted_task = None
        self.update()

    def mousePressEvent(self, event):
        """Handle click — clear highlight if user clicks the highlighted bar, or show badge popup."""
        if event.button() == Qt.MouseButton.LeftButton:
            pos_f = QPointF(event.pos())

            # Check if clicking a ⚠️ blocking badge
            for badge_rect, task, blocking_info in self._badge_rects:
                if badge_rect.contains(pos_f):
                    self._show_blocking_popup(task, blocking_info, event.globalPosition().toPoint())
                    return

            # Clear highlight if clicking on highlighted bar
            if self.highlighted_task and self.highlighted_task.bar_rect.contains(pos_f):
                self.clear_highlight()
                return
        super().mousePressEvent(event)

    def _show_blocking_popup(self, task, info, global_pos):
        """Show a popup dialog with blocking details when ⚠️ badge is clicked."""
        item_id = info.get("item_id", "")
        rev = info.get("rev", "")
        name = info.get("name", "")
        label = f"{item_id}/{rev}" if rev else item_id
        blocking_children = info.get("blocking_children", [])
        is_released = info.get("is_released", False)
        release_text = info.get("release_text", "")
        ecn = info.get("ecn", "")
        nom = info.get("nom", "")
        total = info.get("total_children", 0)
        released = info.get("released_children", 0)
        chains = info.get("blocking_chain", [])
        assignee = info.get("assignee", "")
        current_task = info.get("task", "")

        lines = []
        lines.append(f"<h3>{label} — {name}</h3>")

        if not is_released and not blocking_children:
            # This is the actual problem part
            lines.append("<p style='color:#dc3545;font-weight:bold;'>⚠️ NOT PRODUCTION RELEASED</p>")
            if release_text:
                lines.append(f"<p><b>Has:</b> {release_text}</p>")
            lines.append("<p><b>Missing:</b> PRODUCTION RELEASED</p>")
            if ecn:
                lines.append(f"<p><b>ECN:</b> {ecn}</p>")
            if assignee:
                task_info = f" ({current_task})" if current_task else ""
                lines.append(f"<p><b>Assigned to:</b> {assignee}{task_info}</p>")
        elif blocking_children:
            # This is a parent assembly with blocking children
            lines.append(f"<p style='color:#e67e22;font-weight:bold;'>⚠️ LINK CHECK: {len(blocking_children)} children blocking this assembly</p>")
            if total > 0:
                pct = released / total * 100
                lines.append(f"<p><b>Summary:</b> {released} of {total} children released ({pct:.1f}%)</p>")
            lines.append("<hr>")
            for bc in blocking_children[:15]:  # Limit to 15 to avoid huge popup
                bc_label = f"{bc['item_id']}/{bc['rev']}" if bc.get('rev') else bc['item_id']
                lines.append(f"<p style='margin-left:10px;'>✗ <b>{bc_label}</b> — {bc.get('name', '')}")
                details = []
                if bc.get('ecn'):
                    details.append(f"ECN: {bc['ecn']}")
                if bc.get('release_text'):
                    details.append(f"Has: {bc['release_text']}")
                else:
                    details.append("Has: (none)")
                details.append("Missing: PRODUCTION RELEASED")
                if bc.get('assignee'):
                    task_str = f" ({bc['task']})" if bc.get('task') else ""
                    details.append(f"Assigned to: {bc['assignee']}{task_str}")
                lines.append(f"<br><span style='color:#666;margin-left:20px;'>{'  |  '.join(details)}</span></p>")
            if len(blocking_children) > 15:
                lines.append(f"<p style='color:#888;'>... and {len(blocking_children) - 15} more</p>")

            # Show blocking chains if available
            if chains:
                lines.append("<hr><p><b>Blocking chains:</b></p>")
                for chain in chains[:10]:
                    lines.append(f"<p style='margin-left:10px;color:#666;'>{chain}</p>")
                if len(chains) > 10:
                    lines.append(f"<p style='color:#888;'>... and {len(chains) - 10} more</p>")

        # Create popup
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton
        popup = QDialog(self)
        popup.setWindowTitle(f"⚠️ Blocking Details — {label}")
        popup.setMinimumWidth(500)
        popup.setMaximumWidth(700)
        layout = QVBoxLayout(popup)

        content = QLabel("\n".join(lines))
        content.setWordWrap(True)
        content.setTextFormat(Qt.TextFormat.RichText)
        content.setStyleSheet("padding: 10px; font-size: 12px;")
        layout.addWidget(content)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(popup.close)
        layout.addWidget(close_btn)

        popup.move(global_pos)
        popup.exec()

    def set_project_duration(self, start_date, end_date):
        if not start_date or not end_date: return
        self.project_start_date = start_date
        self.project_end_date = end_date + datetime.timedelta(days=30)
        self.update_canvas_width()

    def set_timescale(self, timescale, viewport_width):
        self.timescale_view = timescale
        if viewport_width <= 0: return

        if self.timescale_view == TimescaleView.YEAR:
            self.pixels_per_day = viewport_width / 365.25
        elif self.timescale_view == TimescaleView.MONTH:
            self.pixels_per_day = viewport_width / 30.5
        else: # WEEK
            self.pixels_per_day = viewport_width / 7
        
        self.update_canvas_width()

    def update_canvas_width(self):
        if self.project_start_date and self.project_end_date:
            total_days = (self.project_end_date - self.project_start_date).days
            canvas_width = total_days * self.pixels_per_day
            self.setMinimumWidth(int(canvas_width))
        self.update()

    def mouseDoubleClickEvent(self, event):
        pos_f = QPointF(event.pos())
        for task in self.tasks_to_display:
            if task.bar_rect.contains(pos_f): self.task_double_clicked.emit(task); return

    def mouseMoveEvent(self, event):
        current_hovered_task = None
        for task in self.tasks_to_display:
            if task.bar_rect.contains(QPointF(event.pos())):
                current_hovered_task = task
                break
        
        if current_hovered_task != self.last_hovered_task:
            self.last_hovered_task = current_hovered_task
            if current_hovered_task:
                tooltip_text = ""
                if hasattr(current_hovered_task, 'is_ecn_item') and current_hovered_task.is_ecn_item:
                    # Get all items under this ECN/PR
                    items = []
                    if isinstance(current_hovered_task, ecn_engine.Item):
                        items = [current_hovered_task]
                    elif isinstance(current_hovered_task, ecn_engine.ECN):
                        items = current_hovered_task.items
                    elif isinstance(current_hovered_task, ecn_engine.ProblemReport):
                        for ecn in current_hovered_task.ecns.values():
                            items.extend(ecn.items)
                    
                    # 1. ECN Description
                    if items:
                        ecn_name = items[0].raw_data.get('ECN Name', 'N/A')
                        tooltip_text += f"Description: {ecn_name}\n"
                    
                    # 2. Collect unique design performers
                    designers = set()
                    for item in items:
                        performer = item.raw_data.get('ECN Design Task Performer', '')
                        if performer and pd.notna(performer):
                            designers.add(str(performer).strip())
                    
                    if designers:
                        tooltip_text += f"Designer(s): {', '.join(sorted(designers))}\n"
                    
                    # 3. Progress
                    progress_attr = 'effective_progress' if hasattr(current_hovered_task, 'effective_progress') else 'progress'
                    tooltip_text += f"Progress: {getattr(current_hovered_task, progress_attr, 0):.0f}%\n"

                    # 3b. Current step info (for individual Items)
                    if isinstance(current_hovered_task, ecn_engine.Item):
                        step_name = getattr(current_hovered_task, 'current_step_name', '')
                        if step_name:
                            tooltip_text += f"Current Step: {step_name}\n"
                        step_performer = getattr(current_hovered_task, 'current_step_performer', '')
                        if step_performer:
                            tooltip_text += f"Responsible: {step_performer}\n"
                        completed = getattr(current_hovered_task, 'completed_steps', 0)
                        total = getattr(current_hovered_task, 'total_steps', 0)
                        if total > 0:
                            tooltip_text += f"Steps: {completed}/{total}\n"

                    # 4. Latest comment
                    if hasattr(current_hovered_task, 'comments') and current_hovered_task.comments:
                        latest = current_hovered_task.comments[-1]
                        tooltip_text += f"Latest Comment: {latest['text']}"
                else:
                    tooltip_text += f"Name: {current_hovered_task.name}\n"
                    progress_attr = 'effective_progress' if hasattr(current_hovered_task, 'effective_progress') else 'progress'
                    tooltip_text += (
                        f"Start: {current_hovered_task.start.strftime('%Y-%m-%d')}\n"
                        f"Finish: {current_hovered_task.finish.strftime('%Y-%m-%d')}\n"
                        f"Progress: {getattr(current_hovered_task, progress_attr, 0):.0f}%"
                    )
                self.setToolTip(tooltip_text.strip())
            else:
                self.setToolTip("")

    def contextMenuEvent(self, event):
        pos = event.pos()
        task_found = None
        for task in self.tasks_to_display:
            if task.bar_rect.contains(QPointF(pos)):
                task_found = task
                break
        
        if not task_found:
             return

        menu = QMenu(self)

        # Build menu additively
        if hasattr(task_found, 'is_ecn_item') and task_found.is_ecn_item and not isinstance(task_found, ecn_engine.Item):
            summary_action = QAction("Show Summary...", self)
            summary_action.triggered.connect(lambda: self.summary_requested.emit(task_found))
            menu.addAction(summary_action)

        if hasattr(task_found, 'comments'):
            comment_action = QAction("Comments & Reminders...", self)
            comment_action.triggered.connect(lambda: self.comment_requested.emit(task_found))
            menu.addAction(comment_action)

        if isinstance(task_found, Task):
            edit_action = QAction("Edit Item...", self)
            edit_action.triggered.connect(lambda: self.edit_task_requested.emit(task_found))
            menu.addAction(edit_action)
        elif isinstance(task_found, Project):
            edit_action = QAction("Edit Project Name...", self)
            edit_action.triggered.connect(lambda: self.edit_project_requested.emit(task_found))
            menu.addAction(edit_action)

        # Add to Notes View - works for ANY task type
        if menu.actions():
            menu.addSeparator()
        add_to_notes_action = QAction("📝 Add to Notes View", self)
        add_to_notes_action.triggered.connect(lambda: self.add_to_notes_requested.emit(task_found))
        menu.addAction(add_to_notes_action)

        if menu.actions():
             menu.addSeparator()

        if hasattr(task_found, 'is_active'):
            active_zone_action = QAction("Remove from Active Zones" if task_found.is_active else "Set as Active Zone", self)
            active_zone_action.triggered.connect(lambda: self.active_zone_changed.emit(task_found, not task_found.is_active))
            menu.addAction(active_zone_action)

        if isinstance(task_found, Task):
            if task_found.is_milestone:
                action = QAction("Mark as Complete" if task_found.progress < 100 else "Mark as Incomplete", self)
                action.triggered.connect(partial(self.set_task_progress, task_found, 100 if task_found.progress < 100 else 0))
                menu.addAction(action)
            elif task_found.children:
                action = QAction("Mark Phase as Complete", self)
                action.triggered.connect(lambda: self.phase_complete_requested.emit(task_found))
                menu.addAction(action)
            else:
                progress_menu = QMenu("Set Progress", self)
                for level in [0, 25, 50, 75, 100]:
                    action = QAction(f"{level}%", self)
                    action.triggered.connect(partial(self.set_task_progress, task_found, level))
                    progress_menu.addAction(action)
                menu.addMenu(progress_menu)
        
        if menu.actions():
            menu.addSeparator()

        delete_action = QAction("Delete Item...", self)
        delete_action.triggered.connect(lambda: self.delete_task_requested.emit(task_found))
        menu.addAction(delete_action)
        
        if menu.actions():
            menu.exec(self.mapToGlobal(pos))

    def set_task_progress(self, task, progress_level): task.progress = progress_level; self.progress_changed.emit()

    def get_x_for_date(self, date):
        if not self.project_start_date: return 0
        return (date - self.project_start_date).days * self.pixels_per_day

    def _item_milestone(self, item):
        """Returns milestone level for an Item: 3=MCN, 2=ProdReleased, 1=LeadEngApproved, 0=Design"""
        row = item.raw_data
        if pd.notna(pd.to_datetime(row.get('MCN Released Date'), errors='coerce')):
            return 3
        if pd.notna(pd.to_datetime(row.get('Item Production Released Date'), errors='coerce')):
            return 2
        # Check Lead Eng Review (1) end date — approved if completed
        lead_end = self.parent().ecn_engine.find_flexible_column(row.keys(), 'Item Lead Eng End Date (1)') if hasattr(self, 'parent') and callable(self.parent) and self.parent() and hasattr(self.parent(), 'ecn_engine') else None
        if lead_end is None:
            # Fallback: try direct key lookup
            for key in row.keys():
                if 'lead eng end date' in key.lower():
                    lead_end = key
                    break
        if lead_end and pd.notna(pd.to_datetime(row.get(lead_end), errors='coerce')):
            return 1
        return 0

    _MILESTONE_COLORS = {
        3: "#28a745",  # Green — MCN Released
        2: "#007bff",  # Blue — Production Released
        1: "#e67e22",  # Dark Orange — Lead Eng Approved
        0: "#ffc107",  # Yellow — Design Phase
    }

    def get_task_color(self, task):
        if hasattr(task, 'is_ecn_item') and task.is_ecn_item:
            # Milestone-based coloring: Yellow → Orange → Blue → Green
            # Rejection uses red outline only (handled in paintEvent), not fill color

            if isinstance(task, ecn_engine.Item):
                level = self._item_milestone(task)
                return QColor(self._MILESTONE_COLORS[level])

            elif isinstance(task, ecn_engine.ECN):
                if task.items:
                    min_level = min(self._item_milestone(i) for i in task.items)
                    return QColor(self._MILESTONE_COLORS[min_level])
                return QColor("#ffc107")

            elif isinstance(task, ecn_engine.ProblemReport):
                all_ecns = list(task.ecns.values())
                if all_ecns:
                    levels = []
                    for ecn in all_ecns:
                        if ecn.items:
                            levels.append(min(self._item_milestone(i) for i in ecn.items))
                        else:
                            levels.append(0)
                    min_level = min(levels)
                    return QColor(self._MILESTONE_COLORS[min_level])
                return QColor("#ffc107")

            # Fallback for any other ECN-type object
            return QColor("#ffc107")

        today = datetime.date.today()
        if hasattr(task, 'progress') and task.progress >= 100:
             return QColor("#28a745")
        if hasattr(task, 'effective_progress') and task.effective_progress >= 100:
            return QColor("#28a745")
        
        if task.finish and task.finish.date() < today and not getattr(task, 'is_milestone', False):
            return QColor("#dc3545")
        
        priority_colors = {Priority.LOW: QColor("#17a2b8"), Priority.MEDIUM: QColor("#007bff"), Priority.HIGH: QColor("#ffc107")}
        
        if isinstance(task, Project):
            return QColor("#2c3e50")
        if task.children:
            return QColor("#5d6d7e")

        return priority_colors.get(task.priority, QColor("#007bff"))

    def paintEvent(self, event):
        painter = QPainter(self); painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if not self.project_start_date or not self.tasks_to_display:
            viewport_rect = self.parentWidget().rect() if self.parentWidget() else self.rect()
            painter.drawText(viewport_rect, Qt.AlignmentFlag.AlignCenter, "Import a project file to begin.")
            return
        visible_rect = event.rect()
        self.draw_header_and_grid(painter, visible_rect)
        self.draw_task_bars_and_labels(painter, visible_rect)
        self.draw_today_line(painter, visible_rect)

    def draw_today_line(self, painter, visible_rect):
        today = datetime.date.today()
        x_pos = self.get_x_for_date(today)
        if visible_rect.left() <= x_pos <= visible_rect.right():
            painter.setPen(QPen(QColor("#d9534f"), 1, Qt.PenStyle.DashLine))
            painter.drawLine(int(x_pos), self.header_height, int(x_pos), self.height())

    def draw_header_and_grid(self, painter, visible_rect):
        painter.fillRect(visible_rect.left(), 0, visible_rect.width(), self.header_height, QColor("#f0f0f0"))
        painter.setPen(QColor("#d0d0d0")); painter.drawLine(visible_rect.left(), self.header_height, visible_rect.right(), self.header_height)
        month_font, day_font = QFont("Arial", 9, QFont.Weight.Bold), QFont("Arial", 8)
        start_day_offset = int(visible_rect.left() / self.pixels_per_day) if self.pixels_per_day > 0 else 0
        end_day_offset = int(visible_rect.right() / self.pixels_per_day) + 1 if self.pixels_per_day > 0 else 1
        current_date = self.project_start_date + datetime.timedelta(days=start_day_offset)
        end_date_to_draw = self.project_start_date + datetime.timedelta(days=end_day_offset)
        last_drawn_month = -1
        while current_date <= end_date_to_draw:
            x_pos = self.get_x_for_date(current_date)
            if self.timescale_view != TimescaleView.YEAR:
                if current_date.month != last_drawn_month:
                    painter.setFont(month_font); painter.setPen(QColor("#000000"))
                    painter.drawText(int(x_pos) + 5, 20, f"{current_date.strftime('%B ''%y')}")
                    painter.setPen(QColor("#d0d0d0")); painter.drawLine(int(x_pos), 0, int(x_pos), self.header_height)
                    last_drawn_month = current_date.month
                painter.setFont(day_font); painter.drawText(int(x_pos) + 3, 40, f"{current_date.day}")
                painter.setPen(QColor("#e0e0e0")); painter.drawLine(int(x_pos), self.header_height, int(x_pos), self.height())
            else:
                 if current_date.month != last_drawn_month:
                    painter.setFont(month_font); painter.setPen(QColor("#000000"))
                    painter.drawText(int(x_pos) + 5, 30, f"{current_date.strftime('%b ''%y')}")
                    painter.setPen(QColor("#e0e0e0")); painter.drawLine(int(x_pos), self.header_height, int(x_pos), self.height())
                    last_drawn_month = current_date.month
            current_date += datetime.timedelta(days=1)

    def draw_task_bars_and_labels(self, painter, visible_rect):
        y_pos, font = self.header_height + 10, QFont("Arial", 9); painter.setFont(font)
        comment_icon = "💬"
        self._badge_rects = []  # Reset badge rects each paint
        for item in self.tasks_to_display:
            painter.setOpacity(1.0 if getattr(item, 'is_active', True) else 0.4)
            if not item.start or not item.finish: continue
            
            needs_attention = getattr(item, 'needs_attention', False)
            is_highlighted = (item == self.highlighted_task)

            item_start_date, item_end_date = item.start.date(), item.finish.date()
            bar_x = self.get_x_for_date(item_start_date)
            duration_days = (item_end_date - item_start_date).days + 1
            # Ensure minimum 30-day bar width for better readability
            min_duration_days = max(duration_days, 30)
            bar_width = min_duration_days * self.pixels_per_day
            if bar_x + bar_width < visible_rect.left() and bar_x > visible_rect.right():
                y_pos += self.row_height; continue
            item.bar_rect = QRectF()
            if getattr(item, 'is_milestone', False):
                milestone_x = self.get_x_for_date(item_start_date)
                center_y, diamond_size = y_pos + (self.row_height * self.bar_height_ratio) / 2, 12
                poly = QPolygonF([QPointF(milestone_x, center_y - diamond_size / 2), QPointF(milestone_x + diamond_size / 2, center_y), QPointF(milestone_x, center_y + diamond_size / 2), QPointF(milestone_x - diamond_size / 2, center_y)])
                
                color = self.get_task_color(item)
                painter.setBrush(color if item.progress == 100 else Qt.GlobalColor.white)
                
                pen = QPen(color)
                if is_highlighted:
                    pen.setColor(QColor("#ff6b35"))  # Bright orange for highlight
                    pen.setWidth(4)
                elif needs_attention:
                    pen.setColor(QColor("#dc3545"))
                    pen.setWidth(2)
                else:
                    pen.setWidth(1)
                painter.setPen(pen)

                painter.drawPolygon(poly)
                
                item.bar_rect = QRectF(milestone_x - diamond_size/2, center_y - diamond_size/2, diamond_size, diamond_size)
                painter.setPen(QColor("#212529")); painter.drawText(int(milestone_x + diamond_size), int(y_pos + self.row_height * 0.7), item.name)
                if getattr(item, 'comments', []): painter.drawText(int(milestone_x + diamond_size + painter.fontMetrics().horizontalAdvance(item.name) + 5), int(y_pos + self.row_height * 0.7), comment_icon)
            else:
                bar_rect = QRectF(bar_x, y_pos, bar_width, self.row_height * self.bar_height_ratio)
                item.bar_rect = bar_rect
                bar_color = self.get_task_color(item)
                
                # Draw highlight background if this is the highlighted task
                if is_highlighted:
                    highlight_rect = QRectF(bar_x - 3, y_pos - 3, bar_width + 6, self.row_height * self.bar_height_ratio + 6)
                    painter.setBrush(QColor("#ff6b35"))
                    painter.setPen(QPen(QColor("#ff6b35"), 4))
                    painter.drawRoundedRect(highlight_rect, 8, 8)
                
                painter.setBrush(bar_color)

                # Determine outline style
                if is_highlighted:
                    pen = QPen(QColor("#ff6b35"))
                    pen.setWidth(4)
                    painter.setPen(pen)
                elif needs_attention:
                    # For ECN items: Red outline when item has pending MCN tasks or is rejected
                    if hasattr(item, 'is_ecn_item') and item.is_ecn_item:
                        if getattr(item, 'has_pending_mcn_task', False) or getattr(item, 'is_rejected', False):
                            pen = QPen(QColor("#dc3545"))
                            pen.setWidth(2)
                            painter.setPen(pen)
                        else:
                            painter.setPen(Qt.PenStyle.NoPen)
                    else:
                        # For non-ECN items with needs_attention
                        pen = QPen(QColor("#dc3545"))
                        pen.setWidth(2)
                        painter.setPen(pen)
                else:
                    painter.setPen(Qt.PenStyle.NoPen)
                
                painter.drawRoundedRect(bar_rect, 5, 5)

                painter.setPen(Qt.PenStyle.NoPen) 
                progress = getattr(item, 'effective_progress', item.progress)
                if progress > 0:
                    painter.setBrush(bar_color.lighter(120))
                    painter.drawRoundedRect(QRectF(bar_x, y_pos, bar_width * (progress / 100.0), self.row_height * self.bar_height_ratio), 5, 5)
                if bar_width > 15:
                    painter.setPen(Qt.GlobalColor.white); text_rect = QRectF(bar_rect); text_rect.adjust(5, 0, -5, 0)
                    painter.drawText(text_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, item.name)
                if getattr(item, 'comments', []): painter.drawText(int(bar_x + bar_width + 5), int(y_pos + self.row_height * 0.7), comment_icon)

            # ⚠️ BOM Blocking Badge
            blocking_info = self._get_blocking_for_task(item)
            if blocking_info and (blocking_info.get("blocking_children") or not blocking_info.get("is_released")):
                badge_font = QFont("Segoe UI Emoji", 10)
                painter.setFont(badge_font)
                painter.setPen(QColor("#e67e22"))
                # Position badge to the right of the bar
                badge_x = item.bar_rect.right() + 3
                comment_offset = 20 if getattr(item, 'comments', []) else 0
                badge_x += comment_offset
                badge_y = int(y_pos + self.row_height * 0.72)
                painter.drawText(int(badge_x), badge_y, "⚠️")
                # Count of blocking children
                bc_count = len(blocking_info.get("blocking_children", []))
                if bc_count > 0:
                    count_font = QFont("Arial", 7, QFont.Weight.Bold)
                    painter.setFont(count_font)
                    painter.setPen(QColor("#dc3545"))
                    painter.drawText(int(badge_x + 18), badge_y, str(bc_count))
                # Store badge rect for click detection
                badge_rect = QRectF(badge_x - 2, y_pos, 30, self.row_height)
                self._badge_rects.append((badge_rect, item, blocking_info))
                painter.setFont(font)  # Restore original font

            y_pos += self.row_height
        painter.setOpacity(1.0)

# ============================================================================
# BOM EXPLORER VIEW
# ============================================================================

class BomExplorerView(QWidget):
    """TC-like tree view of case BOMs with release status, ECN, and workflow visibility."""

    COL_ITEM = 0
    COL_NAME = 1
    COL_QTY = 2
    COL_WORKFLOW = 3
    COL_ECN = 4
    COL_RELEASED = 5
    COL_ASSIGNEE = 6
    COL_TASK = 7
    COL_LINK = 8

    HEADERS = ["Item ID / Rev", "Name", "Qty", "Workflow", "ECN", "Status", "Assignee", "Current Task", "Link Status"]

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self._main_window = main_window
        self._bom_database = {}  # {nomenclature: bom_result}
        self._link_check_data = {}  # {item_id: link_info} populated by _on_link_check()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # --- Toolbar ---
        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(8, 6, 8, 6)

        add_btn = QPushButton("➕ Add Case")
        add_btn.setStyleSheet("padding: 5px 12px; font-weight: bold;")
        add_btn.clicked.connect(self._on_add_case)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("🔄 Refresh All")
        refresh_btn.setStyleSheet("padding: 5px 12px;")
        refresh_btn.clicked.connect(self._on_refresh_all)
        toolbar.addWidget(refresh_btn)

        export_btn = QPushButton("💾 Export CSV")
        export_btn.setStyleSheet("padding: 5px 12px;")
        export_btn.clicked.connect(self._on_export_csv)
        toolbar.addWidget(export_btn)

        toolbar.addSpacing(10)

        missing_btn = QPushButton("🔍 Find Missing BOMs")
        missing_btn.setStyleSheet("padding: 5px 12px; color: #7719AA; font-weight: bold;")
        missing_btn.clicked.connect(self._on_find_missing_boms)
        toolbar.addWidget(missing_btn)

        link_btn = QPushButton("🔗 Link Check")
        link_btn.setStyleSheet("padding: 5px 12px; color: #0066cc; font-weight: bold;")
        link_btn.clicked.connect(self._on_link_check)
        toolbar.addWidget(link_btn)

        toolbar.addSpacing(20)

        toolbar.addWidget(QLabel("Filter:"))
        self._filter_combo = QComboBox()
        self._filter_combo.addItems(["All", "Blocked Only", "Unreleased Only", "In ECN"])
        self._filter_combo.currentIndexChanged.connect(self._apply_filter)
        toolbar.addWidget(self._filter_combo)

        toolbar.addSpacing(10)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("🔍 Search part, name, or ECN...")
        self._search_edit.setMaximumWidth(250)
        self._search_edit.textChanged.connect(self._apply_filter)
        toolbar.addWidget(self._search_edit)

        toolbar.addStretch()

        toolbar_frame = QFrame()
        toolbar_frame.setLayout(toolbar)
        toolbar_frame.setStyleSheet("background-color: #f5f5f5; border-bottom: 1px solid #ddd;")
        layout.addWidget(toolbar_frame)

        # --- Progress bar (hidden by default) ---
        self._progress_bar = QProgressBar()
        self._progress_bar.setVisible(False)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setMaximumHeight(22)
        layout.addWidget(self._progress_bar)

        # --- Tree widget ---
        self._tree = QTreeWidget()
        self._tree.setHeaderLabels(self.HEADERS)
        self._tree.setColumnCount(len(self.HEADERS))
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(True)
        self._tree.setUniformRowHeights(True)
        self._tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        self._tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._tree.customContextMenuRequested.connect(self._on_context_menu)
        self._tree.setStyleSheet("""
            QTreeWidget {
                font-size: 12px;
                font-family: 'Consolas', 'Courier New', monospace;
                border: none;
            }
            QTreeWidget::item {
                padding: 2px 4px;
            }
            QHeaderView::section {
                background-color: #7719AA;
                color: white;
                padding: 4px 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        # Set column widths
        self._tree.setColumnWidth(self.COL_ITEM, 150)
        self._tree.setColumnWidth(self.COL_NAME, 300)
        self._tree.setColumnWidth(self.COL_WORKFLOW, 120)
        self._tree.setColumnWidth(self.COL_ECN, 100)
        self._tree.setColumnWidth(self.COL_RELEASED, 100)
        self._tree.setColumnWidth(self.COL_ASSIGNEE, 120)
        self._tree.setColumnWidth(self.COL_TASK, 150)
        self._tree.setColumnWidth(self.COL_LINK, 250)
        layout.addWidget(self._tree, 1)

        # --- Status bar ---
        self._status_label = QLabel("No cases loaded. Click 'Add Case' to fetch a BOM from Teamcenter.")
        self._status_label.setStyleSheet("padding: 4px 10px; background: #f0f0f0; border-top: 1px solid #ddd; font-size: 11px; color: #555;")
        layout.addWidget(self._status_label)

        # --- Empty state ---
        self._empty_label = QLabel("No cases loaded.\nClick '➕ Add Case' to fetch a BOM from Teamcenter.")
        self._empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_label.setStyleSheet("font-size: 16px; color: #999; padding: 60px;")
        layout.addWidget(self._empty_label)
        self._empty_label.setVisible(True)
        self._tree.setVisible(False)

    # ------------------------------------------------------------------
    # PUBLIC: Load saved cases from config
    # ------------------------------------------------------------------

    def load_saved_cases(self, bom_data_dict):
        """Restore BOM data from saved config (dict of {nom: bom_result})."""
        if not bom_data_dict:
            return
        self._bom_database = bom_data_dict
        self._rebuild_tree()

    def get_bom_database(self):
        """Return current BOM database for saving."""
        return self._bom_database

    def get_saved_case_names(self):
        """Return list of case nomenclatures currently loaded."""
        return list(self._bom_database.keys())

    # ------------------------------------------------------------------
    # ECN DATA LOOKUP
    # ------------------------------------------------------------------

    def _get_ecn_item_lookup(self):
        """Build {item_id: ecn_engine.Item} lookup from already-loaded ECN projects.

        This reuses data loaded via the ECN Data dropdown (load_ecn_data), which
        populates ecn_projects with Item objects that have current_step_name,
        current_step_performer, effective_progress, etc.
        """
        lookup = {}
        ecn_projects = getattr(self._main_window, 'ecn_projects', []) or []
        for pr in ecn_projects:
            ecns = getattr(pr, 'ecns', {}) or {}
            for ecn in ecns.values():
                items = getattr(ecn, 'items', []) or []
                for item in items:
                    # Item.id is the part number (e.g., "3243443")
                    item_id = str(getattr(item, 'id', '')).strip()
                    if item_id:
                        lookup[item_id] = item
                    # Also index by name which is "3243443/D" → extract "3243443"
                    name_id = str(getattr(item, 'name', '')).split("/")[0].strip()
                    if name_id and name_id not in lookup:
                        lookup[name_id] = item
        return lookup

    # ------------------------------------------------------------------
    # LINK CHECK & MISSING BOM DETECTION
    # ------------------------------------------------------------------

    def _build_item_to_nomenclature_map(self):
        """Scan all loaded BOMs, map every item_id → top-level nomenclature."""
        item_to_nom = {}
        for nom, bom_result in self._bom_database.items():
            if bom_result is None:
                continue
            nodes = (bom_result or {}).get("bom", {}).get("nodes", []) or []
            for node in nodes:
                item_id = node.get("item_id", "")
                if item_id:
                    item_to_nom[item_id] = nom
        return item_to_nom

    def _on_find_missing_boms(self):
        """Show dialog listing ECN items not in any loaded BOM, grouped by case."""
        ecn_lookup = self._get_ecn_item_lookup()
        if not ecn_lookup:
            QMessageBox.information(self, "Find Missing BOMs",
                                    "No ECN data loaded. Load ECN data first via the ECN Data menu.")
            return

        item_to_nom = self._build_item_to_nomenclature_map()

        # Categorize: found (in a BOM) vs missing (not in any BOM)
        found = {}    # {item_id: nomenclature}
        missing = {}  # {item_id: Item}
        for item_id, item in ecn_lookup.items():
            if item_id in item_to_nom:
                found[item_id] = item_to_nom[item_id]
            else:
                missing[item_id] = item

        if not missing:
            QMessageBox.information(self, "Find Missing BOMs",
                                    f"All {len(found)} ECN items are covered by loaded BOMs. No missing items.")
            return

        # Group missing items by ECN number for display
        by_ecn = {}
        for item_id, item in missing.items():
            ecn_num = getattr(item, 'ecn_number', 'Unknown')
            by_ecn.setdefault(ecn_num, []).append((item_id, item))

        # Build dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Missing BOMs — {len(missing)} items not in any loaded BOM")
        dialog.setMinimumSize(700, 500)
        dlayout = QVBoxLayout(dialog)

        info_label = QLabel(f"<b>{len(found)}</b> ECN items found in loaded BOMs. "
                            f"<b>{len(missing)}</b> items missing (grouped by ECN).<br>"
                            f"Check items to fetch, enter nomenclature if needed, then click Fetch.")
        info_label.setWordWrap(True)
        dlayout.addWidget(info_label)

        # Table: checkbox | item_id | name | ECN | nomenclature input
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Fetch", "Item ID", "Name", "ECN", "Nomenclature (enter case)"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setColumnWidth(0, 50)
        table.setColumnWidth(1, 120)
        table.setColumnWidth(2, 200)
        table.setColumnWidth(3, 160)

        rows = []
        for ecn_num in sorted(by_ecn.keys()):
            for item_id, item in sorted(by_ecn[ecn_num], key=lambda x: x[0]):
                rows.append((item_id, item, ecn_num))

        table.setRowCount(len(rows))
        check_boxes = []
        nom_edits = []
        for i, (item_id, item, ecn_num) in enumerate(rows):
            cb = QTableWidgetItem()
            cb.setFlags(cb.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            cb.setCheckState(Qt.CheckState.Checked)
            table.setItem(i, 0, cb)
            check_boxes.append(cb)

            table.setItem(i, 1, QTableWidgetItem(str(getattr(item, 'name', item_id))))
            table.setItem(i, 2, QTableWidgetItem(str(getattr(item, 'id', ''))))
            table.setItem(i, 3, QTableWidgetItem(ecn_num))

            nom_edit = QLineEdit()
            nom_edit.setPlaceholderText("e.g. RLN4MA")
            table.setCellWidget(i, 4, nom_edit)
            nom_edits.append(nom_edit)

        dlayout.addWidget(table, 1)

        # Buttons
        btn_layout = QHBoxLayout()
        fetch_btn = QPushButton("Fetch Selected")
        fetch_btn.setStyleSheet("padding: 8px 20px; font-weight: bold; background-color: #7719AA; color: white;")
        cancel_btn = QPushButton("Close")
        cancel_btn.setStyleSheet("padding: 8px 20px;")
        btn_layout.addStretch()
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(fetch_btn)
        dlayout.addLayout(btn_layout)

        cancel_btn.clicked.connect(dialog.reject)

        def _do_fetch():
            # Collect unique nomenclatures to fetch
            noms_to_fetch = set()
            for i, (item_id, item, ecn_num) in enumerate(rows):
                if check_boxes[i].checkState() == Qt.CheckState.Checked:
                    nom_text = nom_edits[i].text().strip()
                    if nom_text:
                        noms_to_fetch.add(nom_text.upper())
            if not noms_to_fetch:
                QMessageBox.warning(dialog, "No Nomenclatures",
                                    "Enter nomenclature for at least one checked item.")
                return
            dialog.accept()
            # Fetch using existing mechanism
            self._fetch_cases_by_name(list(noms_to_fetch))

        fetch_btn.clicked.connect(_do_fetch)
        dialog.exec()

    def _fetch_cases_by_name(self, case_names):
        """Fetch BOM cases by nomenclature list, reusing existing TC fetch logic."""
        from tc_connector import TcLoginDialog, TcEcnDataFetcher
        client = getattr(self._main_window, '_tc_client', None)
        if not client or not getattr(client, 'is_connected', False):
            tc_config = self._main_window.ecn_config.get("teamcenter", {})
            login_dialog = TcLoginDialog(tc_config, self._main_window)
            if login_dialog.exec() != QDialog.DialogCode.Accepted:
                return
            client = login_dialog.get_client()
            self._main_window._tc_client = client
        self._fetch_cases(client, case_names)

    def _on_link_check(self):
        """Run link check: overlay ECN workflow stage + release status on BOM tree."""
        ecn_lookup = self._get_ecn_item_lookup()
        if not self._bom_database:
            QMessageBox.information(self, "Link Check", "No BOM data loaded. Add cases first.")
            return

        # ECN Completed step position — count steps up to and including "ECN Completed"
        # From workflow_config, ECN Completed is typically step index ~14-15
        # We use completed_steps from the Item object: if completed_steps >= ecn_completed_pos, it's done
        ECN_COMPLETED_STEP_NAME = "ECN Completed"

        link_data = {}  # {item_id: {status, text, color, blocked_children_count}}

        for nom, bom_result in self._bom_database.items():
            if bom_result is None:
                continue
            bom = bom_result.get("bom", {}) or {}
            nodes = bom.get("nodes", []) or []
            edges = bom.get("edges", []) or []
            if not nodes:
                continue

            node_by_uid = {n["uid"]: n for n in nodes if isinstance(n, dict) and n.get("uid")}
            children_of = {}
            for e in edges:
                if not isinstance(e, dict):
                    continue
                p = e.get("parent_uid", "")
                c = e.get("child_uid", "")
                if p and c:
                    children_of.setdefault(p, []).append(c)

            def _get_all_descendants(uid):
                desc = []
                for child_uid in children_of.get(uid, []):
                    desc.append(child_uid)
                    desc.extend(_get_all_descendants(child_uid))
                return desc

            def _is_prod_released(node):
                release_text = self._release_text(node.get("release_status_list", ""))
                return "PRODUCTION RELEASED" in release_text.upper()

            def _is_ecn_done(item_id):
                """Check if item's ECN workflow has reached or passed ECN Completed."""
                ecn_item = ecn_lookup.get(item_id)
                if not ecn_item:
                    return False  # No ECN data — cannot confirm done
                step_name = getattr(ecn_item, 'current_step_name', '') or ''
                if step_name.upper() in ('COMPLETED',):
                    return True
                # Check if completed_steps indicates we're past ECN Completed
                # The workflow_config has ECN Completed around position 15
                # If the item has MCN-level steps active, it's definitely past ECN Completed
                completed = getattr(ecn_item, 'completed_steps', 0)
                total = getattr(ecn_item, 'total_steps', 0)
                # Heuristic: if progress > 60% or completed_steps > 10, likely past ECN Completed
                # More precise: check if current step is in the MCN section
                mcn_steps = {'MCN Created', 'Supply Chain', 'MFG ENGG', 'Copper Programmer',
                             'Sheet Metal Programmer', 'Production Control SM', 'First Article (FAI)',
                             'Costing', 'PPAP Needed', 'Set Effectivity', 'MCN Released',
                             'Item Production Released', 'MFG Operations', 'Analyst Review',
                             'Costing Rework', 'PPAP Update', 'PPAP Update MTY', 'Sourceability',
                             'Plant Coding Sourcing', 'Part Implementation Completed'}
                if step_name in mcn_steps:
                    return True
                # If step is "ECN Completed" itself, it's done
                if step_name == ECN_COMPLETED_STEP_NAME:
                    return True
                return False

            # Analyze each node
            for node in nodes:
                uid = node.get("uid", "")
                item_id = node.get("item_id", "")
                if not uid or not item_id:
                    continue

                is_released = _is_prod_released(node)
                ecn_item = ecn_lookup.get(item_id)
                ecn_done = is_released or _is_ecn_done(item_id)

                # Count blocked children
                desc_uids = _get_all_descendants(uid)
                blocked_children = []
                for d_uid in desc_uids:
                    d_node = node_by_uid.get(d_uid, {})
                    d_item_id = d_node.get("item_id", "")
                    if not d_item_id:
                        continue
                    d_released = _is_prod_released(d_node)
                    d_ecn_done = d_released or _is_ecn_done(d_item_id)
                    if not d_released and not d_ecn_done:
                        d_ecn_item = ecn_lookup.get(d_item_id)
                        d_step = getattr(d_ecn_item, 'current_step_name', '') if d_ecn_item else ''
                        d_performer = getattr(d_ecn_item, 'current_step_performer', '') if d_ecn_item else ''
                        d_progress = getattr(d_ecn_item, 'effective_progress', 0) if d_ecn_item else 0
                        blocked_children.append({
                            "item_id": d_item_id,
                            "rev": d_node.get("item_revision_id", ""),
                            "step": d_step,
                            "performer": d_performer,
                            "progress": d_progress,
                        })

                # Determine link status text and color
                if is_released:
                    status_text = "✓ PROD RELEASED"
                    status_color = "#D4EDDA"  # green
                elif ecn_done:
                    status_text = "✓ ECN Completed"
                    status_color = "#D4EDDA"
                elif ecn_item:
                    step = getattr(ecn_item, 'current_step_name', '') or 'Unknown'
                    performer = getattr(ecn_item, 'current_step_performer', '') or ''
                    progress = getattr(ecn_item, 'effective_progress', 0)
                    if getattr(ecn_item, 'is_rejected', False):
                        status_text = f"✗ Rejected at {step}"
                        if performer:
                            status_text += f" — {performer}"
                        status_color = "#F8D7DA"  # red
                    else:
                        status_text = f"⏳ {step} ({progress:.0f}%)"
                        if performer:
                            status_text += f" — {performer}"
                        status_color = "#FFF3CD"  # yellow
                else:
                    status_text = "✗ Not in any ECN"
                    status_color = "#F8D7DA"  # red

                # Add blocked children info for parent nodes
                if blocked_children and (is_released or ecn_done):
                    n = len(blocked_children)
                    status_text = f"⚠ {n} child{'ren' if n > 1 else ''} blocked"
                    status_color = "#FFE0B2"  # orange
                elif blocked_children:
                    n = len(blocked_children)
                    status_text += f" | ⚠ {n} blocked"

                link_data[item_id] = {
                    "text": status_text,
                    "color": status_color,
                    "blocked_children": blocked_children,
                    "is_clear": is_released or (ecn_done and not blocked_children),
                }

        self._link_check_data = link_data
        self._rebuild_tree()

        # Summary
        total = len(link_data)
        clear = sum(1 for v in link_data.values() if v["is_clear"])
        blocked = total - clear
        QMessageBox.information(self, "Link Check Complete",
                                f"Analyzed {total} parts across {len(self._bom_database)} cases.\n\n"
                                f"✓ {clear} parts clear\n"
                                f"✗ {blocked} parts blocked or in progress\n\n"
                                f"Link Status column has been updated.")

    # ------------------------------------------------------------------
    # TOOLBAR ACTIONS
    # ------------------------------------------------------------------

    def _on_add_case(self):
        """Add case(s) by fetching BOM from Teamcenter."""
        from tc_connector import TcLoginDialog, TcEcnDataFetcher

        # Get TC client (reuse session or login)
        client = getattr(self._main_window, '_tc_client', None)
        if not client or not getattr(client, 'is_connected', False):
            tc_config = self._main_window.ecn_config.get("teamcenter", {})
            login_dialog = TcLoginDialog(tc_config, self._main_window)
            if login_dialog.exec() != QDialog.DialogCode.Accepted:
                return
            client = login_dialog.get_client()
            self._main_window._tc_client = client
            self._main_window.ecn_config["teamcenter"] = login_dialog.get_config()
            self._main_window.save_ecn_config()

        # Ask for nomenclature(s)
        text, ok = QInputDialog.getText(
            self, "Add Case(s)",
            "Case Nomenclature(s) (comma-separated):",
            QLineEdit.EchoMode.Normal,
            ""
        )
        if not ok or not text.strip():
            return

        cases = [c.strip().upper() for c in text.split(",") if c.strip()]
        self._fetch_cases(client, cases)

    def _on_refresh_all(self):
        """Refresh all cases using part-level diff — only update nodes that changed in TC."""
        if not self._bom_database:
            QMessageBox.information(self, "Nothing to Refresh", "No cases loaded yet.")
            return

        from tc_connector import TcLoginDialog
        client = getattr(self._main_window, '_tc_client', None)
        if not client or not getattr(client, 'is_connected', False):
            tc_config = self._main_window.ecn_config.get("teamcenter", {})
            login_dialog = TcLoginDialog(tc_config, self._main_window)
            if login_dialog.exec() != QDialog.DialogCode.Accepted:
                return
            client = login_dialog.get_client()
            self._main_window._tc_client = client

        # Attempt smart part-level refresh; fall back to full fetch if something goes wrong
        try:
            self._smart_refresh(client)
        except Exception as e:
            reply = QMessageBox.question(
                self, "Smart Refresh Failed",
                f"Part-level refresh encountered an error:\n{e}\n\nFall back to full re-fetch?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                cases = list(self._bom_database.keys())
                self._fetch_cases(client, cases, replace=True)

    def _smart_refresh(self, client):
        """
        For each cached case, batch-fetch current TC properties for every cached node UID
        and update only the nodes whose properties have changed.
        """
        from PyQt6.QtWidgets import QApplication

        # Properties to compare (must be in the node dict after a full fetch)
        COMPARE_PROPS = ["release_status_list", "process_stage", "h4_ECN_Number", "object_name"]

        self._progress_bar.setVisible(True)
        cases = [nom for nom, result in self._bom_database.items() if result is not None]
        self._progress_bar.setRange(0, max(len(cases), 1))
        self._progress_bar.setValue(0)

        summary_lines = []

        for i, nom in enumerate(cases):
            self._progress_bar.setFormat(f"Checking {nom}... ({i+1}/{len(cases)})")
            self._progress_bar.setValue(i)
            QApplication.processEvents()

            bom_result = self._bom_database[nom]
            nodes = bom_result.get("bom", {}).get("nodes", [])
            if not nodes:
                summary_lines.append(f"{nom}: no cached nodes, skipped")
                continue

            # Collect all UIDs from cached nodes
            all_uids = [n["uid"] for n in nodes if n.get("uid")]
            if not all_uids:
                summary_lines.append(f"{nom}: no UIDs, skipped")
                continue

            # Batch-fetch current properties from TC (50 UIDs per call)
            BATCH = 50
            tc_props = {}
            for start in range(0, len(all_uids), BATCH):
                batch = all_uids[start:start + BATCH]
                try:
                    result = client.get_properties(batch, COMPARE_PROPS)
                    if result:
                        tc_props.update(result)
                except Exception:
                    pass  # keep going with what we have

            # Diff each node against TC data
            updated_count = 0
            for node in nodes:
                uid = node.get("uid")
                if not uid or uid not in tc_props:
                    continue
                tc_node = tc_props[uid]
                changed = False
                for prop in COMPARE_PROPS:
                    old_val = node.get(prop, "")
                    new_val = tc_node.get(prop, "")
                    # Normalise to string for comparison
                    if str(old_val).strip() != str(new_val).strip():
                        node[prop] = new_val
                        changed = True
                if changed:
                    updated_count += 1

            unchanged_count = len(nodes) - updated_count
            summary_lines.append(f"{nom}: {updated_count} updated, {unchanged_count} unchanged")

        self._progress_bar.setValue(len(cases))
        self._progress_bar.setVisible(False)

        # Persist updated cache and rebuild tree
        self._write_bom_cache_files()
        self._rebuild_tree()
        try:
            self.push_blocking_to_timeline()
        except Exception:
            pass

        QMessageBox.information(
            self, "Refresh Complete",
            "Part-level refresh results:\n\n" + "\n".join(summary_lines)
        )

    def _fetch_cases(self, client, cases, replace=False):
        """Fetch BOM for one or more cases from TC."""
        from tc_connector import TcEcnDataFetcher

        self._progress_bar.setVisible(True)
        self._progress_bar.setRange(0, len(cases))
        self._progress_bar.setValue(0)

        from PyQt6.QtWidgets import QApplication

        fetcher = TcEcnDataFetcher(client)

        for i, nom in enumerate(cases):
            self._progress_bar.setFormat(f"Fetching {nom}... ({i+1}/{len(cases)})")
            self._progress_bar.setValue(i)
            QApplication.processEvents()

            try:
                bom_result = fetcher.fetch_bom_by_nomenclature(nom, recursive=True)
                if bom_result and bom_result.get("bom", {}).get("nodes"):
                    self._bom_database[nom.upper()] = bom_result
                    # Save tree files
                    try:
                        fetcher.save_bom_tree_files(bom_result)
                    except Exception:
                        pass
                else:
                    QMessageBox.warning(self, "Not Found", f"No BOM data found for '{nom}'.")
            except Exception as e:
                QMessageBox.warning(self, "Fetch Error", f"Error fetching '{nom}': {e}")

        self._progress_bar.setValue(len(cases))
        self._progress_bar.setVisible(False)

        # Save case list to config
        self._save_cases_to_config()
        self._rebuild_tree()

        # Push blocking data to timeline for ⚠️ badges
        try:
            self.push_blocking_to_timeline()
        except Exception as e:
            print(f"[BOM] Warning: Could not push blocking data to timeline: {e}")

    def refresh_from_ecn_data(self):
        """Refresh the BOM Explorer tree to pick up newly loaded ECN data.

        Called after ECN data is loaded from the main dropdown so the
        BOM Explorer shows updated workflow status without a manual action.
        """
        if self._bom_database:
            self._rebuild_tree()

    def _on_export_csv(self):
        """Export flat CSV of all loaded BOM data."""
        if not self._bom_database:
            QMessageBox.information(self, "Nothing to Export", "No cases loaded.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export BOM Data", "bom_explorer_export.csv", "CSV Files (*.csv)"
        )
        if not file_path:
            return

        import csv
        ecn_lookup = self._get_ecn_item_lookup()
        rows = []
        for nom, bom_result in self._bom_database.items():
            if bom_result is None:
                continue
            nodes = (bom_result.get("bom", {}) or {}).get("nodes", []) or []
            for node in nodes:
                item_id = node.get("item_id", "")
                ecn_item = ecn_lookup.get(item_id)
                release_text = self._release_text(node.get("release_status_list", ""))
                is_released = "PRODUCTION RELEASED" in release_text.upper()
                workflow = self._parse_workflow(node.get("process_stage", ""))
                ecn_step = getattr(ecn_item, 'current_step_name', '') if ecn_item else ""
                ecn_performer = getattr(ecn_item, 'current_step_performer', '') if ecn_item else ""
                eff_progress = getattr(ecn_item, 'effective_progress', 0) if ecn_item else 0
                rows.append({
                    "Case": nom,
                    "Item ID": item_id,
                    "Revision": node.get("item_revision_id", ""),
                    "Name": node.get("object_name", ""),
                    "BOM Depth": node.get("depth", 0),
                    "Workflow (TC)": workflow,
                    "ECN": node.get("h4_ECN_Number", ""),
                    "Prod Released": "Yes" if is_released else "No",
                    "Release Status": release_text,
                    "Current Task": ecn_step,
                    "Assigned To": ecn_performer,
                    "Progress": f"{eff_progress:.0f}%" if ecn_item else "",
                })

        if rows:
            fields = list(rows[0].keys())
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fields)
                writer.writeheader()
                writer.writerows(rows)
            QMessageBox.information(self, "Exported", f"Saved {len(rows)} rows to:\n{file_path}")

    # ------------------------------------------------------------------
    # TREE BUILDING
    # ------------------------------------------------------------------

    def _rebuild_tree(self):
        """Rebuild the QTreeWidget from the BOM database."""
        self._tree.clear()

        if not self._bom_database:
            self._empty_label.setVisible(True)
            self._tree.setVisible(False)
            self._status_label.setText("No cases loaded.")
            return

        self._empty_label.setVisible(False)
        self._tree.setVisible(True)

        total_parts = 0
        total_released = 0
        total_blocked = 0

        for nom, bom_result in self._bom_database.items():
            if bom_result is None:
                # Placeholder case (saved but not yet fetched) — show as empty root
                root_item = QTreeWidgetItem(self._tree)
                root_item.setText(self.COL_ITEM, nom)
                root_item.setText(self.COL_NAME, "(not yet fetched — click Refresh All)")
                root_item.setForeground(self.COL_NAME, QColor("#888888"))
                continue
            bom = bom_result.get("bom", {}) or {}
            nodes = bom.get("nodes", []) or []
            edges = bom.get("edges", []) or []

            if not nodes:
                continue

            # Build parent→children map
            node_by_uid = {n["uid"]: n for n in nodes if isinstance(n, dict) and n.get("uid")}
            children_of = {}
            child_set = set()
            for e in edges:
                if not isinstance(e, dict):
                    continue
                p = e.get("parent_uid", "")
                c = e.get("child_uid", "")
                if p and c:
                    children_of.setdefault(p, []).append(c)
                    child_set.add(c)

            root_uids = [n["uid"] for n in nodes if n.get("uid") and n["uid"] not in child_set]

            ecn_lookup = self._get_ecn_item_lookup()

            def _add_node(parent_widget, uid):
                nonlocal total_parts, total_released, total_blocked
                node = node_by_uid.get(uid, {})
                item_id = node.get("item_id", "")
                rev = node.get("item_revision_id", "")
                name = node.get("object_name", "")
                release_text = self._release_text(node.get("release_status_list", ""))
                is_released_tc = "PRODUCTION RELEASED" in release_text.upper()
                workflow_tc = self._parse_workflow(node.get("process_stage", ""))
                is_blocked = "on-hold" in workflow_tc.lower() or "onhold" in workflow_tc.lower()
                ecn = node.get("h4_ECN_Number", "")
                qty = node.get("quantity", "")

                # Match against loaded ECN data (single source of truth for workflow status)
                ecn_item = ecn_lookup.get(item_id)
                ecn_step = ""
                ecn_performer = ""
                ecn_progress = ""
                if ecn_item:
                    ecn_step = getattr(ecn_item, 'current_step_name', '') or ""
                    ecn_performer = getattr(ecn_item, 'current_step_performer', '') or ""
                    completed = getattr(ecn_item, 'completed_steps', 0)
                    total_s = getattr(ecn_item, 'total_steps', 0)
                    eff_progress = getattr(ecn_item, 'effective_progress', 0)
                    if total_s > 0:
                        ecn_progress = f"{completed}/{total_s} ({eff_progress:.0f}%)"

                # Determine displayed status — prefer ECN data, then TC data
                if ecn_item and ecn_step:
                    if ecn_step.upper() in ("COMPLETED",):
                        display_status = "✓ Completed"
                        is_released = True
                    elif getattr(ecn_item, 'is_rejected', False):
                        display_status = f"✗ Rejected at {ecn_step}"
                        is_released = False
                    else:
                        display_status = ecn_step  # e.g. "Sheetmetal Review"
                        is_released = False
                else:
                    is_released = is_released_tc
                    display_status = "✓ Prod Released" if is_released_tc else "✗ Not Released"

                # Use TC workflow if ECN data doesn't have a step
                workflow_display = ecn_step if ecn_step else workflow_tc
                if is_blocked and not ecn_step:
                    workflow_display = f"⏸ {workflow_tc}"

                total_parts += 1
                if is_released:
                    total_released += 1
                if is_blocked:
                    total_blocked += 1

                item_label = f"{item_id}/{rev}" if rev else item_id
                tree_item = QTreeWidgetItem()
                tree_item.setText(self.COL_ITEM, item_label)
                tree_item.setText(self.COL_NAME, name)
                tree_item.setText(self.COL_QTY, str(qty) if qty != "" else "")
                tree_item.setText(self.COL_WORKFLOW, workflow_display)
                tree_item.setText(self.COL_ECN, ecn)
                tree_item.setText(self.COL_RELEASED, display_status)
                tree_item.setText(self.COL_ASSIGNEE, ecn_performer)
                tree_item.setText(self.COL_TASK, ecn_progress)

                # Link status column (populated after link check)
                link_info = self._link_check_data.get(item_id)
                if link_info:
                    tree_item.setText(self.COL_LINK, link_info["text"])
                    tree_item.setBackground(self.COL_LINK, QColor(link_info["color"]))
                    if link_info.get("blocked_children"):
                        bc_lines = [f"Blocked children:"]
                        for bc in link_info["blocked_children"][:10]:
                            bc_line = f"  ✗ {bc['item_id']}/{bc['rev']}"
                            if bc['step']:
                                bc_line += f" — {bc['step']}"
                            if bc['performer']:
                                bc_line += f" ({bc['performer']})"
                            bc_lines.append(bc_line)
                        if len(link_info["blocked_children"]) > 10:
                            bc_lines.append(f"  ... and {len(link_info['blocked_children']) - 10} more")
                        tree_item.setToolTip(self.COL_LINK, "\n".join(bc_lines))

                # Color coding
                if getattr(ecn_item, 'is_rejected', False) if ecn_item else False:
                    for col in range(len(self.HEADERS)):
                        tree_item.setBackground(col, QColor("#FADBD8"))  # red tint for rejected
                elif is_blocked:
                    for col in range(len(self.HEADERS)):
                        tree_item.setBackground(col, QColor("#FFF3CD"))  # yellow
                elif not is_released:
                    for col in range(len(self.HEADERS)):
                        tree_item.setBackground(col, QColor("#F8D7DA"))  # light red
                else:
                    for col in range(len(self.HEADERS)):
                        tree_item.setBackground(col, QColor("#D4EDDA"))  # light green

                # Status column color
                if is_released:
                    tree_item.setForeground(self.COL_RELEASED, QColor("#28a745"))
                else:
                    tree_item.setForeground(self.COL_RELEASED, QColor("#dc3545"))

                # Tooltip
                tooltip_lines = [
                    f"{item_label} — {name}",
                    "",
                    f"Release Status (TC): {release_text or '(none)'}",
                ]
                if workflow_tc:
                    tooltip_lines.append(f"Workflow (TC): {workflow_tc}")
                if ecn:
                    tooltip_lines.append(f"ECN: {ecn}")
                tooltip_lines.append(f"BOM Depth: Level {node.get('depth', 0)} of {nom}")
                if ecn_item:
                    tooltip_lines.append("")
                    tooltip_lines.append("── ECN Workflow Status ──")
                    if ecn_step:
                        tooltip_lines.append(f"Current Task: {ecn_step}")
                    if ecn_performer:
                        tooltip_lines.append(f"Assigned To: {ecn_performer}")
                    if ecn_progress:
                        tooltip_lines.append(f"Progress: {ecn_progress}")
                    if getattr(ecn_item, 'is_rejected', False):
                        tooltip_lines.append("⚠️ REJECTED")

                tree_item.setToolTip(self.COL_ITEM, "\n".join(tooltip_lines))
                tree_item.setToolTip(self.COL_NAME, "\n".join(tooltip_lines))

                # Store data for filtering
                tree_item.setData(self.COL_ITEM, Qt.ItemDataRole.UserRole, {
                    "uid": uid, "item_id": item_id, "is_released": is_released,
                    "is_blocked": is_blocked, "ecn": ecn, "nom": nom,
                })

                if isinstance(parent_widget, QTreeWidget):
                    parent_widget.addTopLevelItem(tree_item)
                else:
                    parent_widget.addChild(tree_item)

                # Recurse children
                for child_uid in children_of.get(uid, []):
                    _add_node(tree_item, child_uid)

                return tree_item

            for root_uid in root_uids:
                _add_node(self._tree, root_uid)

        # Update status bar
        pct = (total_released / total_parts * 100) if total_parts else 0
        case_count = len(self._bom_database)
        self._status_label.setText(
            f"{case_count} case{'s' if case_count != 1 else ''} │ "
            f"{total_parts} parts │ "
            f"{total_released} released ({pct:.1f}%) │ "
            f"{total_blocked} blocked"
        )

        self._apply_filter()

    # ------------------------------------------------------------------
    # FILTERING & SEARCH
    # ------------------------------------------------------------------

    def _apply_filter(self):
        """Apply filter and search to tree items."""
        filter_mode = self._filter_combo.currentText()
        search_text = self._search_edit.text().strip().lower()

        def _set_visible_recursive(item, force_visible=False):
            """Returns True if this item or any descendant should be visible."""
            data = item.data(self.COL_ITEM, Qt.ItemDataRole.UserRole) or {}
            item_text = f"{item.text(self.COL_ITEM)} {item.text(self.COL_NAME)} {item.text(self.COL_ECN)}".lower()

            # Check search match
            search_match = not search_text or search_text in item_text

            # Check filter match
            if filter_mode == "Blocked Only":
                filter_match = data.get("is_blocked", False) or not data.get("is_released", True)
            elif filter_mode == "Unreleased Only":
                filter_match = not data.get("is_released", True)
            elif filter_mode == "In ECN":
                filter_match = bool(data.get("ecn", ""))
            else:
                filter_match = True

            # Check children
            any_child_visible = False
            for i in range(item.childCount()):
                if _set_visible_recursive(item.child(i)):
                    any_child_visible = True

            visible = (filter_match and search_match) or any_child_visible or force_visible
            item.setHidden(not visible)
            return visible

        for i in range(self._tree.topLevelItemCount()):
            _set_visible_recursive(self._tree.topLevelItem(i), force_visible=(filter_mode == "All" and not search_text))

    # ------------------------------------------------------------------
    # CONTEXT MENU
    # ------------------------------------------------------------------

    def _on_context_menu(self, pos):
        item = self._tree.itemAt(pos)
        if not item:
            return

        menu = QMenu(self)
        copy_id = menu.addAction("Copy Item ID")
        copy_id_rev = menu.addAction("Copy Item ID / Rev")
        copy_row = menu.addAction("Copy Row (all columns)")
        menu.addSeparator()
        expand_all = menu.addAction("Expand All Below")
        collapse_all = menu.addAction("Collapse All Below")

        # Remove case only for top-level items
        remove_action = None
        if not item.parent():
            menu.addSeparator()
            remove_action = menu.addAction("Remove Case")

        action = menu.exec(self._tree.viewport().mapToGlobal(pos))
        if not action:
            return

        if action == copy_id:
            item_id = item.text(self.COL_ITEM).split("/")[0]
            QApplication.clipboard().setText(item_id)
        elif action == copy_id_rev:
            QApplication.clipboard().setText(item.text(self.COL_ITEM))
        elif action == copy_row:
            row_text = "\t".join(item.text(c) for c in range(len(self.HEADERS)))
            QApplication.clipboard().setText(row_text)
        elif action == expand_all:
            self._expand_recursive(item, True)
        elif action == collapse_all:
            self._expand_recursive(item, False)
        elif action == remove_action:
            data = item.data(self.COL_ITEM, Qt.ItemDataRole.UserRole) or {}
            nom = data.get("nom", "")
            if nom and nom in self._bom_database:
                del self._bom_database[nom]
                self._save_cases_to_config()
                self._rebuild_tree()

    def _expand_recursive(self, item, expand):
        item.setExpanded(expand)
        for i in range(item.childCount()):
            self._expand_recursive(item.child(i), expand)

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------

    @staticmethod
    def _release_text(value):
        if isinstance(value, list):
            return ", ".join(str(v) for v in value if v)
        return str(value or "")

    @staticmethod
    def _parse_workflow(process_stage):
        """Extract workflow task name from process_stage text."""
        text = str(process_stage or "").strip()
        if not text:
            return ""
        # process_stage looks like: "On-Hold Auto Complete/3247229/A;1-MODULE:2"
        # Extract the task name (first part before /)
        parts = text.split(",")
        names = set()
        for part in parts:
            part = part.strip()
            if "/" in part:
                task_name = part.split("/")[0].strip()
                if task_name:
                    names.add(task_name)
        return ", ".join(sorted(names)) if names else text[:50]

    def compute_blocking_data(self):
        """Analyze BOM database and compute blocking info for each item.

        Returns dict: item_id -> {
            "item_id", "rev", "name", "nom", "is_released", "release_text",
            "ecn", "workflow", "assignee", "task",
            "blocking_children": [{item_id, rev, name, ecn, assignee, task, release_text}, ...],
            "blocking_chain": ["RLN4MA -> 3247229 -> 3247230 (not PROD RELEASED)"],
            "total_children", "released_children"
        }
        """
        blocking = {}  # item_id -> info dict

        for nom, bom_result in self._bom_database.items():
            if bom_result is None:
                continue
            bom = bom_result.get("bom", {}) or {}
            nodes = bom.get("nodes", []) or []
            edges = bom.get("edges", []) or []
            if not nodes:
                continue

            node_by_uid = {n["uid"]: n for n in nodes if isinstance(n, dict) and n.get("uid")}
            children_of = {}
            parent_of = {}
            child_set = set()
            for e in edges:
                if not isinstance(e, dict):
                    continue
                p = e.get("parent_uid", "")
                c = e.get("child_uid", "")
                if p and c:
                    children_of.setdefault(p, []).append(c)
                    parent_of[c] = p
                    child_set.add(c)

            def _get_all_descendants(uid):
                """Get all descendant UIDs recursively."""
                desc = []
                for child_uid in children_of.get(uid, []):
                    desc.append(child_uid)
                    desc.extend(_get_all_descendants(child_uid))
                return desc

            def _get_ancestor_chain(uid):
                """Get list of ancestor UIDs from immediate parent to root."""
                chain = []
                current = uid
                while current in parent_of:
                    current = parent_of[current]
                    chain.append(current)
                return chain

            def _is_prod_released(node):
                release_text = self._release_text(node.get("release_status_list", ""))
                return "PRODUCTION RELEASED" in release_text.upper()

            ecn_lookup = self._get_ecn_item_lookup()

            # First pass: find all unreleased leaf/child parts
            for node in nodes:
                uid = node.get("uid", "")
                if not uid:
                    continue
                item_id = node.get("item_id", "")
                if not item_id:
                    continue

                is_released = _is_prod_released(node)
                rev = node.get("item_revision_id", "")
                name = node.get("object_name", "")
                ecn = node.get("h4_ECN_Number", "")
                workflow = self._parse_workflow(node.get("process_stage", ""))
                ecn_item = ecn_lookup.get(item_id)
                release_text = self._release_text(node.get("release_status_list", ""))

                # Count children stats
                desc_uids = _get_all_descendants(uid)
                total_children = len(desc_uids)
                released_children = sum(1 for d in desc_uids if _is_prod_released(node_by_uid.get(d, {})))

                # Find unreleased direct+indirect children
                blocking_children = []
                for d_uid in desc_uids:
                    d_node = node_by_uid.get(d_uid, {})
                    if not _is_prod_released(d_node):
                        d_item_id = d_node.get("item_id", "")
                        if d_item_id:
                            d_ecn_item = ecn_lookup.get(d_item_id)
                            blocking_children.append({
                                "item_id": d_item_id,
                                "rev": d_node.get("item_revision_id", ""),
                                "name": d_node.get("object_name", ""),
                                "ecn": d_node.get("h4_ECN_Number", ""),
                                "release_text": self._release_text(d_node.get("release_status_list", "")),
                                "assignee": getattr(d_ecn_item, 'current_step_performer', '') if d_ecn_item else "",
                                "task": getattr(d_ecn_item, 'current_step_name', '') if d_ecn_item else "",
                            })

                # Build blocking chains (paths from this node to unreleased leaves)
                blocking_chains = []
                for bc in blocking_children:
                    # Find the path uid for this blocking child
                    for d_uid in desc_uids:
                        d_node = node_by_uid.get(d_uid, {})
                        if d_node.get("item_id", "") == bc["item_id"]:
                            # Build path from current node to this child
                            path_parts = []
                            walk = d_uid
                            while walk and walk != uid:
                                w_node = node_by_uid.get(walk, {})
                                path_parts.insert(0, w_node.get("item_id", walk))
                                walk = parent_of.get(walk)
                            path_parts.insert(0, item_id)
                            chain_str = " → ".join(path_parts) + " (not PROD RELEASED)"
                            blocking_chains.append(chain_str)
                            break

                has_blocking = not is_released or len(blocking_children) > 0

                if has_blocking:
                    info = {
                        "item_id": item_id,
                        "rev": rev,
                        "name": name,
                        "nom": nom,
                        "is_released": is_released,
                        "release_text": release_text,
                        "ecn": ecn,
                        "workflow": workflow,
                        "assignee": getattr(ecn_item, 'current_step_performer', '') if ecn_item else "",
                        "task": getattr(ecn_item, 'current_step_name', '') if ecn_item else "",
                        "blocking_children": blocking_children,
                        "blocking_chain": blocking_chains,
                        "total_children": total_children,
                        "released_children": released_children,
                    }
                    # Store by item_id (if same item_id in multiple cases, last wins)
                    blocking[item_id] = info

        return blocking

    def push_blocking_to_timeline(self):
        """Compute blocking data and push it to the timeline canvas."""
        blocking = self.compute_blocking_data()
        canvas = self._main_window.timeline_canvas
        canvas.set_bom_blocking_data(blocking)
        canvas.update()

    def _save_cases_to_config(self):
        """Save case list to VPM config and persist full BOM data to disk."""
        config = self._main_window.ecn_config
        config["bom_explorer_cases"] = list(self._bom_database.keys())
        self._main_window.save_ecn_config()
        self._write_bom_cache_files()

    # ------------------------------------------------------------------
    # BOM disk cache helpers
    # ------------------------------------------------------------------

    def _cache_path(self, nom):
        """Return the path to the raw bom_result cache file for a nomenclature."""
        import re
        token = re.sub(r"[^a-zA-Z0-9]+", "_", nom).strip("_") or "bom"
        return os.path.join("tc_bom_output", f"{token}_bom_cache.json")

    def _write_bom_cache_files(self):
        """Write full bom_result dicts to disk so they survive app restarts."""
        try:
            os.makedirs("tc_bom_output", exist_ok=True)
            for nom, result in self._bom_database.items():
                if result is not None:
                    path = self._cache_path(nom)
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
        except Exception:
            pass  # non-fatal; data still in memory

    def _load_bom_cache(self, nom):
        """Load a previously cached bom_result dict from disk. Returns None on miss/error."""
        try:
            path = self._cache_path(nom)
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return None


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ecn_engine = ecn_engine.EcnDashboardEngine(self)
        self.plugins = []
        self.load_plugins()
        self.setWindowTitle("Visual Project Manager"); self.setGeometry(100, 100, 1200, 600)
        self.projects = []
        self.ecn_projects = []
        self.drill_down_stack = []
        self.task_map = {}
        self.current_file_path, self.current_timescale = None, TimescaleView.YEAR
        
        self.ecn_config_path = "vpm_config.json"
        self.ecn_config = {}
        
        self.create_toolbar()
        central_widget = QWidget(); self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget); main_layout.setContentsMargins(0,0,0,0); main_layout.setSpacing(0)
        self.breadcrumb_frame = QFrame(); self.breadcrumb_layout = QHBoxLayout(self.breadcrumb_frame)
        self.breadcrumb_layout.setContentsMargins(5,5,5,5); self.breadcrumb_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        main_layout.addWidget(self.breadcrumb_frame)
        self.timeline_canvas = TimelineCanvas()
        self.timeline_canvas.comment_requested.connect(self.show_comment_dialog)
        self.timeline_canvas.task_double_clicked.connect(self.perform_drill_down)
        self.timeline_canvas.progress_changed.connect(self.update_view)
        self.timeline_canvas.edit_task_requested.connect(self.edit_item)
        self.timeline_canvas.phase_complete_requested.connect(self.mark_phase_complete)
        self.timeline_canvas.active_zone_changed.connect(self.set_active_zone)
        self.timeline_canvas.delete_task_requested.connect(self.delete_item)
        self.timeline_canvas.summary_requested.connect(self.show_ecn_summary)
        self.timeline_canvas.edit_project_requested.connect(self.edit_project)
        self.timeline_canvas.add_to_notes_requested.connect(self.add_task_to_notes)
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setWidget(self.timeline_canvas)
        self.scroll_area.setStyleSheet("""
            QScrollBar:vertical {
                border: 1px solid #d0d0d0;
                background: #f0f0f0;
                width: 15px;
                margin: 0;
            }
            QScrollBar::handle:vertical {
                background: #b0b0b0;
                min-height: 20px;
                border-radius: 7px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
                border: none;
                background: none;
            }
            QScrollBar:horizontal {
                border: 1px solid #d0d0d0;
                background: #f0f0f0;
                height: 15px;
                margin: 0;
            }
            QScrollBar::handle:horizontal {
                background: #b0b0b0;
                min-width: 20px;
                border-radius: 7px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #a0a0a0;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
                border: none;
                background: none;
            }
        """)

        # Create Notes View
        self.notes_view = NotesView(self)

        # Create status checking timer (runs every 60 seconds)
        self.status_check_timer = QTimer(self)
        self.status_check_timer.timeout.connect(self.check_note_statuses)
        self.status_check_timer.start(60000)  # 60 seconds

        # Create stacked widget to switch between Timeline and Notes views
        self.view_stack = QStackedWidget()
        self.view_stack.addWidget(self.scroll_area)  # Index 0: Timeline/Gantt
        self.view_stack.addWidget(self.notes_view)    # Index 1: Notes
        self.bom_explorer = BomExplorerView(self)
        self.view_stack.addWidget(self.bom_explorer)  # Index 2: BOM Explorer
        self.view_stack.setCurrentIndex(0)  # Start with Timeline view

        main_layout.addWidget(self.view_stack, 1)

        self.load_ecn_config()

        # Restore saved BOM Explorer cases from config (load full data from disk cache)
        saved_cases = self.ecn_config.get("bom_explorer_cases", [])
        if saved_cases:
            for nom in saved_cases:
                cached = self.bom_explorer._load_bom_cache(nom)
                self.bom_explorer._bom_database[nom.upper()] = cached  # None if no cache yet
            self.bom_explorer._rebuild_tree()

    def load_plugins(self):
        plugin_dir = "plugins"
        if not os.path.exists(plugin_dir):
            os.makedirs(plugin_dir)
        for filename in os.listdir(plugin_dir):
            if filename.endswith('.py') and not filename.startswith('__'):
                path = os.path.join(plugin_dir, filename)
                spec = importlib.util.spec_from_file_location(filename, path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                self.plugins.append(module)

    def create_toolbar(self):
        toolbar = QToolBar("Main Toolbar"); self.addToolBar(toolbar)
        open_action = QAction("&Open...", self); open_action.triggered.connect(self.open_project)
        save_action = QAction("&Save", self); save_action.triggered.connect(self.save_project)
        save_as_action = QAction("Save &As...", self); save_as_action.triggered.connect(self.save_project_as)
        toolbar.addActions([open_action, save_action, save_as_action]); toolbar.addSeparator()

        # Import menu (always show as menu to include Excel template option)
        import_menu_button = QPushButton("Import")
        import_menu = QMenu(self)

        # Excel Template Import
        excel_import_action = QAction("Import from Excel Template...", self)
        excel_import_action.triggered.connect(self.import_from_excel_template)
        import_menu.addAction(excel_import_action)

        # MS Project XML Import
        xml_import_action = QAction("Import from MS Project XML...", self)
        xml_import_action.triggered.connect(self.import_ms_project)
        import_menu.addAction(xml_import_action)

        # Plugin imports
        if self.plugins:
            import_menu.addSeparator()
            for plugin in self.plugins:
                plugin_action = QAction(plugin.get_importer_name(), self)
                plugin_action.triggered.connect(partial(self.run_plugin_importer, plugin))
                import_menu.addAction(plugin_action)

        import_menu_button.setMenu(import_menu)
        toolbar.addWidget(import_menu_button)

        # Export menu
        export_menu_button = QPushButton("Export")
        export_menu = QMenu(self)

        excel_export_action = QAction("Export to Excel Template...", self)
        excel_export_action.triggered.connect(self.export_to_excel_template)
        export_menu.addAction(excel_export_action)

        export_menu_button.setMenu(export_menu)
        toolbar.addWidget(export_menu_button)

        add_items_button = QPushButton("Add Items")
        add_items_button.clicked.connect(self.show_add_item_dialog)
        toolbar.addWidget(add_items_button)
        toolbar.addSeparator()
        
        load_ecn_button = QPushButton("ECN Data \u25bc")
        ecn_menu = QMenu(self)
        excel_action = QAction("Load from Excel...", self)
        excel_action.triggered.connect(self.load_ecn_data)
        ecn_menu.addAction(excel_action)
        tc_action = QAction("Connect to Teamcenter...", self)
        tc_action.triggered.connect(self.connect_to_teamcenter)
        ecn_menu.addAction(tc_action)
        load_ecn_button.setMenu(ecn_menu)
        toolbar.addWidget(load_ecn_button)
        
        toolbar.addSeparator()
        
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Search...")
        self.search_bar.setMaximumWidth(200)
        self.search_bar.returnPressed.connect(self.perform_search)
        toolbar.addWidget(self.search_bar)
        search_button = QPushButton("Search")
        search_button.clicked.connect(self.perform_search)
        toolbar.addWidget(search_button)

        toolbar.addSeparator()

        # View Toggle Buttons
        self.timeline_button = QPushButton("📊 Timeline")
        self.timeline_button.setCheckable(True)
        self.timeline_button.setChecked(True)
        self.timeline_button.clicked.connect(lambda: self.switch_view(0))
        self.timeline_button.setStyleSheet("""
            QPushButton {
                padding: 5px 15px;
                font-size: 13px;
                border: 1px solid #0078d4;
            }
            QPushButton:checked {
                background-color: #0078d4;
                color: white;
            }
        """)
        toolbar.addWidget(self.timeline_button)

        self.notes_button = QPushButton("📝 Notes")
        self.notes_button.setCheckable(True)
        self.notes_button.clicked.connect(lambda: self.switch_view(1))
        self.notes_button.setStyleSheet("""
            QPushButton {
                padding: 5px 15px;
                font-size: 13px;
                border: 1px solid #0078d4;
            }
            QPushButton:checked {
                background-color: #0078d4;
                color: white;
            }
        """)
        toolbar.addWidget(self.notes_button)

        self.bom_explorer_button = QPushButton("🔧 BOM Explorer")
        self.bom_explorer_button.setCheckable(True)
        self.bom_explorer_button.clicked.connect(lambda: self.switch_view(2))
        self.bom_explorer_button.setStyleSheet("""
            QPushButton {
                padding: 5px 15px;
                font-size: 13px;
                border: 1px solid #0078d4;
            }
            QPushButton:checked {
                background-color: #0078d4;
                color: white;
            }
        """)
        toolbar.addWidget(self.bom_explorer_button)

        spacer = QWidget(); spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding); toolbar.addWidget(spacer)
        
        legend_button = QPushButton("Legend")
        legend_button.clicked.connect(self.show_legend_dialog)
        toolbar.addWidget(legend_button)
        
        self.notification_button = QPushButton("🔔", self); self.notification_button.setFlat(True); self.notification_button.clicked.connect(self.show_notifications); self.notification_button.setEnabled(False); toolbar.addWidget(self.notification_button)
        self.timescale_action = QAction("Zoom: Year", self); self.timescale_action.triggered.connect(self.toggle_timescale); toolbar.addAction(self.timescale_action)
        summary_action = QAction("Summary", self); summary_action.triggered.connect(self.show_summary); toolbar.addAction(summary_action)

    def run_plugin_importer(self, plugin):
        try:
            projects_data = plugin.run_importer()
            if not projects_data:
                return

            self.projects = [p for p in self.projects if p.source_plugin != plugin.__name__]
            self.task_map = {uid: task for uid, task in self.task_map.items() if task.project.source_plugin != plugin.__name__}

            # Handle mixed UID types (int for regular tasks, str for ECN items)
            if self.task_map:
                int_uids = [uid for uid in self.task_map.keys() if isinstance(uid, int)]
                next_uid = (max(int_uids) if int_uids else 0) + 1
            else:
                next_uid = 1

            for proj_data in projects_data:
                new_project = Project(proj_data['name'], next_uid)
                next_uid += 1
                new_project.source_plugin = plugin.__name__

                tasks_data = proj_data.get('tasks', [])
                self._recursively_create_tasks(tasks_data, new_project, None, next_uid)
                
                self.projects.append(new_project)
            
            self.update_project_dates_and_view()

        except Exception as e:
            QMessageBox.critical(self, f"Plugin Error: {plugin.get_importer_name()}", f"An error occurred while running the plugin:\n\n{e}")
    
    def load_ecn_config(self):
        if os.path.exists(self.ecn_config_path):
            try:
                with open(self.ecn_config_path, 'r') as f:
                    self.ecn_config = json.load(f)
            except Exception as e:
                QMessageBox.warning(self, "Config Error", f"Could not load ECN config file:\n{e}")
                self.ecn_config = {}
        else:
            self.ecn_config = {}

    def save_ecn_config(self):
        try:
            with open(self.ecn_config_path, 'w') as f:
                json.dump(self.ecn_config, f, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Config Save Error", f"Could not save ECN config file:\n{e}")

    def save_notes_structure(self):
        """Save notes structure to config file."""
        try:
            self.ecn_config['notes_structure'] = self.notes_view.notes_structure.to_dict()
            self.save_ecn_config()
        except Exception as e:
            print(f"Warning: Could not save notes structure: {e}")

    def load_notes_structure(self):
        """Load notes structure from config file."""
        print(f"[DEBUG] load_notes_structure() CALLED - REPLACING notes_structure from disk!")
        import traceback
        traceback.print_stack()
        try:
            if 'notes_structure' in self.ecn_config:
                self.notes_view.notes_structure = NotesStructure.from_dict(
                    self.ecn_config['notes_structure'],
                    self.task_map
                )
                print(f"[DEBUG] Loaded notes_structure from config - ALL OBJECTS REPLACED!")
        except Exception as e:
            print(f"Warning: Could not load notes structure: {e}")
            self.notes_view.notes_structure = NotesStructure()

    def update_pr_parent_map(self, pr_number, parent_uid):
        if 'pr_parent_map' not in self.ecn_config:
            self.ecn_config['pr_parent_map'] = {}
        self.ecn_config['pr_parent_map'][pr_number] = parent_uid
        self.save_ecn_config()
    
    def update_pr_placement_map(self, pr_placement_map):
        """Update the PR placement configuration"""
        if 'pr_placement_map' not in self.ecn_config:
            self.ecn_config['pr_placement_map'] = {}
        self.ecn_config['pr_placement_map'].update(pr_placement_map)
        self.save_ecn_config()
    
    def get_pr_placement_map(self):
        """Get the current PR placement configuration"""
        return self.ecn_config.get('pr_placement_map', {})

    def update_ecn_metadata(self, item):
        if 'ecn_metadata' not in self.ecn_config:
            self.ecn_config['ecn_metadata'] = {}
        
        self.ecn_config['ecn_metadata'][item.uid] = {
            'comments': item.comments
        }
        self.save_ecn_config()

    def load_ecn_data(self):
        # Always let the user pick Excel file(s); remember the last used folder
        start_dir = self.ecn_config.get("ecn_data_path", "")
        files, _ = QFileDialog.getOpenFileNames(self, "Select ECN Excel File(s)", start_dir, "Excel Files (*.xlsx *.xls)")
        if not files:
            return

        # Persist last selected folder for convenience
        last_dir = os.path.dirname(files[0])
        if last_dir:
            self.ecn_config["ecn_data_path"] = last_dir
            self.save_ecn_config()

        # First, parse the Excel files to extract PR information
        pr_data = self._extract_pr_data_from_files(files)
        if not pr_data:
            QMessageBox.warning(self, "No PR Data", "No PR data found in the selected files.")
            return

        # Check for known vs new PRs
        existing_placement_map = self.get_pr_placement_map()
        known_prs = []
        new_prs = []
        
        for pr_info in pr_data:
            pr_id = pr_info.get('pr_id')
            if pr_id in existing_placement_map:
                known_prs.append(pr_info)
            else:
                new_prs.append(pr_info)

        # Show PR placement dialog if there are new PRs or if user wants to reconfigure
        if new_prs:
            if known_prs:
                QMessageBox.information(self, "Mixed PRs", 
                    f"Found {len(new_prs)} new PRs and {len(known_prs)} known PRs.\n\n"
                    "You can configure placement for all PRs or just the new ones.")
            self._show_pr_placement_dialog(new_prs, known_prs)
        elif known_prs:
            # All PRs are known, ask if user wants to reconfigure
            reply = QMessageBox.question(
                self, "All PRs Known", 
                f"All {len(known_prs)} PRs have existing placement configurations.\n\n"
                "Would you like to reconfigure them?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._show_pr_placement_dialog(pr_data, [])
            else:
                # User chose not to reconfigure, proceed with existing configuration
                QMessageBox.information(self, "Using Existing Configuration", 
                    f"Using existing placement configuration for {len(known_prs)} PRs.")

        # Proceed with ECN import using the placement configuration
        current_parent = self.drill_down_stack[-1] if self.drill_down_stack else None
        pr_parent_map = self.ecn_config.get("pr_parent_map", {})
        ecn_metadata = self.ecn_config.get("ecn_metadata", {})

        all_known_parents = self.task_map.copy()
        for p in self.projects + self.ecn_projects:
            all_known_parents[p.uid] = p

        # Get PR placement configuration
        pr_placement_map = self.get_pr_placement_map()
        
        self.ecn_engine.load_and_display_data(
            pr_parent_map,
            current_parent,
            all_known_parents,
            ecn_metadata,
            file_paths=files,
            pr_placement_map=pr_placement_map
        )

        # Refresh BOM Explorer with newly loaded ECN data
        if hasattr(self, 'bom_explorer') and self.bom_explorer:
            try:
                self.bom_explorer.refresh_from_ecn_data()
            except Exception:
                pass

    def connect_to_teamcenter(self):
        """Connect to Teamcenter and fetch ECN data via SOA.
        Persists the TC session so credentials are only needed once."""
        from tc_connector import TcLoginDialog, TcSearchDialog

        # Check if we already have a connected session
        client = getattr(self, '_tc_client', None)
        if client and getattr(client, 'is_connected', False):
            # Reuse existing session — skip login dialog
            pass
        else:
            # Step 1: Show login dialog
            tc_config = self.ecn_config.get("teamcenter", {})
            login_dialog = TcLoginDialog(tc_config, self)
            if login_dialog.exec() != QDialog.DialogCode.Accepted:
                return

            client = login_dialog.get_client()
            self._tc_client = client  # Persist for reuse

            # Save URL/username to config if opted in
            self.ecn_config["teamcenter"] = login_dialog.get_config()
            self.save_ecn_config()

        # Step 2: Show search dialog
        search_dialog = TcSearchDialog(client, self)
        if search_dialog.exec() != QDialog.DialogCode.Accepted:
            # Don't logout — keep session alive for next search
            return

        # Step 3: Process fetched data
        row_dicts = search_dialog.get_results()
        if not row_dicts:
            return

        # Feed into existing pipeline — same as Excel import path
        current_parent = self.drill_down_stack[-1] if self.drill_down_stack else None
        pr_parent_map = self.ecn_config.get("pr_parent_map", {})
        ecn_metadata = self.ecn_config.get("ecn_metadata", {})

        all_known_parents = self.task_map.copy()
        for p in self.projects + self.ecn_projects:
            all_known_parents[p.uid] = p

        pr_placement_map = self.get_pr_placement_map()

        self.ecn_engine.load_from_tc_data(
            row_dicts,
            pr_parent_map,
            current_parent,
            all_known_parents,
            ecn_metadata,
            pr_placement_map=pr_placement_map
        )

        # Refresh BOM Explorer with newly loaded ECN data
        if hasattr(self, 'bom_explorer') and self.bom_explorer:
            try:
                self.bom_explorer.refresh_from_ecn_data()
            except Exception:
                pass

        # Don't logout — keep session alive
        QMessageBox.information(
            self, "TC Data Loaded",
            f"Successfully loaded {len(row_dicts)} items from Teamcenter."
        )

    def disconnect_teamcenter(self):
        """Explicitly disconnect from Teamcenter."""
        client = getattr(self, '_tc_client', None)
        if client and getattr(client, 'is_connected', False):
            try:
                client.logout()
            except Exception:
                pass
        self._tc_client = None

    def closeEvent(self, event):
        """Clean up TC session on app close."""
        self.disconnect_teamcenter()
        super().closeEvent(event)

    def _extract_pr_data_from_files(self, file_paths):
        """Extract PR data from Excel files for placement configuration"""
        import pandas as pd

        pr_data = []
        pr_ids_seen = set()
        
        try:
            for file_path in file_paths:
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # Find PR column (still search for this)
                pr_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if any(term in col_lower for term in ['pr id', 'problem report', 'pr number']):
                        pr_col = col
                        break
                
                if not pr_col:
                    continue
                
                # Always use column C (index 2) for descriptions
                if len(df.columns) < 3:
                    continue
                    
                description_col = df.columns[2]  # Always column C
                
                # Extract unique PRs
                for _, row in df.iterrows():
                    pr_id = str(row[pr_col]).strip() if pd.notna(row[pr_col]) else None
                    project_description = str(row[description_col]).strip() if pd.notna(row[description_col]) else ""
                    
                    if pr_id and pr_id not in pr_ids_seen:
                        pr_ids_seen.add(pr_id)
                        pr_data.append({
                            'pr_id': pr_id,
                            'project_description': project_description
                        })
        
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error extracting PR data: {e}")
            return []
        
        return pr_data
    
    def _show_pr_placement_dialog(self, new_prs, known_prs):
        """Show the PR placement configuration dialog"""
        if not self.projects:
            QMessageBox.warning(self, "No Projects", "Please import or create projects before configuring PR placement.")
            return
        
        # Combine new and known PRs for the dialog
        all_prs = new_prs + known_prs
        
        dialog = PRPlacementDialog(all_prs, self.projects, self)
        
        # Pre-populate known PRs if any
        if known_prs:
            existing_placement_map = self.get_pr_placement_map()
            for row in range(dialog.table.rowCount()):
                pr_id = dialog.table.item(row, 0).text()
                if pr_id in existing_placement_map:
                    placement = existing_placement_map[pr_id]
                    project_uid = placement.get('project_uid')
                    level_uid = placement.get('level_uid')
                    
                    # Set project selection
                    project_combo = dialog.table.cellWidget(row, 2)
                    for i in range(project_combo.count()):
                        if project_combo.itemData(i) == project_uid:
                            project_combo.setCurrentIndex(i)
                            break
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            placement_config = dialog.get_placement_config()
            if placement_config:
                self.update_pr_placement_map(placement_config)
                QMessageBox.information(self, "Configuration Saved", 
                    f"PR placement configuration saved for {len(placement_config)} PRs.")

    def _recursively_create_tasks(self, tasks_data, project, parent_task, next_uid):
        for task_data in tasks_data:
            start_dt = datetime.datetime.combine(task_data["start"], datetime.time.min)
            end_dt = datetime.datetime.combine(task_data["end"], datetime.time.max)
            
            new_task = Task(next_uid, task_data['name'], start_dt, end_dt, 1, task_data.get('priority', Priority.MEDIUM))
            next_uid += 1
            
            new_task.project = project
            new_task.parent = parent_task
            new_task.progress = task_data.get('progress', 0)

            if parent_task:
                parent_task.children.append(new_task)
            else:
                project.root_tasks.append(new_task)
            
            self.task_map[new_task.uid] = new_task
            
            if 'children' in task_data:
                next_uid = self._recursively_create_tasks(task_data['children'], project, new_task, next_uid)
        return next_uid

    def import_ms_project(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Import from MS Project XML", "", "XML Files (*.xml)")
        if not file_name: return
        try:
            new_tasks_flat = self.parse_ms_project_xml(file_name)
            if not new_tasks_flat: return
            
            # Handle mixed UID types (int for regular tasks, str for ECN items)
            if self.task_map:
                int_uids = [uid for uid in self.task_map.keys() if isinstance(uid, int)]
                max_uid = max(int_uids) if int_uids else 0
            else:
                max_uid = 0
            for i, task in enumerate(new_tasks_flat):
                task.uid += max_uid
            
            project_name = file_name.split('/')[-1].replace('.xml', '')
            new_project = Project(project_name, max_uid + len(new_tasks_flat) + 1)
            
            for task in new_tasks_flat:
                task.project = new_project
                if task.parent:
                    task.parent.children.append(task)
                else:
                    new_project.root_tasks.append(task)
            
            new_project.update_dates()
            self.projects.append(new_project)
            self.task_map.update({t.uid: t for t in new_tasks_flat})
            
            self.update_project_dates_and_view()
            self.setWindowTitle(f"Visual Project Manager - {self.current_file_path or 'Unsaved Workspace*'}")
        except Exception as e: QMessageBox.critical(self, "Import Error", f"Failed to import XML file.\nError: {e}")

    def open_project(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Workspace", "", "VPMW Files (*.vpmw)")
        if not file_name: return
        try:
            
            self.projects = []
            self.task_map = {}
            self.drill_down_stack = []
            self.ecn_projects = []

            with open(file_name, 'r') as f: data = json.load(f)

            project_map = {}

            for proj_data in data['projects']:
                project = Project(proj_data['name'], proj_data['uid'])
                project.source_plugin = proj_data.get('source_plugin')
                project_map[project.uid] = project
                tasks_flat = [Task.from_dict(t_data) for t_data in proj_data['tasks']]
                temp_task_map = {t.uid: t for t in tasks_flat}

                for task in tasks_flat:
                    task.project = project
                    if hasattr(task, 'parent_uid') and task.parent_uid in temp_task_map:
                        parent = temp_task_map[task.parent_uid]; task.parent = parent; parent.children.append(task)
                    else:
                        project.root_tasks.append(task)
                
                project.update_dates()
                self.projects.append(project)
                self.task_map.update(temp_task_map)
            
            self.current_file_path = file_name
            
            drill_down_uids = data.get('drill_down_stack_uids', [])
            if drill_down_uids:
                reconstructed_stack = []
                for uid in drill_down_uids:
                    item = self.task_map.get(uid) or project_map.get(uid)
                    if item:
                        reconstructed_stack.append(item)
                self.drill_down_stack = reconstructed_stack

            self.update_project_dates_and_view()
            self.setWindowTitle(f"Visual Project Manager - {self.current_file_path}")
        except Exception as e: QMessageBox.critical(self, "Open Error", f"Failed to open workspace file.\nError: {e}")

    def import_from_excel_template(self):
        """SIMPLIFIED: Import tasks from Excel template - AUTOMATIC, NO DIALOGS"""
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Import from Excel Template",
            "",
            "Excel Files (*.xlsm *.xlsx)"
        )

        if not file_name:
            return

        try:
            from openpyxl import load_workbook

            # AUTOMATIC project name from filename
            project_name = os.path.splitext(os.path.basename(file_name))[0]

            # Load workbook
            wb = load_workbook(file_name, data_only=True)

            # DETECT FORMAT: New Tracker vs. Legacy Template
            if 'Tasks' in wb.sheetnames:
                tasks_sheet = wb['Tasks']
                # Check column headers to detect format
                first_col_header = tasks_sheet.cell(1, 1).value

                if first_col_header == "Task Name":
                    # NEW TRACKER FORMAT
                    print("[DEBUG] Detected NEW TRACKER format")
                    return self._import_new_tracker_format(wb, project_name)
                else:
                    # LEGACY TEMPLATE FORMAT
                    print("[DEBUG] Detected LEGACY TEMPLATE format")
                    pass  # Continue with legacy import below

            # Check for required sheets (LEGACY)
            if 'Settings' not in wb.sheetnames or 'Tasks' not in wb.sheetnames:
                QMessageBox.critical(
                    self,
                    "Invalid Template",
                    "Excel file must contain either:\n" +
                    "- 'Task Name' column (New Tracker format)\n" +
                    "- 'Settings' and 'Tasks' sheets (Legacy format)"
                )
                return

            # Read Settings sheet
            settings_sheet = wb['Settings']
            status_mapping = {}

            for row in range(5, 21):
                status_name = settings_sheet.cell(row, 1).value
                if status_name:
                    progress_value = settings_sheet.cell(row, 4).value
                    status_mapping[status_name] = progress_value if progress_value else 0

            # Read Tasks sheet
            tasks_sheet = wb['Tasks']
            imported_tasks = []

            for row_idx in range(6, tasks_sheet.max_row + 1):
                task_name = tasks_sheet.cell(row_idx, 1).value
                if not task_name:
                    continue

                start_date = tasks_sheet.cell(row_idx, 2).value
                end_date = tasks_sheet.cell(row_idx, 3).value
                status = tasks_sheet.cell(row_idx, 4).value
                responsible = tasks_sheet.cell(row_idx, 5).value
                comments = tasks_sheet.cell(row_idx, 6).value
                task_id = tasks_sheet.cell(row_idx, 7).value
                level = tasks_sheet.cell(row_idx, 8).value
                parent_id = tasks_sheet.cell(row_idx, 9).value
                progress = tasks_sheet.cell(row_idx, 10).value

                # Parse dates
                if isinstance(start_date, str):
                    start_date = self._parse_excel_date(start_date)
                elif start_date:
                    start_date = datetime.datetime.combine(start_date, datetime.time())

                if isinstance(end_date, str):
                    end_date = self._parse_excel_date(end_date)
                elif end_date:
                    end_date = datetime.datetime.combine(end_date, datetime.time())

                # Get progress from status if not set
                if progress is None and status in status_mapping:
                    progress = status_mapping[status]

                imported_tasks.append({
                    'id': task_id if task_id else len(imported_tasks) + 1,
                    'name': task_name.strip(),
                    'start': start_date or datetime.datetime.now(),
                    'end': end_date or datetime.datetime.now() + datetime.timedelta(days=30),
                    'level': level if level else 1,
                    'parent_id': parent_id,
                    'progress': progress if progress is not None else 0,
                    'status': status,
                    'responsible': responsible,
                    'comments': comments
                })

            if not imported_tasks:
                QMessageBox.information(self, "No Tasks", "No tasks found in Excel file.")
                return

            # Build tasks
            task_map = {}

            # Create all tasks
            for task_data in imported_tasks:
                task = Task(
                    uid=task_data['id'],
                    name=task_data['name'],
                    start=task_data['start'],
                    finish=task_data['end'],
                    outline_level=task_data['level'],
                    priority=Priority.MEDIUM
                )
                task.progress = task_data['progress']

                # SIMPLE: Attach grid notes directly to task.notes
                if task_data['comments'] or task_data['responsible']:
                    grid_note = self._create_grid_note_from_comments(
                        task_data['comments'],
                        task_data['responsible']
                    )
                    task.notes.append(grid_note)

                task_map[task_data['id']] = task

            # Build parent-child relationships
            root_tasks = []
            for task_data in imported_tasks:
                task = task_map[task_data['id']]
                parent_id = task_data['parent_id']

                if parent_id and parent_id in task_map:
                    parent_task = task_map[parent_id]
                    task.parent = parent_task
                    parent_task.children.append(task)
                else:
                    root_tasks.append(task)

            # Check if project exists (AUTOMATIC - case-insensitive matching)
            existing_project = None
            for proj in self.projects:
                if proj.name.lower() == project_name.lower():
                    existing_project = proj
                    print(f"[DEBUG] Found existing project: '{proj.name}' matches '{project_name}'")
                    break

            if existing_project:
                print(f"[DEBUG] Updating existing project: {existing_project.name}")
                # SIMPLE UPDATE: Update existing tasks, let auto_populate handle NotesItems
                update_count = 0
                create_count = 0

                for task_id, task in task_map.items():
                    existing_task = self.task_map.get(task_id)

                    if existing_task:
                        # Update existing
                        existing_task.progress = task.progress
                        existing_task.name = task.name
                        existing_task.start = task.start
                        existing_task.finish = task.finish
                        existing_task.notes = task.notes  # Update notes (grid notes attached during task creation)
                        update_count += 1
                    else:
                        # Add new task
                        task.project = existing_project
                        self.task_map[task_id] = task
                        if task.parent is None:
                            existing_project.root_tasks.append(task)
                        create_count += 1

                QMessageBox.information(
                    self,
                    "Import Complete",
                    f"Updated {update_count} tasks, Created {create_count} tasks"
                )
            else:
                # SIMPLE CREATE: Create new project
                project = Project(project_name, self._get_next_project_uid())
                project.root_tasks = root_tasks

                for task_id, task in task_map.items():
                    self.task_map[task_id] = task
                    task.project = project

                self.projects.append(project)

                QMessageBox.information(
                    self,
                    "Import Complete",
                    f"Created project '{project_name}' with {len(imported_tasks)} tasks"
                )

            # Remove duplicate projects (keep first occurrence of each name)
            seen_names = set()
            unique_projects = []
            for proj in self.projects:
                if proj.name.lower() not in seen_names:
                    seen_names.add(proj.name.lower())
                    unique_projects.append(proj)
                else:
                    print(f"[DEBUG] Removing duplicate project: {proj.name}")
            self.projects = unique_projects

            # Refresh view - auto_populate_from_project() will create NotesItems from tasks
            self.update_project_dates_and_view()
            self.save_project()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Import Error",
                f"Failed to import Excel file.\nError: {str(e)}"
            )
            import traceback
            traceback.print_exc()

    def export_to_excel_template(self):
        """Export current project to Excel template format"""

        if not self.drill_down_stack:
            QMessageBox.warning(
                self,
                "No Project Selected",
                "Please drill down into a project first."
            )
            return

        current_item = self.drill_down_stack[-1]

        # Must be a Project
        if not isinstance(current_item, Project):
            QMessageBox.warning(
                self,
                "Not a Project",
                "Please select a project (not a task) to export."
            )
            return

        current_project = current_item

        # File save dialog
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel Template",
            f"{current_project.name}.xlsx",
            "Excel Files (*.xlsx *.xlsm)"
        )

        if not file_name:
            return

        try:
            from openpyxl import load_workbook
            import shutil

            # Use template as base
            template_path = "VPM_Master_Template.xlsx"
            if not os.path.exists(template_path):
                QMessageBox.critical(
                    self,
                    "Template Not Found",
                    f"Template file '{template_path}' not found.\nPlease ensure it's in the same directory as the application."
                )
                return

            # Copy template to destination
            shutil.copy(template_path, file_name)

            # Load the copied file
            wb = load_workbook(file_name)
            tasks_sheet = wb['Tasks']

            # Clear existing data (keep headers)
            for row in range(6, tasks_sheet.max_row + 1):
                for col in range(1, 12):
                    tasks_sheet.cell(row, col).value = None

            # Flatten task hierarchy
            def flatten_tasks(tasks, parent_id=None, level=1):
                flat_list = []
                for task in tasks:
                    flat_list.append({
                        'task': task,
                        'level': level,
                        'parent_id': parent_id
                    })
                    # Recursively add children
                    if task.children:
                        flat_list.extend(
                            flatten_tasks(task.children, task.uid, level + 1)
                        )
                return flat_list

            flat_tasks = flatten_tasks(current_project.root_tasks)

            # Write to Excel
            start_row = 6  # After header

            for idx, item in enumerate(flat_tasks):
                task = item['task']
                level = item['level']
                parent_id = item['parent_id']
                row = start_row + idx

                # Visual indent
                indent = "  " * (level - 1)
                tree_symbol = self._get_tree_symbol(task, level)

                # Visible columns
                tasks_sheet.cell(row, 1).value = f"{tree_symbol}{indent}{task.name}"  # Task Name
                tasks_sheet.cell(row, 2).value = task.start  # Start Date
                tasks_sheet.cell(row, 3).value = task.finish  # End Date
                tasks_sheet.cell(row, 4).value = self._get_status_from_progress(task.progress)  # Status

                # Get responsible from notes
                responsible = self._get_responsible_from_notes(task, current_project)
                tasks_sheet.cell(row, 5).value = responsible  # Responsible

                # Get comments from notes
                comments = self._get_comments_from_notes(task, current_project)
                tasks_sheet.cell(row, 6).value = comments  # Comments

                # Hidden columns
                tasks_sheet.cell(row, 7).value = task.uid  # ID
                tasks_sheet.cell(row, 8).value = level  # Level
                tasks_sheet.cell(row, 9).value = parent_id if parent_id else ""  # Parent_ID
                tasks_sheet.cell(row, 10).value = task.progress  # Progress
                tasks_sheet.cell(row, 11).value = self._get_status_from_progress(task.progress)  # Current_Status

            # Save
            wb.save(file_name)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Exported {len(flat_tasks)} tasks to:\n{file_name}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export to Excel.\nError: {str(e)}"
            )
            import traceback
            traceback.print_exc()

    # Helper functions for Excel import/export

    def _create_grid_note_from_comments(self, comments_text, responsible_person):
        """SIMPLE: Create a grid Note object from Excel comments"""
        import time

        # Create grid note
        note = Note(
            text="",  # Grid notes don't use text field
            heading="Excel Comments",
            note_type="grid",
            x_pos=50,
            y_pos=50,
            width=712,
            height=547
        )

        # Build grid data from comments
        grid_data = {}

        # Header row
        grid_data[(0, 0)] = "Date"
        grid_data[(0, 1)] = "Update"
        grid_data[(0, 2)] = "Responsible"
        grid_data[(0, 3)] = "Status"

        # Parse comments into rows
        if comments_text:
            # Split by newlines
            comment_lines = str(comments_text).split('\n')
            row = 1

            for line in comment_lines:
                line = line.strip()
                if not line:
                    continue

                # Try to parse "MM/DD: text" format
                if ':' in line:
                    parts = line.split(':', 1)
                    date_part = parts[0].strip()
                    update_part = parts[1].strip() if len(parts) > 1 else ""

                    grid_data[(row, 0)] = date_part
                    grid_data[(row, 1)] = update_part
                    grid_data[(row, 2)] = responsible_person if responsible_person else ""
                    grid_data[(row, 3)] = ""
                    row += 1
                else:
                    # No date format, just add as update
                    grid_data[(row, 0)] = ""
                    grid_data[(row, 1)] = line
                    grid_data[(row, 2)] = responsible_person if responsible_person else ""
                    grid_data[(row, 3)] = ""
                    row += 1
        else:
            # No comments, just add responsible person row
            grid_data[(1, 0)] = ""
            grid_data[(1, 1)] = ""
            grid_data[(1, 2)] = responsible_person if responsible_person else ""
            grid_data[(1, 3)] = ""

        note.grid_data = grid_data
        return note

    def _parse_excel_date(self, date_str):
        """Parse date from Excel in various formats"""
        if not date_str:
            return None

        formats = [
            '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d',
            '%d-%b-%Y', '%b %d, %Y', '%m-%d-%Y'
        ]

        for fmt in formats:
            try:
                return datetime.datetime.strptime(str(date_str), fmt)
            except ValueError:
                continue

        return None

    def _get_next_project_uid(self):
        """Get next available project UID"""
        if not self.projects:
            return 1
        return max(p.uid for p in self.projects) + 1

    def _import_new_tracker_format(self, wb, project_name):
        """Import from NEW TRACKER format (VPM_New_Tracker.xlsx)"""

        try:
            tasks_sheet = wb['Tasks']

            # Read tasks starting from row 2 (row 1 is header)
            imported_tasks = []

            for row_idx in range(2, tasks_sheet.max_row + 1):
                task_name = tasks_sheet.cell(row_idx, 1).value  # A: Task Name
                if not task_name:
                    continue

                start_date = tasks_sheet.cell(row_idx, 2).value  # B: Start Date
                end_date = tasks_sheet.cell(row_idx, 3).value    # C: End Date
                # Duration is calculated (column D), skip
                status = tasks_sheet.cell(row_idx, 5).value      # E: Status
                owner = tasks_sheet.cell(row_idx, 6).value       # F: Owner
                notes = tasks_sheet.cell(row_idx, 7).value       # G: Notes
                task_id = tasks_sheet.cell(row_idx, 8).value     # H: ID (hidden)
                level = tasks_sheet.cell(row_idx, 9).value       # I: Level (hidden)
                parent_id = tasks_sheet.cell(row_idx, 10).value  # J: Parent_ID (hidden)
                # Dates_Locked (column K) is for Excel only, not imported

                # Parse dates
                if isinstance(start_date, str):
                    start_date = self._parse_excel_date(start_date)
                elif start_date:
                    start_date = datetime.datetime.combine(start_date, datetime.time())

                if isinstance(end_date, str):
                    end_date = self._parse_excel_date(end_date)
                elif end_date:
                    end_date = datetime.datetime.combine(end_date, datetime.time())

                # Map status to progress
                status_to_progress = {
                    'Not Started': 0,
                    'In Progress': 50,
                    'Completed': 100,
                    'Delayed': 25
                }
                progress = status_to_progress.get(status, 0)

                imported_tasks.append({
                    'id': task_id if task_id else len(imported_tasks) + 1,
                    'name': task_name.strip(),
                    'start': start_date or datetime.datetime.now(),
                    'end': end_date or datetime.datetime.now() + datetime.timedelta(days=30),
                    'level': level if level else 1,
                    'parent_id': parent_id,
                    'progress': progress,
                    'status': status,
                    'owner': owner,
                    'notes': notes
                })

            if not imported_tasks:
                QMessageBox.information(self, "No Tasks", "No tasks found in Excel file.")
                return

            # Build tasks
            task_map = {}

            # Create all tasks
            for task_data in imported_tasks:
                task = Task(
                    uid=task_data['id'],
                    name=task_data['name'],
                    start=task_data['start'],
                    finish=task_data['end'],
                    outline_level=task_data['level'],
                    priority=Priority.MEDIUM
                )
                task.progress = task_data['progress']

                # Create grid note from Notes column
                if task_data['notes']:
                    grid_note = self._create_grid_note_from_notes_column(
                        task_data['notes'],
                        task_data['owner']
                    )
                    task.notes.append(grid_note)

                task_map[task_data['id']] = task

            # Build parent-child relationships
            root_tasks = []
            for task_data in imported_tasks:
                task = task_map[task_data['id']]
                parent_id = task_data['parent_id']

                if parent_id and parent_id in task_map:
                    parent_task = task_map[parent_id]
                    task.parent = parent_task
                    parent_task.children.append(task)
                else:
                    root_tasks.append(task)

            # Check if project exists (case-insensitive)
            existing_project = None
            for proj in self.projects:
                if proj.name.lower() == project_name.lower():
                    existing_project = proj
                    print(f"[DEBUG] Found existing project: '{proj.name}' matches '{project_name}'")
                    break

            if existing_project:
                print(f"[DEBUG] Updating existing project: {existing_project.name}")
                # Update existing tasks
                update_count = 0
                create_count = 0

                for task_id, task in task_map.items():
                    existing_task = self.task_map.get(task_id)

                    if existing_task:
                        # Update existing
                        existing_task.progress = task.progress
                        existing_task.name = task.name
                        existing_task.start = task.start
                        existing_task.finish = task.finish
                        existing_task.notes = task.notes
                        update_count += 1
                    else:
                        # Add new task
                        task.project = existing_project
                        self.task_map[task_id] = task
                        if task.parent is None:
                            existing_project.root_tasks.append(task)
                        create_count += 1

                QMessageBox.information(
                    self,
                    "Import Complete",
                    f"Updated {update_count} tasks, Created {create_count} tasks"
                )
            else:
                # Create new project
                project = Project(project_name, self._get_next_project_uid())
                project.root_tasks = root_tasks

                for task_id, task in task_map.items():
                    self.task_map[task_id] = task
                    task.project = project

                self.projects.append(project)

                QMessageBox.information(
                    self,
                    "Import Complete",
                    f"Created project '{project_name}' with {len(imported_tasks)} tasks"
                )

            # Remove duplicate projects
            seen_names = set()
            unique_projects = []
            for proj in self.projects:
                if proj.name.lower() not in seen_names:
                    seen_names.add(proj.name.lower())
                    unique_projects.append(proj)
                else:
                    print(f"[DEBUG] Removing duplicate project: {proj.name}")
            self.projects = unique_projects

            # Refresh view
            self.update_project_dates_and_view()
            self.save_project()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Import Error",
                f"Failed to import new tracker format.\nError: {str(e)}"
            )
            import traceback
            traceback.print_exc()

    def _create_grid_note_from_notes_column(self, notes_text, owner):
        """Create grid Note from Notes column text ([MM/DD]: format)"""

        note = Note(
            text="",
            heading="Task Notes",
            note_type="grid",
            x_pos=50,
            y_pos=50,
            width=712,
            height=547
        )

        grid_data = {}

        # Header row
        grid_data[(0, 0)] = "Date"
        grid_data[(0, 1)] = "Note"
        grid_data[(0, 2)] = "Owner"

        # Parse notes (format: [MM/DD]: text)
        if notes_text:
            lines = str(notes_text).split('\n')
            row = 1

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Try to parse [MM/DD]: format
                if line.startswith('[') and ']:' in line:
                    parts = line.split(']:', 1)
                    date_part = parts[0].replace('[', '').strip()
                    note_part = parts[1].strip() if len(parts) > 1 else ''

                    grid_data[(row, 0)] = date_part
                    grid_data[(row, 1)] = note_part
                    grid_data[(row, 2)] = owner if owner else ""
                    row += 1
                else:
                    # Plain text line
                    grid_data[(row, 0)] = ""
                    grid_data[(row, 1)] = line
                    grid_data[(row, 2)] = owner if owner else ""
                    row += 1

        note.grid_data = grid_data
        return note

    def _update_task_notes_from_excel(self, task, project, comments_text, responsible_person):
        """Create/update grid notes from Excel comments"""

        print(f"[DEBUG] _update_task_notes_from_excel called for task: {task.name}")
        print(f"[DEBUG]   comments_text: {comments_text}")
        print(f"[DEBUG]   responsible_person: {responsible_person}")

        if not comments_text and not responsible_person:
            print(f"[DEBUG]   Skipping - no comments or responsible person")
            return

        # Get or create notes structure
        if not hasattr(self, 'notes_view'):
            print(f"[DEBUG]   ERROR: notes_view not found!")
            return

        if not self.notes_view.notes_structure:
            print(f"[DEBUG]   ERROR: notes_structure not found!")
            return

        print(f"[DEBUG]   notes_view and notes_structure found")

        # Get project UID
        project_uid = str(project.uid)
        print(f"[DEBUG]   project_uid: {project_uid}")

        # Get project structure
        project_structure = self.notes_view.notes_structure.get_project_structure(project_uid)
        print(f"[DEBUG]   project_structure items count: {len(project_structure['items'])}")

        # Find existing notes item for this task (search recursively including children)
        def find_notes_item_recursive(task_uid, items):
            """Recursively search for NotesItem by task UID in items and their children."""
            for item in items:
                if item.linked_task and item.linked_task.uid == task_uid:
                    return item
                # Search in children
                found = find_notes_item_recursive(task_uid, item.children)
                if found:
                    return found
            return None

        notes_item = find_notes_item_recursive(task.uid, project_structure['items'])

        # Create new notes item if doesn't exist
        if not notes_item:
            print(f"[DEBUG]   Creating new NotesItem for task {task.uid}")
            notes_item = NotesItem(
                task.name,
                f'notes_{task.uid}',
                task,
                False,
                level=2
            )
            project_structure['items'].append(notes_item)
            print(f"[DEBUG]   NotesItem created and added")
        else:
            print(f"[DEBUG]   Found existing NotesItem")

        # Set to grid mode
        notes_item.view_mode = 'grid'

        # Parse comments into grid data
        grid_data = {}

        # Headers (4 columns)
        grid_data[(0, 0)] = 'Date'
        grid_data[(0, 1)] = 'Update'
        grid_data[(0, 2)] = 'Responsible'
        grid_data[(0, 3)] = 'Status'

        if comments_text:
            lines = str(comments_text).split('\n')

            row = 1
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Try to parse "MM/DD: Update text"
                if ':' in line:
                    parts = line.split(':', 1)
                    date_str = parts[0].strip()
                    update_text = parts[1].strip() if len(parts) > 1 else ''

                    grid_data[(row, 0)] = date_str
                    grid_data[(row, 1)] = update_text
                    grid_data[(row, 2)] = responsible_person or ''
                    grid_data[(row, 3)] = ''

                    row += 1
                else:
                    # Line without date separator
                    grid_data[(row, 0)] = ''
                    grid_data[(row, 1)] = line
                    grid_data[(row, 2)] = responsible_person or ''
                    grid_data[(row, 3)] = ''
                    row += 1
        else:
            # No comments yet, but add responsible person
            if responsible_person:
                grid_data[(1, 0)] = ''
                grid_data[(1, 1)] = 'Task assigned'
                grid_data[(1, 2)] = responsible_person
                grid_data[(1, 3)] = ''

        print(f"[DEBUG]   Grid data created with {len(grid_data)} cells")
        print(f"[DEBUG]   Grid data keys: {list(grid_data.keys())[:10]}")  # First 10 keys

        # Create a Note object with the grid data
        # Clear existing notes first
        notes_item.notes.clear()

        # Create grid note with ACTUAL Excel comments
        note = Note(
            text="",  # Grid notes don't use text
            heading="Excel Comments",
            note_type="grid",
            x_pos=50,
            y_pos=50,
            width=712,
            height=547,
            reminder_date=task.finish
        )
        # Use the REAL grid_data parsed from Excel
        note.grid_data = grid_data
        notes_item.add_note(note)
        print(f"[DEBUG] Created grid note from Excel with {len(note.grid_data)} cells")
        print(f"[DEBUG] Grid data sample: {list(grid_data.items())[:5]}")

        print(f"[DEBUG]   Created Note object with grid data")
        print(f"[DEBUG]   Note UID: {note.uid}, type: {note.note_type}")
        print(f"[DEBUG]   NotesItem '{notes_item.name}' now has {len(notes_item.notes)} notes")
        print(f"[DEBUG]   NotesItem id: {id(notes_item)}")

        # Don't save config here - will save after building hierarchy
        print(f"[DEBUG]   NotesItem updated (config will be saved after hierarchy is built)")

        return notes_item

    def _get_responsible_from_notes(self, task, project):
        """Extract responsible person from grid notes"""

        if not hasattr(self, 'notes_view') or not self.notes_view.notes_structure:
            return ""

        project_uid = str(project.uid)
        project_structure = self.notes_view.notes_structure.get_project_structure(project_uid)

        # Find notes for this task
        for item in project_structure['items']:
            if item.linked_task and item.linked_task.uid == task.uid:
                grid_data = item.grid_data
                # Get responsible from first data row
                return grid_data.get((1, 2), '')

        return ""

    def _get_comments_from_notes(self, task, project):
        """Extract comments from grid notes and format for Excel"""

        if not hasattr(self, 'notes_view') or not self.notes_view.notes_structure:
            return ""

        project_uid = str(project.uid)
        project_structure = self.notes_view.notes_structure.get_project_structure(project_uid)

        # Find notes for this task
        for item in project_structure['items']:
            if item.linked_task and item.linked_task.uid == task.uid:
                grid_data = item.grid_data

                # Reconstruct comment lines
                comments = []
                row = 1
                while (row, 0) in grid_data or (row, 1) in grid_data:
                    date_str = grid_data.get((row, 0), '')
                    update_text = grid_data.get((row, 1), '')

                    if date_str or update_text:
                        if date_str:
                            comments.append(f"{date_str}: {update_text}")
                        else:
                            comments.append(update_text)

                    row += 1

                return '\n'.join(comments)

        return ""

    def _get_status_from_progress(self, progress):
        """Map progress % to status name"""
        if progress == 0:
            return "Not Started"
        elif progress == 100:
            return "Completed"
        elif progress > 0 and progress < 100:
            return "In Progress"
        else:
            return "Not Started"

    def _get_tree_symbol(self, task, level):
        """Get tree symbol for visual hierarchy"""
        if level == 1:
            return "▼ "
        elif task.parent and task == task.parent.children[-1]:
            return "└─ "
        else:
            return "├─ "

    # ==================== END EXCEL TEMPLATE FUNCTIONS ====================

    def save_project(self):
        if not (self.projects or self.ecn_projects):
            QMessageBox.warning(self, "No Projects", "There is nothing to save.")
            return
        if self.current_file_path: self._write_to_file(self.current_file_path)
        else: self.save_project_as()

    def save_project_as(self):
        if not (self.projects or self.ecn_projects):
            QMessageBox.warning(self, "No Projects", "There is nothing to save.")
            return
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Workspace As", "", "VPMW Files (*.vpmw)")
        if file_name: self.current_file_path = file_name; self._write_to_file(self.current_file_path)

    def _write_to_file(self, file_path):
        try:
            print("--- DEBUG: Starting Save Process ---")
            projects_to_save = self.projects
            print(f"--- DEBUG: Found {len(projects_to_save)} XML-based project(s) to save.")
            
            data_to_save = {
                'projects': [],
                'drill_down_stack_uids': [item.uid for item in self.drill_down_stack]
            }
            for project in projects_to_save:
                print(f"--- DEBUG: Processing project '{project.name}'.")
                if isinstance(project, Project):
                    tasks_to_serialize = project.get_savable_tasks()
                    print(f"--- DEBUG: Found {len(tasks_to_serialize)} savable tasks in '{project.name}'.")
                    proj_data = {
                        'name': project.name, 'uid': project.uid, 'source_plugin': project.source_plugin,
                        'tasks': [task.to_dict() for task in tasks_to_serialize]
                    }
                    data_to_save['projects'].append(proj_data)

            with open(file_path, 'w') as f: json.dump(data_to_save, f, indent=4)
            print("--- DEBUG: Save process completed successfully. ---")
            self.setWindowTitle(f"Visual Project Manager - {self.current_file_path}")
            self.statusBar().showMessage(f"Workspace saved to {file_path}", 3000)
        except Exception as e: 
            print(f"--- DEBUG: SAVE FAILED. Error: {e} ---")
            import traceback
            print(traceback.format_exc())
            QMessageBox.critical(self, "Save Error", f"Failed to save workspace file.\nError: {e}")

    def parse_ms_project_xml(self, file_path):
        tasks_flat = []
        try:
            tree = ET.parse(file_path); root = tree.getroot()
            ns = {'ns': root.tag.split('}')[0][1:]} if '}' in root.tag else {}
            for task_elem in root.findall('.//ns:Task', ns) if ns else root.findall('.//Task'):
                uid_elem, name_elem, start_elem, finish_elem, level_elem = (task_elem.find(f'ns:{tag}', ns) if ns else task_elem.find(tag) for tag in ['UID', 'Name', 'Start', 'Finish', 'OutlineLevel'])
                if any(e is None for e in [uid_elem, name_elem, start_elem, finish_elem, level_elem]): continue
                
                def parse_date(s): return datetime.datetime.strptime(s.split('.')[0], '%Y-%m-%dT%H:%M:%S') if s else None
                
                if (s_date := parse_date(start_elem.text)) and (f_date := parse_date(finish_elem.text)):
                    task = Task(int(uid_elem.text), name_elem.text, s_date, f_date, int(level_elem.text))
                    tasks_flat.append(task)
            
            for i, task in enumerate(tasks_flat):
                if task.outline_level > 1:
                    for j in range(i - 1, -1, -1):
                        if tasks_flat[j].outline_level == task.outline_level - 1:
                            task.parent = tasks_flat[j]
                            break
        except ET.ParseError as e: raise Exception(f"Invalid XML file. Details: {e}")
        return tasks_flat
    
    def update_project_dates_and_view(self):
        all_projects = self.projects + self.ecn_projects
        for p in all_projects:
            if hasattr(p, 'update_dates'):
                p.update_dates()
            
        valid_projects = [p for p in all_projects if p.start and p.finish]
        if valid_projects:
            start_date = min(p.start.date() for p in valid_projects)
            end_date = max(p.finish.date() for p in valid_projects)
            self.timeline_canvas.set_project_duration(start_date, end_date)
        else:
            self.timeline_canvas.set_project_duration(None, None)
        
        self.toggle_timescale(force_timescale=self.current_timescale)
        self.update_view()

    def update_view(self):
        self.update_breadcrumb()
        self.timeline_canvas.clear_highlight()
        all_display_items = self.projects + self.ecn_projects
        if not self.drill_down_stack:
            self.timeline_canvas.set_tasks(all_display_items)
            print(f"DEBUG: Home screen showing {len(all_display_items)} projects")
        else:
            last_item = self.drill_down_stack[-1]
            valid_children = [c for c in last_item.children if c.start]
            sorted_children = sorted(valid_children, key=lambda t: t.start)

            self.timeline_canvas.set_tasks([last_item] + sorted_children)
            print(f"DEBUG: Drill-down view showing {len([last_item] + sorted_children)} items")
        self.check_for_notifications()

        # REMOVED: Do NOT reload notes_structure here - it wipes out programmatically created notes
        # Notes structure should only be loaded ONCE at startup in __init__

    def switch_view(self, view_index):
        """Switch between Timeline (0), Notes (1), and BOM Explorer (2) views."""
        self.view_stack.setCurrentIndex(view_index)

        # Update button states
        self.timeline_button.setChecked(view_index == 0)
        self.notes_button.setChecked(view_index == 1)
        self.bom_explorer_button.setChecked(view_index == 2)

        if view_index == 0:  # Timeline
            self.breadcrumb_frame.setVisible(True)
        elif view_index == 1:  # Notes
            self.breadcrumb_frame.setVisible(False)
            # Update notes view with current projects
            all_projects = self.projects + self.ecn_projects
            self.notes_view.set_projects(all_projects)
        else:  # BOM Explorer
            self.breadcrumb_frame.setVisible(False)

    def add_task_to_notes(self, task):
        """
        Add any task/item from Gantt to Notes view.
        Works for: Task, ProblemReport, ECN, Item, Project, etc.
        """
        # Find which project this task belongs to
        project_uid = None

        # Try to get project from task
        if hasattr(task, 'project') and task.project:
            project_uid = task.project.uid
        elif hasattr(task, 'uid'):
            # Check if task itself is a project
            for proj in self.projects + self.ecn_projects:
                if proj.uid == task.uid:
                    project_uid = task.uid
                    break
                # Check if task is in this project
                if hasattr(proj, 'root_tasks'):
                    all_tasks = proj.get_all_tasks() if hasattr(proj, 'get_all_tasks') else proj.root_tasks
                    if any(t.uid == task.uid for t in all_tasks):
                        project_uid = proj.uid
                        break

        if not project_uid:
            QMessageBox.warning(self, "Cannot Add", "Could not determine which project this item belongs to.")
            return

        # Add to notes structure
        success = self.notes_view.notes_structure.add_task_manually(task, project_uid)

        if success:
            QMessageBox.information(self, "Added to Notes", f"'{task.name}' has been added to Notes view.")
            # Refresh notes view if currently viewing
            if self.view_stack.currentIndex() == 1:
                self.switch_view(1)
        else:
            QMessageBox.information(self, "Already in Notes", f"'{task.name}' is already in the Notes view.")

    def check_note_statuses(self):
        """
        Check all notes for status updates and trigger UI updates.
        Runs every 60 seconds to update overdue indicators.
        """
        # Update status for all notes in the notes structure
        for project_data in self.notes_view.notes_structure.projects_notes.values():
            for item in project_data['items']:
                for note in item.get_all_notes():
                    note.update_status()

        # Trigger UI update in notes view
        if self.view_stack.currentIndex() == 1:  # If Notes view is active
            # Refresh the canvas if an item is selected
            if self.notes_view.current_item:
                # Force refresh of containers
                for container in self.notes_view.notes_canvas.containers:
                    container.update()

            # Update cascading indicators in middle column
            self.notes_view.items_list.update_indicators()

            # Update cascading indicators in tabs
            self.notes_view.child_tabs.update_indicators()

    def check_for_notifications(self):
        self.late_tasks, self.due_soon_tasks, self.reminders = [], [], []
        
        # Include both regular Task objects and ECN items
        all_tasks = [t for t in self.task_map.values() if isinstance(t, Task)]
        all_ecn_items = [t for t in self.task_map.values() if hasattr(t, 'is_ecn_item') and t.is_ecn_item]
        all_items = all_tasks + all_ecn_items

        today = datetime.date.today(); next_week = today + datetime.timedelta(days=7)
        self.late_tasks = [t for t in all_items if hasattr(t, 'is_active') and t.is_active and hasattr(t, 'children') and not t.children and hasattr(t, 'finish') and t.finish and t.finish.date() < today and hasattr(t, 'effective_progress') and t.effective_progress < 100]
        self.due_soon_tasks = [t for t in all_items if hasattr(t, 'is_active') and t.is_active and hasattr(t, 'children') and not t.children and hasattr(t, 'finish') and t.finish and today <= t.finish.date() <= next_week and hasattr(t, 'effective_progress') and t.effective_progress < 100]
        
        # Check for reminders in all items (both regular tasks and ECN items)
        for task in self.task_map.values():
            if hasattr(task, 'is_active') and task.is_active and hasattr(task, 'comments'):
                for comment in task.comments:
                    if 'reminder' in comment and not comment.get('reminder_complete', False):
                        reminder_date = datetime.datetime.strptime(comment['reminder'], '%Y-%m-%d').date()
                        if reminder_date >= today:
                            self.reminders.append((task, comment))
                            
        notification_count = len(self.late_tasks) + len(self.due_soon_tasks) + len(self.reminders)
        if notification_count > 0:
            self.notification_button.setText(f"🔔 ({notification_count})"); self.notification_button.setStyleSheet("color: red; border: none; font-weight: bold;"); self.notification_button.setEnabled(True)
        else:
            self.notification_button.setText("🔔"); self.notification_button.setStyleSheet("border: none;"); self.notification_button.setEnabled(False)

    def show_notifications(self):
        self.check_for_notifications()
        print(f"DEBUG: Found {len(self.late_tasks)} late tasks, {len(self.due_soon_tasks)} due soon tasks, {len(self.reminders)} reminders")
        dialog = NotificationDialog(self.late_tasks, self.due_soon_tasks, self.reminders, self)
        dialog.tasks_to_complete.connect(self.complete_tasks_from_ids)
        dialog.reminders_to_complete.connect(self.complete_reminders)
        dialog.task_selected.connect(self.navigate_to_task_from_uid)
        dialog.reminder_selected.connect(self.navigate_to_task_from_uid)
        dialog.exec()
    
    def navigate_to_task(self, task_object):
        if not task_object: return
        self.navigate_to_task_from_uid(task_object.uid)
        
    def navigate_to_task_from_uid(self, task_uid):
        if not task_uid: 
            print(f"DEBUG: No task_uid provided")
            return
            
        task = self.task_map.get(task_uid)
        print(f"DEBUG: Looking for task_uid: {task_uid}")
        print(f"DEBUG: Found task: {task}")
        print(f"DEBUG: Task map contains {len(self.task_map)} items")
        print(f"DEBUG: Sample task_map keys: {list(self.task_map.keys())[:5]}")

        if not task: 
            print(f"DEBUG: Task not found in task_map")
            return
        
        new_stack = []
        current = task
        print(f"DEBUG: Starting navigation from task: {current.name if hasattr(current, 'name') else 'Unknown'}")
        
        # Build the drill-down stack by traversing up the hierarchy
        while getattr(current, 'parent', None):
            new_stack.insert(0, current.parent)
            current = current.parent
            print(f"DEBUG: Added parent to stack: {current.name if hasattr(current, 'name') else 'Unknown'}")
        
        # Find the top-level project/PR and add it to the stack
        for p in self.projects + self.ecn_projects:
            if current in p.children or current == p:
                if p not in new_stack:
                    new_stack.insert(0, p)
                    print(f"DEBUG: Added top-level project to stack: {p.name}")
                break

        # Remove duplicates while preserving order
        unique_stack = []
        for item in new_stack:
            if item not in unique_stack:
                unique_stack.append(item)

        print(f"DEBUG: Final drill-down stack: {[item.name if hasattr(item, 'name') else 'Unknown' for item in unique_stack]}")
        self.drill_down_stack = unique_stack
        self.update_view()
        
        # Highlight the searched task
        print(f"DEBUG: Attempting to highlight task: {task_uid}")
        self.timeline_canvas.highlight_task(task_uid)

    def complete_tasks_from_ids(self, uids):
        for uid in uids:
            if task := self.task_map.get(uid): task.progress = 100
        self.update_view()

    def complete_reminders(self, reminders_to_mark):
        for uid, timestamp in reminders_to_mark:
            if task := self.task_map.get(uid):
                for comment in task.comments:
                    if comment['timestamp'] == timestamp: comment['reminder_complete'] = True; break
        self.update_view()

    def update_breadcrumb(self):
        while self.breadcrumb_layout.count():
            if (widget := self.breadcrumb_layout.takeAt(0).widget()) is not None: widget.setParent(None)
        home_btn = QPushButton("Home"); home_btn.setStyleSheet("border:none; color: blue; text-decoration: underline;"); home_btn.clicked.connect(partial(self.navigate_to_breadcrumb, -1)); self.breadcrumb_layout.addWidget(home_btn)
        for i, item in enumerate(self.drill_down_stack):
            self.breadcrumb_layout.addWidget(QLabel(">"))
            level_btn = QPushButton(item.name); level_btn.setStyleSheet("border:none; color: blue; text-decoration: underline;"); level_btn.clicked.connect(partial(self.navigate_to_breadcrumb, i)); self.breadcrumb_layout.addWidget(level_btn)
        self.breadcrumb_layout.addStretch()

    def navigate_to_breadcrumb(self, index):
        self.drill_down_stack = [] if index == -1 else self.drill_down_stack[:index+1]
        self.update_view()
    def perform_drill_down(self, item):
        if hasattr(item, 'is_ecn_item') and item.is_ecn_item:
            self.ecn_engine.handle_drill_down(item)
            return

        if self.drill_down_stack and self.drill_down_stack[-1].uid == item.uid: return
        if item.children:
            self.drill_down_stack.append(item); self.update_view()
        elif isinstance(item, Task):
            if QMessageBox.question(self, 'Create New Level?', f"'{item.name}' has no sub-tasks. Create a new level?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
                self.drill_down_stack.append(item); self.update_view()

    def show_add_item_dialog(self):
        if not self.projects:
            QMessageBox.warning(self, "No Project", "Please import or open a project first.")
            return
        parent_item = self.drill_down_stack[-1] if self.drill_down_stack else None
        if not parent_item and len(self.projects) > 1:
            QMessageBox.warning(self, "Select Project", "Please drill into a specific project before adding items.")
            return

        dialog = AddItemDialog(self)
        if not dialog.exec():
            return
            
        tasks_to_add = dialog.get_tasks()
        if not tasks_to_add:
            return
        
        # Handle mixed UID types (int for regular tasks, str for ECN items)
        if self.task_map:
            int_uids = [uid for uid in self.task_map.keys() if isinstance(uid, int)]
            current_max_uid = max(int_uids) if int_uids else 0
        else:
            current_max_uid = 0
        next_uid = current_max_uid + 1

        target_project = None
        if parent_item and isinstance(parent_item, Task): target_project = parent_item.project
        elif parent_item and isinstance(parent_item, Project): target_project = parent_item
        else:
            if len(self.projects) == 1:
                target_project = self.projects[0]
        
        if not target_project:
             QMessageBox.critical(self, "Error", "Could not determine the target project.")
             return

        for data in tasks_to_add:
            new_outline_level = (parent_item.outline_level + 1) if parent_item and isinstance(parent_item, Task) else 1
            start_dt = datetime.datetime.combine(data["start"], datetime.time.min)
            end_dt = datetime.datetime.combine(data["start"] if data["is_milestone"] else data["end"], datetime.time.max)
            new_task = Task(next_uid, data["name"], start_dt, end_dt, new_outline_level, data["priority"])
            
            if data["description"]:
                comment = {'text': data["description"], 'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                new_task.comments.append(comment)

            new_task.project = target_project
            new_task.parent = parent_item if isinstance(parent_item, Task) else None
            
            if new_task.parent:
                new_task.parent.children.append(new_task)
            else:
                target_project.root_tasks.append(new_task)

            self.task_map[next_uid] = new_task
            next_uid += 1
        
        print(f"DEBUG: Added {len(tasks_to_add)} new tasks to project '{target_project.name}'")
        print(f"DEBUG: Project now has {len(target_project.root_tasks)} root tasks")
        print(f"DEBUG: Task map size: {len(self.task_map)}")
        
        # If we're at home screen and added tasks to a project, drill down to show them
        if not self.drill_down_stack and target_project:
            self.drill_down_stack.append(target_project)
            print(f"DEBUG: Auto-drilled down into project '{target_project.name}' to show new tasks")
        
        self.update_project_dates_and_view()

    def edit_item(self, task_to_edit):
        dialog = AddItemDialog(self)
        dialog.setWindowTitle(f"Edit '{task_to_edit.name}'")
        dialog.table.setItem(0, 0, QTableWidgetItem(task_to_edit.name))
        
        desc = ""
        if task_to_edit.comments:
            desc = task_to_edit.comments[0]['text']
        dialog.table.setItem(0, 1, QTableWidgetItem(desc))
        
        dialog.table.cellWidget(0, 2).findChild(QCheckBox).setChecked(task_to_edit.is_milestone)
        dialog.table.cellWidget(0, 3).setDate(task_to_edit.start.date())
        dialog.table.cellWidget(0, 4).setDate(task_to_edit.finish.date())
        dialog.table.cellWidget(0, 5).setCurrentIndex(task_to_edit.priority.value - 1)
        
        if dialog.exec():
            edited_tasks = dialog.get_tasks()
            if not edited_tasks: return
            data = edited_tasks[0]
            
            task_to_edit.name = data['name']
            task_to_edit.priority = data['priority']
            task_to_edit.start = datetime.datetime.combine(data["start"], datetime.time.min)
            task_to_edit.finish = datetime.datetime.combine(data["start"] if data["is_milestone"] else data["end"], datetime.time.max)
            task_to_edit.is_milestone = (task_to_edit.finish.date() - task_to_edit.start.date()).days <= 0

            if data['description']:
                if task_to_edit.comments:
                    task_to_edit.comments[0]['text'] = data['description']
                else:
                    task_to_edit.comments.append({'text': data["description"], 'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
            elif task_to_edit.comments:
                task_to_edit.comments.pop(0)

            self.update_project_dates_and_view()

    def edit_project(self, project_to_edit):
        """Edit project name using a simple dialog"""
        from PyQt6.QtWidgets import QInputDialog
        
        new_name, ok = QInputDialog.getText(self, 'Edit Project Name', 
                                           'Enter new project name:', 
                                           text=project_to_edit.name)
        
        if ok and new_name.strip():
            old_name = project_to_edit.name
            project_to_edit.name = new_name.strip()
            print(f"DEBUG: Renamed project from '{old_name}' to '{new_name.strip()}'")
            self.update_project_dates_and_view()
            
            # Update window title if this is the current project
            if self.current_file_path:
                self.setWindowTitle(f"Visual Project Manager - {self.current_file_path}")

    def delete_item(self, task_to_delete):
        # Show the delete options dialog
        dialog = DeleteOptionsDialog(task_to_delete.name, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return  # User cancelled
        
        delete_option = dialog.get_delete_option()
        
        if delete_option == 'delete_all':
            # Original behavior: delete everything recursively
            self._recursively_delete_task(task_to_delete)
        else:  # delete_level_only
            # New behavior: delete only this level, move children up
            self._delete_level_only(task_to_delete)
        
        # Clean up ECN metadata if needed
        if hasattr(task_to_delete, 'is_ecn_item') and task_to_delete.is_ecn_item:
            if 'ecn_metadata' in self.ecn_config and task_to_delete.uid in self.ecn_config['ecn_metadata']:
                del self.ecn_config['ecn_metadata'][task_to_delete.uid]
                self.save_ecn_config()
        
        # Clean up drill_down_stack if the deleted item was in it
        if task_to_delete in self.drill_down_stack:
            # Find the index of the deleted item
            deleted_index = self.drill_down_stack.index(task_to_delete)
            # Remove the deleted item and everything after it from the stack
            self.drill_down_stack = self.drill_down_stack[:deleted_index]
            print(f"DEBUG: Cleaned drill_down_stack, now has {len(self.drill_down_stack)} items")
        
        # Force update of project dates and view
        self.update_project_dates_and_view()
        
        # Debug: Print current state
        print(f"DEBUG: After delete, projects: {len(self.projects)}, ecn_projects: {len(self.ecn_projects)}")
        print(f"DEBUG: Task map size: {len(self.task_map)}")
        if self.projects:
            for p in self.projects:
                print(f"DEBUG: Project '{p.name}' has {len(p.root_tasks)} root tasks")

    def _delete_level_only(self, task):
        """Delete only this level and move all children up one level"""
        # Get all children before we modify the structure
        children = list(getattr(task, 'children', []))
        
        # Find the parent of this task
        parent = getattr(task, 'parent', None)
        project = None
        
        # Special handling for project deletion
        if task in self.projects:
            # This is a project being deleted
            # Children should become root tasks of the same project
            # We keep the project but move its children to root level
            project = task
            new_parent = None
            print(f"DEBUG: Deleting project '{task.name}', moving {len(children)} children to root level")
        elif parent:
            # This task has a parent, children will move to the parent
            new_parent = parent
        elif isinstance(task, Task) and hasattr(task, 'project'):
            # This is a root task, children will become root tasks
            project = task.project
            new_parent = None
        else:
            # Fallback: treat as root task
            new_parent = None
        
        # Move all children to the new parent
        for child in children:
            child.parent = new_parent
            if new_parent:
                new_parent.children.append(child)
            elif project:
                # Only add to project.root_tasks if not already there
                if child not in project.root_tasks:
                    project.root_tasks.append(child)
                    safe_name = child.name.encode('ascii', errors='replace').decode('ascii')
                    print(f"DEBUG: Added child '{safe_name}' to project root tasks")
                else:
                    safe_name = child.name.encode('ascii', errors='replace').decode('ascii')
                    print(f"DEBUG: Child '{safe_name}' already in project root tasks, skipping")
            
            # Make sure child is in task_map if it's a Task
            if isinstance(child, Task) and hasattr(child, 'uid') and child.uid not in self.task_map:
                self.task_map[child.uid] = child
        
        # Now remove the task from its current location
        if parent:
            parent.children.remove(task)
        elif isinstance(task, Task) and hasattr(task, 'project') and task in task.project.root_tasks:
            task.project.root_tasks.remove(task)
        elif task in self.projects:
            # For projects, we need to be more careful
            # If this project has children, we don't remove it from self.projects
            # The children are now root tasks of this project
            # Only remove the project if it has no children
            if not children:  # Only remove if no children to preserve
                self.projects.remove(task)
                print(f"DEBUG: Removed empty project '{task.name}' from projects list")
            else:
                print(f"DEBUG: Kept project '{task.name}' with {len(children)} root tasks")
        elif task in self.ecn_projects:
            self.ecn_projects.remove(task)
        
        # Remove from task map
        if hasattr(task, 'uid') and task.uid in self.task_map:
            del self.task_map[task.uid]

    def _recursively_delete_task(self, task):
        for child in list(getattr(task, 'children', [])):
            self._recursively_delete_task(child)

        if getattr(task, 'parent', None):
            task.parent.children.remove(task)
        elif isinstance(task, Task) and hasattr(task, 'project') and task in task.project.root_tasks:
            task.project.root_tasks.remove(task)
        elif task in self.projects:
            self.projects.remove(task)
        elif task in self.ecn_projects:
            self.ecn_projects.remove(task)


        if hasattr(task, 'uid') and task.uid in self.task_map:
            del self.task_map[task.uid]

    def show_comment_dialog(self, task):
        dialog = CommentDialog(task, self)
        dialog.exec()
        if hasattr(task, 'is_ecn_item') and task.is_ecn_item:
            self.update_ecn_metadata(task)
        self.update_view()

    def show_ecn_summary(self, item):
        summary_data = self.ecn_engine.get_summary_data(item)
        if not summary_data:
            QMessageBox.information(self, "No Data", "No active items found to summarize.")
            return
        
        dialog = AdvancedSummaryDialog(summary_data, self)
        dialog.exec()
        
    def show_legend_dialog(self):
        dialog = LegendDialog(self)
        dialog.exec()

    def mark_phase_complete(self, task_to_complete):
        if QMessageBox.question(self, 'Confirm Completion', "This will mark this task and ALL sub-tasks as 100% complete. Are you sure?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            def recurse_complete(task):
                task.progress = 100
                for child in task.children: recurse_complete(child)
            recurse_complete(task_to_complete)
            self.update_view()

    def set_active_zone(self, task, set_active):
        def recurse_set_active(t):
            t.is_active = set_active
            for child in getattr(t, 'children', []): recurse_set_active(child)
        recurse_set_active(task)
        self.update_view()
    
    def perform_search(self):
        search_term = self.search_bar.text().lower()
        if not search_term:
            return

        results = []

        all_items_in_map = list(self.task_map.values())
        items_to_search = []

        if self.drill_down_stack:
            # Include the current drilled item, its children, AND its siblings
            root_item = self.drill_down_stack[-1]
            visible_uids = {root_item.uid}

            # Add all descendants of current item
            items_to_check = [root_item]
            while items_to_check:
                current = items_to_check.pop(0)
                if hasattr(current, 'children'):
                    for child in current.children:
                        if hasattr(child, 'uid'):
                            visible_uids.add(child.uid)
                        items_to_check.append(child)

            # Also add siblings — items at the same level (parent's other children)
            if len(self.drill_down_stack) >= 2:
                parent_item = self.drill_down_stack[-2]
                if hasattr(parent_item, 'children'):
                    for sibling in parent_item.children:
                        if hasattr(sibling, 'uid'):
                            visible_uids.add(sibling.uid)

            items_to_search = [item for item in all_items_in_map if hasattr(item, 'uid') and item.uid in visible_uids]

            # If no results at current level, search globally
            for item in items_to_search:
                self._find_matches_in_item(item, search_term, results)

            if not results:
                # Fall back to global search
                results_global = []
                for item in all_items_in_map:
                    self._find_matches_in_item(item, search_term, results_global)
                if results_global:
                    results = results_global
        else:
            items_to_search = all_items_in_map
            for item in items_to_search:
                self._find_matches_in_item(item, search_term, results)

        dialog = SearchResultsDialog(results, self)
        dialog.task_selected.connect(self.navigate_to_task)
        dialog.exec()

    def _find_matches_in_item(self, item, term, results):
        if hasattr(item, 'name') and term in item.name.lower():
            results.append((item, f"Match in name: '{item.name}'"))
        
        if hasattr(item, 'comments'):
            for comment in item.comments:
                if term in comment['text'].lower():
                    results.append((item, f"Match in comment on '{item.name}': '{comment['text'][:40]}...'"))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.toggle_timescale(force_timescale=self.current_timescale)

    def toggle_timescale(self, checked=False, force_timescale=None):
        all_projects = self.projects + self.ecn_projects
        if not all_projects: return
        if force_timescale: self.current_timescale = force_timescale
        else:
            if self.current_timescale == TimescaleView.WEEK: self.current_timescale = TimescaleView.MONTH
            elif self.current_timescale == TimescaleView.MONTH: self.current_timescale = TimescaleView.YEAR
            else: self.current_timescale = TimescaleView.WEEK
        
        if self.scroll_area.viewport():
            self.timeline_canvas.set_timescale(self.current_timescale, self.scroll_area.viewport().width())
        self.timescale_action.setText(f"Zoom: {self.current_timescale.name.capitalize()}")
        self.statusBar().showMessage(f"Zoom level set to: {self.current_timescale.name}", 2000)

    def show_summary(self):
        all_projects = self.projects + self.ecn_projects
        if not all_projects: QMessageBox.warning(self, "No Project", "Please load a project to see a summary."); return
        total_tasks = len(self.task_map)
        
        project_objects = self.projects
        overall_progress = sum(p.effective_progress for p in project_objects) / len(project_objects) if project_objects else 0
        
        all_tasks = self.task_map.values()
        late_tasks = [t for t in all_tasks if isinstance(t, Task) and t.finish and t.finish.date() < datetime.date.today() and t.effective_progress < 100 and not t.children]
        
        report = f"WORKSPACE STATUS SUMMARY\n========================\n\nOverall Progress: {overall_progress:.1f}%\nTotal Projects: {len(all_projects)}\nTotal Tasks: {total_tasks}\n\n"
        if late_tasks:
            report += "LATE TASKS:\n"
            for task in late_tasks: 
                if hasattr(task, 'project') and task.project:
                    report += f"- {task.project.name} / {task.name} (Due: {task.finish.strftime('%Y-%m-%d')})\n"
        else: report += "No tasks are currently late.\n"
        dialog = SummaryDialog(report, self); dialog.exec()

def exception_hook(exc_type, exc_value, exc_traceback):
    """Global exception handler to catch all unhandled exceptions"""
    import traceback
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    print(f"\n{'='*60}")
    print("UNHANDLED EXCEPTION - Application Crash")
    print(f"{'='*60}")
    print(error_msg)
    print(f"{'='*60}\n")
    
    # Try to show a message box if QApplication exists
    try:
        app = QApplication.instance()
        if app:
            QMessageBox.critical(
                None, 
                "Application Crash", 
                f"An unhandled exception occurred:\n\n{exc_type.__name__}: {exc_value}\n\n"
                f"See console for full traceback."
            )
    except:
        pass  # If we can't show a message box, at least we logged it

if __name__ == '__main__':
    # Install global exception handler
    sys.excepthook = exception_hook
    
    try:
        app = QApplication(sys.argv)
        # Set light theme palette
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.ToolTipText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
        palette.setColor(QPalette.ColorRole.Link, QColor(0, 0, 255))
        palette.setColor(QPalette.ColorRole.Highlight, QColor(0, 120, 212))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
        app.setPalette(palette)
        # Set comprehensive light theme stylesheet for dropdowns and menus
        app.setStyleSheet("""
            QComboBox {
                background-color: #ffffff;
                color: #000000;
                border: 1px solid #cccccc;
            }
            QComboBox:hover {
                border: 1px solid #999999;
            }
            QComboBox::drop-down {
                border: none;
                background-color: #f0f0f0;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #000000;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                color: #000000;
                selection-background-color: #0078d4;
                selection-color: #ffffff;
                border: 1px solid #cccccc;
            }
            QMenu {
                background-color: #ffffff;
                color: #000000;
                border: 1px solid #cccccc;
            }
            QMenu::item {
                background-color: #ffffff;
                color: #000000;
                padding: 4px 20px 4px 20px;
            }
            QMenu::item:selected {
                background-color: #0078d4;
                color: #ffffff;
            }
            QMenu::item:hover {
                background-color: #e8e8e8;
            }
        """)
        window = MainWindow()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"\n{'='*60}")
        print("FATAL ERROR DURING APPLICATION STARTUP")
        print(f"{'='*60}")
        print(error_msg)
        print(f"{'='*60}\n")
        
        # Try to show error dialog
        try:
            app = QApplication.instance() or QApplication(sys.argv)
            QMessageBox.critical(
                None,
                "Startup Error",
                f"Failed to start application:\n\n{type(e).__name__}: {e}\n\n"
                f"See console for full traceback."
            )
        except:
            pass
        sys.exit(1)


