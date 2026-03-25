import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

file_path = r'c:\Users\silam\OneDrive\Documents\GitHub\VPM-1\VPM_New_Tracker.xlsx'

# 1. Create Data Structure
columns = [
    "ID", "Task Name", "Start Date", "End Date", "Duration", 
    "Status", "Owner", "Notes", "Level", "Parent ID"
]

# Create a sample dataframe with one root task to start
data = [
    [1, "Project Root", None, None, None, "Not Started", "User", "Initial Task", 1, None]
]

df = pd.DataFrame(data, columns=columns)

# 2. Create Excel File with openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Project Tracker"

# Write header
for col_num, column_title in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Write initial data
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# 3. Create Table
tab = Table(displayName="ProjectTable", ref=f"A1:J{len(data)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

# 4. Column Widths & Formatting
column_widths = {
    "A": 8,   # ID
    "B": 50,  # Task Name (Wide)
    "C": 15,  # Start
    "D": 15,  # End
    "E": 10,  # Duration
    "F": 15,  # Status
    "G": 15,  # Owner
    "H": 60,  # Notes (Very Wide)
    "I": 8,   # Level (Hidden)
    "J": 8    # Parent ID (Hidden)
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Hide Level and Parent ID columns
ws.column_dimensions['I'].hidden = True
ws.column_dimensions['J'].hidden = True

# Wrap Text for Notes
ws.column_dimensions['H'].alignment = Alignment(wrap_text=True)

# 5. Conditional Formatting
# Status Colors
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
yellow_font = Font(color="9C5700")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006")

# Status Column is F
ws.conditional_formatting.add("F2:F1000", CellIsRule(operator="equal", formula=['"Completed"'], stopIfTrue=True, fill=green_fill, font=green_font))
ws.conditional_formatting.add("F2:F1000", CellIsRule(operator="equal", formula=['"In Progress"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
ws.conditional_formatting.add("F2:F1000", CellIsRule(operator="equal", formula=['"Delayed"'], stopIfTrue=True, fill=red_fill, font=red_font))

# 6. Data Validation for Status
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Delayed"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("F2:F1000")

# Save
wb.save(file_path)
print(f"Created {file_path}")
