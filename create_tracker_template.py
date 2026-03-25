"""
Script to generate VPM_New_Tracker.xlsx base template
Run this script to create the Excel file, then add VBA macros manually
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, timedelta

def create_vpm_tracker_template():
    """Create VPM_New_Tracker.xlsx with all formatting"""

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Define headers
    headers = [
        "Task Name",      # A
        "Start Date",     # B
        "End Date",       # C
        "Duration",       # D
        "Status",         # E
        "Owner",          # F
        "Notes",          # G
        "ID",             # H (hidden)
        "Level",          # I (hidden)
        "Parent_ID",      # J (hidden)
        "Dates_Locked",   # K (hidden)
        "",               # L (reserved)
        "Owner List"      # M (hidden - owner names)
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 40  # Task Name
    ws.column_dimensions['B'].width = 12  # Start Date
    ws.column_dimensions['C'].width = 12  # End Date
    ws.column_dimensions['D'].width = 10  # Duration
    ws.column_dimensions['E'].width = 15  # Status
    ws.column_dimensions['F'].width = 15  # Owner
    ws.column_dimensions['G'].width = 50  # Notes
    ws.column_dimensions['H'].width = 8   # ID (will be hidden)
    ws.column_dimensions['I'].width = 8   # Level (will be hidden)
    ws.column_dimensions['J'].width = 10  # Parent_ID (will be hidden)
    ws.column_dimensions['K'].width = 12  # Dates_Locked (will be hidden)
    ws.column_dimensions['L'].width = 10  # Reserved (will be hidden)
    ws.column_dimensions['M'].width = 15  # Owner List (will be hidden)

    # Hide columns H, I, J, K, L, M
    ws.column_dimensions['H'].hidden = True
    ws.column_dimensions['I'].hidden = True
    ws.column_dimensions['J'].hidden = True
    ws.column_dimensions['K'].hidden = True
    ws.column_dimensions['L'].hidden = True
    ws.column_dimensions['M'].hidden = True

    # Add sample data (3 rows)
    sample_data = [
        {
            'Task Name': 'Project Phase 1',
            'Start Date': datetime.now().date(),
            'End Date': (datetime.now() + timedelta(days=30)).date(),
            'Status': 'In Progress',
            'Owner': '',
            'Notes': '',
            'ID': 1,
            'Level': 1,
            'Parent_ID': '',
            'Dates_Locked': 'FALSE'
        },
        {
            'Task Name': '  > Task 1.1',
            'Start Date': datetime.now().date(),
            'End Date': (datetime.now() + timedelta(days=10)).date(),
            'Status': 'Not Started',
            'Owner': '',
            'Notes': '',
            'ID': 2,
            'Level': 2,
            'Parent_ID': 1,
            'Dates_Locked': 'FALSE'
        },
        {
            'Task Name': '  > Task 1.2',
            'Start Date': (datetime.now() + timedelta(days=11)).date(),
            'End Date': (datetime.now() + timedelta(days=30)).date(),
            'Status': 'Not Started',
            'Owner': '',
            'Notes': '',
            'ID': 3,
            'Level': 2,
            'Parent_ID': 1,
            'Dates_Locked': 'FALSE'
        }
    ]

    # Write sample data
    for row_idx, data in enumerate(sample_data, start=2):
        ws.cell(row_idx, 1).value = data['Task Name']
        ws.cell(row_idx, 2).value = data['Start Date']
        ws.cell(row_idx, 2).number_format = 'mm/dd/yyyy'
        ws.cell(row_idx, 3).value = data['End Date']
        ws.cell(row_idx, 3).number_format = 'mm/dd/yyyy'

        # Duration formula
        ws.cell(row_idx, 4).value = f"=C{row_idx}-B{row_idx}"
        ws.cell(row_idx, 4).number_format = '0'

        ws.cell(row_idx, 5).value = data['Status']
        ws.cell(row_idx, 6).value = data['Owner']
        ws.cell(row_idx, 7).value = data['Notes']
        ws.cell(row_idx, 7).alignment = Alignment(wrap_text=True, vertical='top')

        ws.cell(row_idx, 8).value = data['ID']
        ws.cell(row_idx, 9).value = data['Level']
        ws.cell(row_idx, 10).value = data['Parent_ID']
        ws.cell(row_idx, 11).value = data['Dates_Locked']

    # Add data validation for Status column (E2:E1000)
    status_validation = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Delayed"',
        allow_blank=False
    )
    status_validation.error = 'Invalid status'
    status_validation.errorTitle = 'Invalid Entry'
    status_validation.prompt = 'Select a status'
    status_validation.promptTitle = 'Status'
    ws.add_data_validation(status_validation)
    status_validation.add('E2:E1000')

    # Add owner list in column M (hidden)
    owner_names = ['Brett', 'John', 'Chris', 'Sunil']
    for idx, owner in enumerate(owner_names, start=2):
        ws.cell(idx, 13).value = owner  # Column M = 13

    # Create named range for owner list
    from openpyxl.workbook.defined_name import DefinedName
    owner_range = DefinedName('OwnerList', attr_text='Tasks!$M$2:$M$11')
    wb.defined_names.add(owner_range)

    # Add data validation for Owner column (F2:F1000)
    owner_validation = DataValidation(
        type="list",
        formula1='=OwnerList',
        allow_blank=True
    )
    owner_validation.error = 'Please select from the owner list'
    owner_validation.errorTitle = 'Invalid Owner'
    owner_validation.prompt = 'Select an owner (or right-click to edit list)'
    owner_validation.promptTitle = 'Owner'
    ws.add_data_validation(owner_validation)
    owner_validation.add('F2:F1000')

    # Conditional Formatting - Priority Order:
    # 1. Overdue (red) - highest priority
    # 2. Completed (green)
    # 3. Leaf tasks (light orange) - action items
    # 4. Parent tasks (light gray) - containers

    # PRIORITY 1: Overdue Tasks (Red background, highest priority)
    overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    overdue_font = Font(color="9C0006", bold=True)
    overdue_rule = FormulaRule(
        formula=['AND($C2<TODAY(),$E2<>"Completed")'],
        fill=overdue_fill,
        font=overdue_font,
        stopIfTrue=True  # This rule takes priority over all others
    )
    ws.conditional_formatting.add('A2:G1000', overdue_rule)

    # PRIORITY 2: Completed Tasks - Light Green
    completed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    completed_font = Font(color="006100")
    completed_rule = FormulaRule(
        formula=['$E2="Completed"'],
        fill=completed_fill,
        font=completed_font,
        stopIfTrue=True  # Stop after completed
    )
    ws.conditional_formatting.add('A2:G1000', completed_rule)

    # PRIORITY 3: Leaf Tasks (no children) - Light Orange (action items)
    # Formula: Count how many tasks have this task's ID as their Parent_ID
    # If count = 0, it's a leaf task (no children)
    leaf_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    leaf_rule = FormulaRule(
        formula=['COUNTIF($J:$J,$H2)=0'],
        fill=leaf_fill,
        stopIfTrue=True  # Stop after leaf tasks
    )
    ws.conditional_formatting.add('A2:G1000', leaf_rule)

    # PRIORITY 4: Parent Tasks (have children) - Light Gray (organizational)
    # Formula: If any tasks have this task's ID as their Parent_ID
    parent_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    parent_rule = FormulaRule(
        formula=['COUNTIF($J:$J,$H2)>0'],
        fill=parent_fill
    )
    ws.conditional_formatting.add('A2:G1000', parent_rule)

    # Conditional Formatting - Child dates exceed parent hard dates (Orange background on dates)
    warning_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    # This will be added via VBA dynamically since it requires parent lookup

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Set default row height
    ws.row_dimensions[1].height = 20
    for row in range(2, 1000):
        ws.row_dimensions[row].height = 15

    # Save workbook
    filename = "VPM_New_Tracker.xlsx"
    wb.save(filename)
    print(f"[OK] Created {filename}")
    print(f"[OK] Sample data: 3 tasks (1 parent, 2 children)")
    print(f"[OK] Owner list: Brett, John, Chris, Sunil (in column M)")
    print(f"\nNext steps:")
    print(f"1. Open {filename} in Excel")
    print(f"2. Press Alt+F11 to open VBA Editor")
    print(f"3. Copy/paste macros from VPM_New_Tracker_Macros_V2.vba")
    print(f"4. Save as .xlsm (Excel Macro-Enabled Workbook)")
    print(f"\nColumns:")
    print(f"  Visible: A-G (Task Name, Dates, Status, Owner, Notes)")
    print(f"  Hidden: H-M (ID, Level, Parent_ID, Dates_Locked, Reserved, Owner List)")
    print(f"\nFeatures:")
    print(f"  - Subtask symbol: >")
    print(f"  - Owner dropdown with validation")
    print(f"  - Right-click to edit owner list")

if __name__ == "__main__":
    create_vpm_tracker_template()
