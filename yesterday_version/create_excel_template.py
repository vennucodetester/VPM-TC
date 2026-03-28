#!/usr/bin/env python3
"""
Script to create DG-Cassette2.0 Excel template
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_template():
    """Create the DG-Cassette2.0.xlsm template"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    settings_sheet = wb.create_sheet("Settings", 0)
    tasks_sheet = wb.create_sheet("Tasks", 1)
    instructions_sheet = wb.create_sheet("Instructions", 2)

    # ==================== SETTINGS SHEET ====================
    create_settings_sheet(settings_sheet)

    # ==================== TASKS SHEET ====================
    create_tasks_sheet(tasks_sheet)

    # ==================== INSTRUCTIONS SHEET ====================
    create_instructions_sheet(instructions_sheet)

    # Save file
    filename = "VPM_Master_Template.xlsx"
    wb.save(filename)
    print(f"[OK] Created Excel template: {filename}")
    print(f"[OK] Template has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    print("\nNOTE: To enable macros, you need to:")
    print("1. Open the file in Excel")
    print("2. Save As -> Excel Macro-Enabled Workbook (.xlsm)")
    print("3. Add VBA code from the implementation plan")


def create_settings_sheet(sheet):
    """Create the Settings configuration sheet"""

    # Title
    sheet['A1'] = 'VPM STATUS CONFIGURATION'
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    sheet.merge_cells('A1:D1')

    # Subtitle
    sheet['A2'] = 'Edit this table to customize your statuses'
    sheet['A2'].font = Font(size=10, italic=True)
    sheet.merge_cells('A2:D2')

    # Headers
    headers = ['Status Name', 'Background Color', 'Text Color', 'Progress %']
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Default statuses
    default_statuses = [
        ['Not Started', 'D3D3D3', '000000', 0],
        ['In Progress', 'FFFF00', '000000', 50],
        ['Completed', '00B050', 'FFFFFF', 100],
        ['Blocked', 'FF0000', 'FFFFFF', 0]
    ]

    for row, status_data in enumerate(default_statuses, start=5):
        sheet.cell(row=row, column=1).value = status_data[0]
        sheet.cell(row=row, column=2).value = status_data[1]
        sheet.cell(row=row, column=3).value = status_data[2]
        sheet.cell(row=row, column=4).value = status_data[3]

    # Add borders to table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(4, 9):
        for col in range(1, 5):
            sheet.cell(row=row, column=col).border = thin_border

    # Timeline alert settings
    sheet['A12'] = 'TIMELINE ALERT SETTINGS'
    sheet['A12'].font = Font(size=14, bold=True, color='FFFFFF')
    sheet['A12'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    sheet.merge_cells('A12:D12')

    sheet['A14'] = 'Days before due date to show warning:'
    sheet['C14'] = 7
    sheet['C14'].alignment = Alignment(horizontal='center')

    sheet['A16'] = 'Color coding:'
    sheet['A17'] = '  • Green: More than 7 days remaining'
    sheet['A18'] = '  • Yellow: 1-7 days remaining (warning)'
    sheet['A19'] = '  • Red: Overdue (past end date)'

    # Instructions
    sheet['A22'] = 'HOW TO CUSTOMIZE:'
    sheet['A22'].font = Font(size=12, bold=True)

    instructions = [
        '1. Add new row to the table above to add a new status',
        '2. Delete row to remove a status',
        '3. Change Status Name to rename',
        '4. Change colors (use hex codes like FF0000 for red)',
        '5. Change Progress % (0-100)',
        '',
        'COLOR OPTIONS:',
        '  • Use hex codes: FF0000 (red), 00FF00 (green), 0000FF (blue)',
        '  • Common colors: D3D3D3 (gray), FFFF00 (yellow), FFA500 (orange)',
        '',
        'After making changes, save the file and re-import to VPM'
    ]

    for i, instruction in enumerate(instructions, start=23):
        sheet[f'A{i}'] = instruction

    # Set column widths
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15


def create_tasks_sheet(sheet):
    """Create the Tasks data entry sheet"""

    # Title/Toolbar area
    sheet['A1'] = 'QUICK ACTIONS:'
    sheet['A1'].font = Font(size=14, bold=True)

    sheet['A2'] = '[Instructions: Use buttons or add rows manually below. See Instructions sheet for details]'
    sheet['A2'].font = Font(size=9, italic=True, color='666666')
    sheet.merge_cells('A2:F2')

    # Column headers (row 5)
    headers = ['Task Name', 'Start Date', 'End Date', 'Status', 'Responsible', 'Comments']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Hidden columns headers (for VPM integration)
    hidden_headers = ['ID', 'Level', 'Parent_ID', 'Progress', 'Current_Status']
    for col, header in enumerate(hidden_headers, start=7):
        cell = sheet.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=9, color='999999')
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

    # Sample data (one example task)
    sheet['A6'] = 'Example Task'
    sheet['B6'] = '11/16/2025'
    sheet['C6'] = '12/31/2025'
    sheet['D6'] = 'Not Started'
    sheet['E6'] = 'Team'
    sheet['F6'] = '11/16: This is a sample task'
    sheet['G6'] = 1  # ID
    sheet['H6'] = 1  # Level
    sheet['I6'] = ''  # Parent_ID
    sheet['J6'] = 0  # Progress
    sheet['K6'] = 'Not Started'  # Current_Status

    # Add child example
    sheet['A7'] = '  └─ Child Task Example'
    sheet['B7'] = '11/16/2025'
    sheet['C7'] = '12/15/2025'
    sheet['D7'] = 'Not Started'
    sheet['E7'] = 'John'
    sheet['F7'] = '11/16: This is a child task'
    sheet['G7'] = 2  # ID
    sheet['H7'] = 2  # Level
    sheet['I7'] = 1  # Parent_ID
    sheet['J7'] = 0  # Progress
    sheet['K7'] = 'Not Started'  # Current_Status

    # Set column widths
    sheet.column_dimensions['A'].width = 40  # Task Name
    sheet.column_dimensions['B'].width = 12  # Start Date
    sheet.column_dimensions['C'].width = 12  # End Date
    sheet.column_dimensions['D'].width = 15  # Status
    sheet.column_dimensions['E'].width = 15  # Responsible
    sheet.column_dimensions['F'].width = 50  # Comments

    # Hidden columns (narrower)
    for col in range(7, 12):
        sheet.column_dimensions[get_column_letter(col)].width = 10
        sheet.column_dimensions[get_column_letter(col)].hidden = True

    # Add data validation for Status column (reads from Settings sheet)
    status_validation = DataValidation(
        type="list",
        formula1="Settings!$A$5:$A$8",
        allow_blank=False
    )
    status_validation.error = 'Please select a status from the list'
    status_validation.errorTitle = 'Invalid Status'
    sheet.add_data_validation(status_validation)
    status_validation.add(f'D6:D1000')  # Apply to Status column

    # Format dates
    for row in range(6, 1000):
        sheet[f'B{row}'].number_format = 'MM/DD/YYYY'
        sheet[f'C{row}'].number_format = 'MM/DD/YYYY'

    # Text wrapping for Comments column
    for row in range(6, 1000):
        sheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze panes (freeze header row)
    sheet.freeze_panes = 'A6'


def create_instructions_sheet(sheet):
    """Create the Instructions/Help sheet"""

    # Title
    sheet['A1'] = 'DG-Cassette2.0 - User Guide'
    sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    sheet.merge_cells('A1:E1')

    instructions = [
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'ONE-TIME SETUP',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        '1. Go to Tasks sheet',
        '2. Delete the example tasks (rows 6-7)',
        '3. Add your tasks:',
        '   • Type task name in column A',
        '   • Enter start and end dates (MM/DD/YYYY format)',
        '   • Select status from dropdown',
        '   • Enter responsible person name',
        '   • Leave comments empty for now',
        '',
        '4. For child tasks:',
        '   • Add 2-4 spaces before the task name (visual indent)',
        '   • Fill hidden columns:',
        '     - Level: 2 for child, 3 for grandchild, etc.',
        '     - Parent_ID: Enter the ID number of the parent task',
        '',
        '5. Save file as .xlsm (Excel Macro-Enabled Workbook)',
        '',
        '6. Import to VPM:',
        '   • Open VPM application',
        '   • File → Import from Excel Template',
        '   • Select this file',
        '   • All tasks will be created with hierarchy',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'WEEKLY UPDATES',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        '1. Open this Excel file',
        '',
        '2. Update tasks:',
        '   • Click in Comments cell',
        '   • Press Alt+Enter to add new line (or just continue typing)',
        '   • Type date (MM/DD) and your update',
        '   • Example: "11/20: Completed testing phase"',
        '',
        '3. Update Status if changed:',
        '   • Click Status dropdown',
        '   • Select new status (In Progress, Completed, etc.)',
        '',
        '4. Save file',
        '',
        '5. Import to VPM:',
        '   • VPM will update existing tasks',
        '   • Comments will appear in Notes grid view',
        '   • Progress will update based on status',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'ADDING COMMENTS (DETAILED)',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        'Current cell content:',
        '  "11/16: Started testing"',
        '',
        'To add a new update:',
        '  1. Click in the Comments cell',
        '  2. Press END key (go to end of text)',
        '  3. Press Alt+Enter (creates new line in same cell)',
        '  4. Type: "11/20: Parts arrived"',
        '',
        'Result:',
        '  "11/16: Started testing',
        '   11/20: Parts arrived"',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'ADDING NEW TASKS MID-PROJECT',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        '1. Find the last row with data',
        '2. Add new row below',
        '3. Fill in task details',
        '4. Assign next ID number (look at last ID and add 1)',
        '5. Set Level and Parent_ID if it\'s a child task',
        '6. Save and import to VPM',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'HIERARCHY SETUP',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        'Example structure:',
        '',
        'Task Name              | Level | Parent_ID',
        'Testing                |   1   |',
        '  Performance/UL       |   2   |   1',
        '    5DR Testing        |   3   |   2',
        '    4DR Testing        |   3   |   2',
        '  Structural testing   |   2   |   1',
        'ECN release            |   1   |',
        '',
        'Rules:',
        '  • Level 1 = Top-level task (no parent)',
        '  • Level 2 = Child of Level 1',
        '  • Level 3 = Child of Level 2',
        '  • Parent_ID = ID of the parent task (leave blank for Level 1)',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'GRID NOTES IN VPM',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        'After importing, comments appear in VPM as a grid:',
        '',
        '  Date  | Update                    | Responsible | Status',
        '  11/16 | Started testing           | John        |',
        '  11/20 | Parts arrived             | John        |',
        '  11/25 | Testing in progress       | John        |',
        '',
        'The grid also shows:',
        '  • Timeline icon with end date',
        '  • Color coding based on deadline:',
        '    - Green: More than 7 days remaining',
        '    - Yellow: 1-7 days until deadline',
        '    - Red: Past deadline (overdue)',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'CUSTOMIZING STATUSES',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        '1. Go to Settings sheet',
        '2. Edit the Status Configuration table:',
        '   • Add new row for new status',
        '   • Delete row to remove status',
        '   • Change colors (use hex codes)',
        '   • Change progress % mapping',
        '3. Save file',
        '4. Changes take effect on next import',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'TIPS & TRICKS',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        '• Keep ID numbers sequential (1, 2, 3, 4...)',
        '• Don\'t skip ID numbers',
        '• Use consistent date format (MM/DD/YYYY)',
        '• Add dates to comments for tracking history',
        '• Update Responsible person if ownership changes',
        '• Save regularly to avoid losing work',
        '• Share file in Teams for collaborative updates',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
        'TROUBLESHOOTING',
        '═══════════════════════════════════════════════════════════════════════════',
        '',
        'Problem: Tasks not importing correctly',
        'Solution: Check that ID, Level, and Parent_ID columns are filled',
        '',
        'Problem: Hierarchy is wrong in VPM',
        'Solution: Verify Parent_ID matches the parent task\'s ID',
        '',
        'Problem: Dates showing as numbers',
        'Solution: Format cells as Date (MM/DD/YYYY)',
        '',
        'Problem: Status dropdown not working',
        'Solution: Make sure Settings sheet has status definitions',
        '',
        '═══════════════════════════════════════════════════════════════════════════',
    ]

    for i, line in enumerate(instructions, start=2):
        sheet[f'A{i}'] = line
        if '═══' in line or line.startswith('ONE-TIME') or line.startswith('WEEKLY') or \
           line.startswith('ADDING') or line.startswith('HIERARCHY') or \
           line.startswith('GRID NOTES') or line.startswith('CUSTOMIZING') or \
           line.startswith('TIPS') or line.startswith('TROUBLESHOOTING'):
            sheet[f'A{i}'].font = Font(bold=True, size=11)

    # Set column width
    sheet.column_dimensions['A'].width = 80


if __name__ == '__main__':
    create_template()
